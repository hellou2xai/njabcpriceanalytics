"""Compare Prices — side-by-side price comparison across 2-3 distributors.

Strict UPC intersection: only products carried by ALL selected distributors
(same normalised UPC + size + pack + vintage rules as /catalog/cross-distributor)
so the grid never shows blank cells.

Three price layers per distributor per product:
  - frontline_case_price  (list)
  - best_case_price       (price after the best CPL quantity discount)
  - effective_case_price  (after best QD + best full-month RIP — the canonical
                           "the price" per FOUNDATION.md)
plus the full QD/RIP tier ladder on demand (/tiers) via pricing.attach_tiers —
no pricing math re-implemented here.

Endpoints:
  GET /api/compare/options   — distributors + product counts for the picker
  GET /api/compare/products  — the comparison grid + smart analysis summary
  GET /api/compare/tiers     — full side-by-side tier ladders for one product
  GET /api/compare/rips      — RIP-outcome comparison (landed curve + break-even)
  GET /api/compare/price360  — one holistic per-product label ranking every offer
"""
import io
import math
import re
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from backend import pricing as _pricing
from backend.auth import get_optional_user, get_current_user
from backend.db import get_duckdb, read_parquet
from backend.enrichment_join import attach_sku_mapping as _attach_sku_mapping
from backend.enrichment_join import attach_enrichment_image as _attach_image
from backend.size_std import _to_ml

router = APIRouter(prefix="/api/compare", tags=["compare"])

_MAX_WHOLESALERS = 3
_TIE_EPS = 0.005

# A deal counts toward today's price only when its window is live NOW. Past
# (expired) and future (upcoming) deals are excluded so the grid reflects what a
# buyer can actually get today and old promos don't confuse the comparison.
_ACTIVE_NOW = {"active", "whole_month", "evergreen"}


def _has_rip_signal(rec: dict) -> bool:
    """Cheap pre-tier candidate test: does this CPL row plausibly have a RIP?
    True when the precomputed has_rip flag is set OR a real rip_code is present.
    The rip_code check catches RIPs the precomputed flag can lag on — e.g. Opici
    files a descriptive text code and a multi-listing UPC's has_rip can read
    False in a stale cache. This only widens the candidate set fed to
    attach_tiers; the precise gate (the distributor actually has RIP tiers) is
    applied after the canonical ladder is built."""
    if rec.get("has_rip"):
        return True
    code = str(rec.get("rip_code") or "").strip()
    return bool(code) and code.lower() not in ("0", "none", "nan")

# Same validity rule as catalog: a real barcode, not stub/placeholder filler.
_VALID_UPC = (
    "upc IS NOT NULL AND upc <> '' AND upc <> '0'"
    " AND NOT regexp_matches(upc, '^(0+|9+|1+)$')"
    " AND NOT upc LIKE '999999%'"
    " AND LENGTH(LTRIM(upc, '0')) >= 8"
)


def _parse_wholesalers(raw: str, con) -> list[str]:
    """Validate the comma-separated wholesaler list against the data."""
    slugs = [w.strip() for w in raw.split(",") if w.strip()]
    if not (2 <= len(slugs) <= _MAX_WHOLESALERS):
        raise HTTPException(400, f"Pick 2-{_MAX_WHOLESALERS} distributors")
    if len(set(slugs)) != len(slugs):
        raise HTTPException(400, "Duplicate distributor")
    src = read_parquet(con, "cpl_enriched")
    known = {r[0] for r in con.execute(f"SELECT DISTINCT wholesaler FROM {src}").fetchall()}
    bad = [s for s in slugs if s not in known]
    if bad:
        raise HTTPException(400, f"Unknown distributor(s): {', '.join(bad)}")
    return slugs


def _editions_for(con, src: str, slugs: list[str]) -> dict[str, str]:
    """Current edition per wholesaler (latest edition <= current ET month,
    falling back to the latest available)."""
    current_ym = _pricing.current_yyyy_mm()
    placeholders = ",".join("?" * len(slugs))
    rows = con.execute(
        f"""
        SELECT wholesaler,
               MAX(CASE WHEN edition <= ? THEN edition END) AS cur_ed,
               MAX(edition) AS latest_ed
        FROM {src}
        WHERE wholesaler IN ({placeholders})
        GROUP BY wholesaler
        """,
        [current_ym] + slugs,
    ).fetchall()
    eds = {r[0]: (r[1] or r[2]) for r in rows}
    missing = [s for s in slugs if s not in eds]
    if missing:
        raise HTTPException(400, f"No data for: {', '.join(missing)}")
    return eds


def _edition_pred(slugs: list[str], eds: dict[str, str]) -> tuple[str, list]:
    parts, params = [], []
    for s in slugs:
        parts.append("(wholesaler = ? AND edition = ?)")
        params.extend([s, eds[s]])
    return "(" + " OR ".join(parts) + ")", params


def _nan_clean(rec: dict) -> dict:
    for k, v in list(rec.items()):
        if isinstance(v, float) and math.isnan(v):
            rec[k] = None
    return rec


_PACK_COMBO = re.compile(r"^\s*(\d+)\s*[/xX\s]\s*(\d+)")


def _pack_norm(uq) -> str:
    """Bottles-per-case as a canonical string. Handles '24', '24.0', 24.0 and
    multi-pack spellings like '2 12-Packs' / '4 6-packs' (-> 24). '' if unknown."""
    if uq is None:
        return ""
    s = str(uq).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f != f:  # NaN
            return ""
        return str(int(f)) if f.is_integer() else str(f)
    except ValueError:
        pass
    m = _PACK_COMBO.match(s)
    if m:
        return str(int(m.group(1)) * int(m.group(2)))
    return s.upper()


def _vintage_key(vintage_norm, vintage_sensitive) -> str:
    """The vintage component of a match_key, always a STRING. Shared by
    _common_rows and _prev_prices so the two key the same way. In postgres-cache
    mode vintage_norm can arrive as a float (e.g. 2022.0); coerce it to '2022'
    so the key matches the parquet/str form (and "|".join never sees a float)."""
    if not vintage_sensitive:
        return ""
    v = vintage_norm
    if v is None:
        return ""
    if isinstance(v, float):
        if v != v:  # NaN
            return ""
        return str(int(v)) if v.is_integer() else str(v)
    return str(v)


def _size_key(raw) -> str:
    """Physical-size bucket so '12OZ', '12oz' and '355ML' all match (same can),
    and '15.5GAL' matches '1984OZ' (same 1/2-BBL keg). Unparseable sizes fall
    back to the cleaned raw string so they can still match their own spelling."""
    ml, _fam = _to_ml(str(raw or ""))
    if ml is None:
        return (str(raw or "").strip().upper().replace(" ", "")) or "?"
    if ml >= 15000:  # kegs / bulk: bucket to 100 ml
        return f"K{round(ml / 100)}"
    return f"M{round(ml / 5) * 5}"  # bottles/cans: bucket to 5 ml


def _common_rows(con, src: str, slugs: list[str], eds: dict[str, str],
                 require_all: bool = True) -> list[dict]:
    """One best-offer row per (identity key, wholesaler), restricted to
    identity keys present at ALL selected wholesalers.

    Pass ``require_all=False`` to keep the best offer at EVERY wholesaler that
    carries each key (no intersection) — used by the Best RIPs board, which
    shows a RIP even when only one distributor carries the product.

    Identity = normalised UPC + physical-size bucket + bottles-per-case +
    vintage (vintage-sensitive categories only). Size and pack are normalised
    in Python (_size_key / _pack_norm) because distributors spell the same
    physical size many ways ('12OZ' vs '355ML', '24' vs '2 12-Packs')."""
    ed_pred, ed_params = _edition_pred(slugs, eds)
    vn = _pricing.vintage_norm_sql("vintage")
    # Exclude a row as "part of a combo bundle" ONLY when its combo_code is a
    # real code in that wholesaler's COMBO sheet. Some distributors (Shore
    # Point, Jersey Beverage) repurpose the CPL combo-code column for internal
    # cross-reference codes — a blanket non-empty test would silently drop
    # most of their catalogue from every comparison.
    combo_pred = "TRUE"
    has_combo_tbl = bool(con.execute(
        "SELECT 1 FROM information_schema.tables WHERE table_name = 'combo'"
    ).fetchone())
    if has_combo_tbl:
        combo_src = read_parquet(con, "combo")
        combo_pred = f"""NOT EXISTS (
            SELECT 1 FROM {combo_src} cb
            WHERE cb.wholesaler = e.wholesaler AND cb.edition = e.edition
              AND cb.combo_code = e.combo_code
        )"""
    sql = f"""
        SELECT wholesaler, edition, upc, product_name, product_type, brand,
               unit_qty, unit_volume, unit_type, vintage, abv_proof,
               from_date, to_date,
               frontline_case_price, frontline_unit_price,
               best_case_price, best_unit_price,
               effective_case_price, rip_savings, total_savings_per_case,
               has_discount, has_rip, rip_code, rip_windows,
               discount_1_qty, discount_1_amt, discount_2_qty, discount_2_amt,
               discount_3_qty, discount_3_amt, discount_4_qty, discount_4_amt,
               discount_5_qty, discount_5_amt,
               LTRIM(upc, '0') AS upc_norm,
               TRY_CAST(unit_qty AS DOUBLE) AS uqd,
               {vn} AS vintage_norm,
               UPPER(product_type) IN ('WINE','SPARKLING','VERMOUTH') AS vintage_sensitive
        FROM {src} e
        WHERE {ed_pred}
          AND {_VALID_UPC}
          AND (combo_code IS NULL OR combo_code = '' OR combo_code = '0'
               OR {combo_pred})
    """
    df = con.execute(sql, ed_params).df()
    recs = [_nan_clean(r) for r in df.to_dict(orient="records")]

    # Build identity keys in Python (size/pack spellings vary by distributor).
    for r in recs:
        pack = _pack_norm(r.get("unit_qty"))
        r["pack_norm"] = pack
        r["match_key"] = "|".join([
            r["upc_norm"],
            _size_key(r.get("unit_volume")),
            pack,
            _vintage_key(r.get("vintage_norm"), r.get("vintage_sensitive")),
        ])

    # Drop ambiguous identities: a key mapping to >1 distinct product name
    # within ONE wholesaler is an unreliable barcode (same rule as
    # /catalog/cross-distributor). Split-case duplicate listings of the same
    # product share a name prefix, so compare on a cleaned name.
    names: dict[tuple, set] = {}
    for r in recs:
        names.setdefault((r["wholesaler"], r["match_key"]), set()).add(
            re.sub(r"\s+", " ", (r["product_name"] or "").strip().upper())[:12]
        )
    ambiguous = {k for k, v in names.items() if len(v) > 1}
    recs = [r for r in recs if (r["wholesaler"], r["match_key"]) not in ambiguous]

    # Best offer per (key, wholesaler). PREFER a row whose window is live TODAY
    # (so the tier ladder we build from it reflects today's deals, not a future
    # promo window), then cheapest effective, then frontline. Distributors that
    # split a SKU into consecutive dated windows (Shore Point) would otherwise
    # leave the active window's deal mis-stamped as "upcoming".
    for r in recs:
        st = _pricing.window_status(r.get("from_date"), r.get("to_date"))["status"]
        r["_active_today"] = st in _ACTIVE_NOW
    best: dict[tuple, dict] = {}
    for r in recs:
        k = (r["match_key"], r["wholesaler"])
        cur = best.get(k)
        def _rank(x):
            return (
                0 if x.get("_active_today") else 1,
                x.get("effective_case_price") if x.get("effective_case_price") is not None else float("inf"),
                x.get("frontline_case_price") if x.get("frontline_case_price") is not None else float("inf"),
            )
        if cur is None or _rank(r) < _rank(cur):
            best[k] = r

    if not require_all:
        return list(best.values())

    # Keep only keys present at ALL selected wholesalers.
    per_key: dict[str, int] = {}
    for (mk, _w) in best:
        per_key[mk] = per_key.get(mk, 0) + 1
    n = len(slugs)
    return [r for (mk, _w), r in best.items() if per_key[mk] == n]


def _price_obj(d: dict) -> dict:
    """One distributor's price layers (computed LIVE for today) plus the
    explanation metadata the UI needs:

      - qd_save            : $/case the best QD takes off the list price
      - qd_time_sensitive  : the deal driving Best QD is a dated promo that ENDS
                             this month (window_status == 'active'), so the buyer
                             should know the price won't last
      - deal_window        : {from, to, status} of that dated deal

    Prices are already set on `d` by _live_best (active-today tiers only), so
    Best Net is never higher than Best QD here."""
    front = d.get("frontline_case_price")
    bq = d.get("best_case_price")
    net = d.get("effective_case_price")
    at = d.get("_applied_qd_tier") or {}
    qd_save = (round(front - bq, 2)
               if front is not None and bq is not None and bq < front - _TIE_EPS else 0.0)
    # time-sensitive = the applied QD rides on a dated window that ends this month
    ts = bool(qd_save > 0 and at.get("window_status") == "active")
    win = ({"from": at.get("from_date"), "to": at.get("to_date"),
            "status": at.get("window_status")} if ts else None)
    return {
        "upc": d.get("upc"),
        "edition": d.get("edition"),
        "product_name": d.get("product_name"),
        "frontline": front,
        "after_qd": bq,
        "effective": net,
        "btl_effective": (round(net / d["uqd"], 2)
                          if net and d.get("uqd") else None),
        "rip_savings": d.get("rip_savings"),
        "qd_save": qd_save or None,
        "qd_time_sensitive": ts,
        "deal_window": win,
        "has_discount": bool(d.get("has_discount")),
        "has_rip": bool(d.get("has_rip")),
        # Prior-edition layers for the two-month view (None unless months=2).
        "prev": ({
            "edition": pv.get("edition"),
            "frontline": pv.get("frontline"),
            "after_qd": pv.get("after_qd"),
            "effective": pv.get("effective"),
            "btl_effective": (round(pv["effective"] / d["uqd"], 2)
                              if pv.get("effective") and d.get("uqd") else None),
        } if (pv := d.get("_prev")) else None),
    }


def _winner(per: dict[str, dict], field: str) -> dict:
    """Cheapest wholesaler for a price field. Ties within $0.005 -> 'tie'."""
    vals = {w: d.get(field) for w, d in per.items() if d.get(field) is not None and d.get(field) > 0}
    if not vals:
        return {"winner": None, "spread": None}
    lo = min(vals.values())
    hi = max(vals.values())
    winners = [w for w, v in vals.items() if abs(v - lo) < _TIE_EPS]
    return {
        "winner": "tie" if len(winners) > 1 else winners[0],
        "spread": round(hi - lo, 2),
    }


@router.get("/options")
def options(user: Optional[dict] = Depends(get_optional_user)):
    """Distributors available for comparison, with current-edition product
    counts so the picker can hint at catalogue size."""
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        slugs = [r[0] for r in con.execute(
            f"SELECT DISTINCT wholesaler FROM {src} ORDER BY wholesaler").fetchall()]
        eds = _editions_for(con, src, slugs)
        ed_pred, ed_params = _edition_pred(slugs, eds)
        rows = con.execute(f"""
            SELECT wholesaler, COUNT(DISTINCT LTRIM(upc,'0') || '|' || COALESCE(unit_volume,'')) AS n
            FROM {src}
            WHERE {ed_pred} AND {_VALID_UPC}
            GROUP BY wholesaler ORDER BY wholesaler
        """, ed_params).fetchall()
        return [
            {"wholesaler": w, "edition": eds.get(w), "products": n}
            for w, n in rows
        ]


def _active_qd_from_raw(con, slugs: list[str], eds: dict[str, str],
                        n_cases: Optional[float], upcs: list[str]) -> dict[tuple, tuple]:
    """Deepest quantity discount LIVE TODAY for each SKU, read straight from the
    RAW cpl (all window rows), keyed by (wholesaler, upc_norm, size_key, pack).

    cpl_enriched keeps only ONE row per SKU, so when a distributor splits a SKU
    into consecutive dated windows (e.g. Shore Point: Jun 1-21 then Jun 22-30) the
    enriched row — and any ladder built from it — can miss the window that's
    actually live now. Reading raw cpl avoids that: we look at every row whose
    window contains today (or is full-month/evergreen) and take the deepest
    discount, volume-capped when `n_cases` is set.

    Scoped to the page's matched UPCs and filtered to live windows IN SQL, so a
    big head-to-head (allied vs fedway) doesn't scan the whole catalogue in
    Python."""
    if not upcs:
        return {}
    ed_pred, ed_params = _edition_pred(slugs, eds)
    today = _pricing.eastern_today().isoformat()
    upc_ph = ",".join("?" * len(upcs))
    # a window is live today when it contains today, or has no dates (evergreen)
    active_sql = ("(from_date IS NULL OR to_date IS NULL OR "
                  "(CAST(from_date AS DATE) <= CAST(? AS DATE) "
                  "AND CAST(? AS DATE) <= CAST(to_date AS DATE)))")
    df = con.execute(f"""
        SELECT wholesaler, upc, unit_volume, unit_qty, from_date, to_date,
               discount_1_qty, discount_1_amt, discount_2_qty, discount_2_amt,
               discount_3_qty, discount_3_amt, discount_4_qty, discount_4_amt,
               discount_5_qty, discount_5_amt
        FROM cpl
        WHERE {ed_pred} AND {_VALID_UPC}
          AND LTRIM(upc, '0') IN ({upc_ph})
          AND {active_sql}
    """, ed_params + list(upcs) + [today, today]).df()
    out: dict[tuple, tuple] = {}
    for r in df.to_dict("records"):
        # SQL already restricted to live windows; classify only for the marker
        st = "active" if _pricing.is_time_sensitive_window(
            r.get("from_date"), r.get("to_date")) else "whole_month"
        best_amt = 0.0
        for i in range(1, 6):
            a = r.get(f"discount_{i}_amt")
            try:
                af = float(a)
            except (TypeError, ValueError):
                continue
            if af != af or af <= 0:
                continue
            if n_cases is not None:
                m = re.match(r"^\s*(\d+(?:\.\d+)?)", str(r.get(f"discount_{i}_qty") or ""))
                thr = float(m.group(1)) if m else None
                if thr is None or n_cases + 1e-9 < thr:
                    continue
            if af > best_amt:
                best_amt = af
        if best_amt <= 0:
            continue
        key = (r["wholesaler"], str(r.get("upc") or "").lstrip("0"),
               _size_key(r.get("unit_volume")), _pack_norm(r.get("unit_qty")))
        cur = out.get(key)
        if cur is None or best_amt > cur[0]:
            out[key] = (best_amt, _pricing._iso(r.get("from_date")),
                        _pricing._iso(r.get("to_date")), st)
    return out


def _active_rip_rebate(tiers: list, n_cases: Optional[float], pack: float) -> float:
    """Best RIP-only rebate $/case from RIP tiers live today (volume-capped when
    n_cases is set). RIP windows come from the rip sheet rows, which attach_tiers
    classifies reliably, so the ladder is trustworthy for the RIP layer."""
    best = 0.0
    for t in tiers:
        if t.get("source") != "rip" or t.get("window_status") not in _ACTIVE_NOW:
            continue
        if n_cases is not None:
            thr = _cases_threshold(t, pack)
            if thr is None or n_cases + 1e-9 < thr:
                continue
        save = t.get("rip_only_save_per_case")
        if save is None:
            save = t.get("save_per_case") or 0.0
        if save and save > best:
            best = save
    return best


def _prev_prices(con, src: str, slugs: list[str], eds: dict[str, str],
                 page_upcs: list[str]) -> tuple[dict[str, str], dict[tuple, dict]]:
    """Prior-edition price layers for the Price-Comparison 2-month view.

    Returns (prev_editions, lookup) where prev_editions maps wholesaler -> the
    edition immediately BEFORE its current one, and lookup maps
    (wholesaler, match_key) -> {edition, frontline, after_qd, effective} read
    straight from that edition's PRECOMPUTED cpl_enriched columns (List =
    frontline_case_price, Best QD = best_case_price, Best Net =
    effective_case_price — the canonical whole-month price for a past month you
    can no longer buy live). match_key is built exactly like _common_rows so it
    joins the current-month rows 1:1."""
    prev_eds: dict[str, str] = {}
    for w in slugs:
        r = con.execute(
            f"SELECT MAX(edition) FROM {src} WHERE wholesaler = ? AND edition < ?",
            [w, eds[w]]).fetchone()
        if r and r[0]:
            prev_eds[w] = r[0]
    if not prev_eds or not page_upcs:
        return prev_eds, {}
    parts, params = [], []
    for w, e in prev_eds.items():
        parts.append("(wholesaler = ? AND edition = ?)")
        params += [w, e]
    upc_ph = ",".join("?" * len(page_upcs))
    vn = _pricing.vintage_norm_sql("vintage")
    df = con.execute(f"""
        SELECT wholesaler, edition, LTRIM(upc, '0') AS upc_norm,
               unit_volume, unit_qty,
               frontline_case_price, best_case_price, effective_case_price,
               {vn} AS vintage_norm,
               UPPER(product_type) IN ('WINE','SPARKLING','VERMOUTH') AS vintage_sensitive
        FROM {src}
        WHERE ({" OR ".join(parts)}) AND {_VALID_UPC}
          AND LTRIM(upc, '0') IN ({upc_ph})
    """, params + list(page_upcs)).df()

    def _nz(v):
        # NaN -> None, and coerce numpy/Decimal scalars to native float so the
        # JSON response never trips on a non-serialisable type.
        if v is None or (isinstance(v, float) and v != v):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v

    lookup: dict[tuple, dict] = {}
    for r in df.to_dict("records"):
        pack = _pack_norm(r.get("unit_qty"))
        mk = "|".join([
            r["upc_norm"], _size_key(r.get("unit_volume")), pack,
            _vintage_key(r.get("vintage_norm"), r.get("vintage_sensitive")),
        ])
        k = (r["wholesaler"], mk)
        eff = _nz(r.get("effective_case_price"))
        cur = lookup.get(k)
        # cheapest effective wins (same rule as the current-month best offer)
        if cur is None or (eff is not None and (cur["effective"] is None or eff < cur["effective"])):
            lookup[k] = {
                "edition": r["edition"],
                "frontline": _nz(r.get("frontline_case_price")),
                "after_qd": _nz(r.get("best_case_price")),
                "effective": eff,
            }
    return prev_eds, lookup


@router.get("/products")
def compare_products(
    wholesalers: str = Query(..., description="2-3 comma-separated slugs"),
    q: str = Query("", description="Product name contains"),
    product_type: str = Query(""),
    only_differences: bool = Query(False),
    min_spread: float = Query(0.0, ge=0),
    cases: float = Query(0, ge=0, description="Buy quantity; 0 = each distributor's best deal"),
    sort: str = Query("spread", description="spread | spread_pct | product | effective"),
    order: str = Query("desc"),
    limit: int = Query(2000, ge=1, le=50000),
    months: int = Query(1, ge=1, le=2, description="1 = current month only; 2 = also attach each distributor's PRIOR-edition price layers (prev) for the two-month Price Comparison view."),
    user: Optional[dict] = Depends(get_optional_user),
):
    """The comparison grid: products common to ALL selected distributors with
    the three price layers per distributor, per-layer winners, and a smart
    analysis summary.

    `cases` controls whether the After-QD / After-RIP columns are each
    distributor's BEST deal (cases=0, deepest tier) or the landed price AT THAT
    VOLUME — so a low-volume buyer isn't shown a winner they can't actually reach.
    At-volume uses the canonical tier ladder (attach_tiers) + `_applied_tier_at`,
    never a re-implemented formula."""
    n_cases = cases if cases and cases > 0 else None
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        slugs = _parse_wholesalers(wholesalers, con)
        eds = _editions_for(con, src, slugs)
        raw = _common_rows(con, src, slugs, eds)
        # Set each distributor's Best QD / Best Net from the deals that are LIVE
        # TODAY (expired + upcoming excluded), so the grid shows today's real
        # price and old/future promos don't confuse the winner. The QD comes from
        # RAW cpl (every window row), which is robust to a SKU split into
        # consecutive dated windows; the RIP layer comes from the canonical tier
        # ladder. In at-volume mode a deal must also be reachable at `cases`.
        page_upcs = sorted({r["upc_norm"] for r in raw if r.get("upc_norm")})
        qd_live = _active_qd_from_raw(con, slugs, eds, n_cases, page_upcs)
        # RIP layer: at-volume needs the volume-aware tier ladder; best-deal uses
        # the cheap precomputed rip_windows overlay (no per-product RIP queries,
        # so a big allied-vs-fedway grid stays fast).
        if n_cases:
            _pricing.attach_tiers(con, raw)
        else:
            _pricing.attach_live_rip(con, raw)
        for d in raw:
            pack = float(d.get("uqd") or 0)
            tiers = d.get("tiers") or []
            front = d.get("frontline_case_price")
            parts = d["match_key"].split("|")
            key = (d["wholesaler"], d["upc_norm"],
                   parts[1] if len(parts) > 1 else "", parts[2] if len(parts) > 2 else "")
            qd_hit = qd_live.get(key)
            qd_amt = qd_hit[0] if qd_hit else 0.0
            best_qd = round(front - qd_amt, 2) if front is not None and qd_amt > 0 else front
            rip_amt = (_active_rip_rebate(tiers, n_cases, pack) if n_cases
                       else (d.get("live_rip_amt") or 0.0))
            best_net = best_qd
            if best_qd is not None and rip_amt > 0:
                best_net = max(round(best_qd - rip_amt, 2), 0.0)
            d["best_case_price"] = best_qd
            d["effective_case_price"] = best_net
            d["rip_savings"] = (round(best_qd - best_net, 2)
                                if best_qd is not None and best_net is not None
                                and best_qd - best_net > _TIE_EPS else None)
            # window of the live QD deal, for the "ends this month" marker
            d["_applied_qd_tier"] = (
                {"window_status": qd_hit[3], "from_date": qd_hit[1], "to_date": qd_hit[2]}
                if qd_hit and qd_amt > 0 else None)

        # Two-month Price Comparison: attach each distributor's PRIOR-edition
        # price layers (prev) from the precomputed cpl_enriched columns.
        prev_eds: dict[str, str] = {}
        if months >= 2:
            try:
                prev_eds, prev_lookup = _prev_prices(con, src, slugs, eds, page_upcs)
                for d in raw:
                    d["_prev"] = prev_lookup.get((d["wholesaler"], d["match_key"]))
            except Exception as e:
                # Never 500 the whole grid over the optional prior-month layer;
                # degrade to the current month and log for follow-up.
                print(f"[compare] prev-month fetch failed, showing current only: {e}")
                prev_eds = {}

    # Pivot: match_key -> {wholesaler: row}
    by_key: dict[str, dict[str, dict]] = {}
    for r in raw:
        by_key.setdefault(r["match_key"], {})[r["wholesaler"]] = r

    rows = []
    for key, per in by_key.items():
        if len(per) != len(slugs):
            continue  # defensive; SQL already guarantees this
        any_row = per[slugs[0]]
        name = min((d["product_name"] for d in per.values()), key=len)
        w_front = _winner(per, "frontline_case_price")
        w_qd = _winner(per, "best_case_price")
        w_eff = _winner(per, "effective_case_price")
        effs = [d.get("effective_case_price") for d in per.values()
                if d.get("effective_case_price")]
        best_eff = min(effs) if effs else None
        per_prices = {w: _price_obj(d) for w, d in per.items()}
        parts = key.split("|")
        row = {
            "match_key": key,
            "upc_norm": parts[0],
            "size_key": parts[1] if len(parts) > 1 else "",
            "product_name": name,
            "product_type": any_row.get("product_type"),
            "brand": any_row.get("brand"),
            "unit_qty": any_row.get("unit_qty"),
            "unit_volume": any_row.get("unit_volume"),
            "unit_type": any_row.get("unit_type"),
            "vintage": any_row.get("vintage"),
            "upc": any_row.get("upc"),
            "prices": per_prices,
            "winner_frontline": w_front["winner"],
            "winner_after_qd": w_qd["winner"],
            "winner_effective": w_eff["winner"],
            "spread": w_eff["spread"],
            "spread_pct": (
                round(w_eff["spread"] / best_eff * 100, 1)
                if w_eff["spread"] is not None and best_eff else None
            ),
            # winner changes once deals are applied -> volume/deals flip it
            "deal_flip": (
                w_front["winner"] is not None and w_eff["winner"] is not None
                and w_front["winner"] != w_eff["winner"]
            ),
            # at least one distributor's shown price rides on a dated deal that
            # ENDS this month — surfaced so the buyer knows it won't last.
            "has_expiring": any(
                p.get("qd_time_sensitive") for p in per_prices.values()
            ),
        }
        rows.append(row)

    # ---- search/category filters narrow BOTH the grid and the summary ------
    if q:
        qq = q.lower()
        # UPC-aware: a barcode never appears in the product NAME, so a digit
        # query must match the (normalised) UPC or it returns nothing.
        qd = re.sub(r"\D", "", q).lstrip("0")
        rows = [r for r in rows if qq in (r["product_name"] or "").lower()
                or qq in (r["brand"] or "").lower()
                or (len(qd) >= 6 and (r.get("upc_norm") == qd
                                      or qd in ((r.get("upc") or "").lstrip("0"))))]
    if product_type:
        rows = [r for r in rows if (r["product_type"] or "").lower() == product_type.lower()]

    # ---- smart analysis (over the full common set for this search context;
    #      only_differences / min_spread are display filters and must NOT
    #      zero out the ties / common-products scoreboard) -------------------
    wins = {w: 0 for w in slugs}
    wins_front = {w: 0 for w in slugs}
    ties = 0
    flips = []
    total_spread = 0.0
    by_type: dict[str, dict[str, int]] = {}
    for r in rows:
        w = r["winner_effective"]
        if w == "tie":
            ties += 1
        elif w in wins:
            wins[w] += 1
            t = r["product_type"] or "Other"
            by_type.setdefault(t, {x: 0 for x in slugs})[w] += 1
        wf = r["winner_frontline"]
        if wf in wins_front:
            wins_front[wf] += 1
        if r["deal_flip"]:
            flips.append(r)
        total_spread += r["spread"] or 0.0

    top_spreads = sorted(rows, key=lambda r: -(r["spread"] or 0))[:5]
    insights = []
    if rows:
        overall = max(wins, key=lambda w: wins[w])
        if wins[overall] > 0:
            insights.append(
                f"{overall} is cheapest on {wins[overall]} of {len(rows)} shared products "
                f"(effective price, after all deals)."
            )
        for t, tw in sorted(by_type.items(), key=lambda kv: -sum(kv[1].values()))[:4]:
            tot = sum(tw.values())
            if tot >= 3:
                lead = max(tw, key=lambda w: tw[w])
                if tw[lead] > tot / 2:
                    insights.append(f"{lead} wins {tw[lead]}/{tot} in {t}.")
        if flips:
            ex = max(flips, key=lambda r: r["spread"] or 0)
            if ex["winner_frontline"] == "tie":
                ex_txt = (f"e.g. {ex['product_name']}: tied at list price, "
                          f"but {ex['winner_effective']} wins after deals.")
            elif ex["winner_effective"] == "tie":
                ex_txt = (f"e.g. {ex['product_name']}: {ex['winner_frontline']} is cheaper "
                          f"at list, but deals level it to a tie.")
            else:
                ex_txt = (f"e.g. {ex['product_name']}: {ex['winner_frontline']} is cheaper "
                          f"at list but {ex['winner_effective']} wins after deals.")
            insights.append(
                f"{len(flips)} product(s) change winner once QD/RIP deals apply: {ex_txt}"
            )
        if total_spread > 0:
            insights.append(
                f"Buying every shared product from its cheapest distributor saves "
                f"${total_spread:,.2f} per case-each versus always picking the most expensive."
            )

    total = len(rows)  # common universe for this search context

    # ---- display filters (grid only — summary above is already computed) ---
    if only_differences:
        rows = [r for r in rows if r["winner_effective"] not in (None, "tie")]
    if min_spread > 0:
        rows = [r for r in rows if (r["spread"] or 0) >= min_spread]

    # ---- sort + limit ------------------------------------------------------
    keymap = {
        "spread": lambda r: r["spread"] or 0,
        "spread_pct": lambda r: r["spread_pct"] or 0,
        "product": lambda r: (r["product_name"] or "").lower(),
        "effective": lambda r: min(
            (p["effective"] for p in r["prices"].values() if p["effective"]),
            default=0,
        ),
    }
    rows.sort(key=keymap.get(sort, keymap["spread"]),
              reverse=(order != "asc") if sort != "product" else (order == "desc"))
    rows = rows[:limit]

    return {
        "wholesalers": slugs,
        "editions": eds,
        "prev_editions": prev_eds,
        "total_common": total,
        "cases": (n_cases or 0),
        "volume_basis": ("at_volume" if n_cases else "best_deal"),
        "rows": rows,
        "summary": {
            "common_products": total,
            "wins_effective": wins,
            "wins_frontline": wins_front,
            "ties": ties,
            "deal_flips": len(flips),
            "total_spread": round(total_spread, 2),
            "top_spreads": [
                {
                    "product_name": r["product_name"],
                    "spread": r["spread"],
                    "winner": r["winner_effective"],
                    "unit_volume": r["unit_volume"],
                } for r in top_spreads if (r["spread"] or 0) > 0
            ],
            "by_type": by_type,
            "insights": insights,
        },
    }


def _pretty_w(slug: str) -> str:
    """'shore_point' -> 'Shore Point' for export headers."""
    return str(slug or "").replace("_", " ").title()


@router.get("/export")
def compare_export(
    wholesalers: str = Query(..., description="2-3 comma-separated slugs"),
    q: str = Query(""),
    product_type: str = Query(""),
    only_differences: bool = Query(False),
    min_spread: float = Query(0.0, ge=0),
    cases: float = Query(0, ge=0),
    sort: str = Query("spread"),
    order: str = Query("desc"),
    user: Optional[dict] = Depends(get_optional_user),
):
    """The current comparison grid as an .xlsx download. Reuses the exact
    /products logic (same filters, volume basis, winners) so the spreadsheet
    matches what's on screen — no separate math."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    data = compare_products(
        wholesalers=wholesalers, q=q, product_type=product_type,
        only_differences=only_differences, min_spread=min_spread, cases=cases,
        sort=sort, order=order, limit=50000, user=user,
    )
    slugs = data["wholesalers"]
    eds = data["editions"]
    at_vol = bool(data.get("cases"))
    qd_label = f"QD @{int(data['cases'])}cs" if at_vol else "Best QD"
    net_label = f"Net @{int(data['cases'])}cs" if at_vol else "Best Net"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compare Prices"

    # Summary grid only (the on-screen table): one line per matched product,
    # carrying its UPC, then each distributor's three price layers.
    headers = ["Product", "Size", "Vintage", "UPC"]
    for w in slugs:
        nm = _pretty_w(w)
        headers += [f"{nm} List", f"{nm} {qd_label}", f"{nm} {net_label}"]
    headers += ["Spread $", "Spread %", "Winner", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    def _money(v):
        return round(float(v), 2) if isinstance(v, (int, float)) else None

    def _upc_txt(v):
        return str(v) if v not in (None, "") else None

    for r in data["rows"]:
        prices = r["prices"]
        size = f"{r.get('unit_qty') or ''} x {r.get('unit_volume') or ''}".strip(" x")
        win = r.get("winner_effective")
        win_txt = "Tie" if win == "tie" else (_pretty_w(win) if win else "")
        notes = []
        # A >100% spread is almost always a distributor filing/data error (e.g. a
        # pack-size mismatch under one shared barcode), not a real deal — flag it
        # in the export the same way the UI shows the "check" sticker.
        sp_pct = r.get("spread_pct")
        if sp_pct is not None and sp_pct > 100:
            notes.append(f"CHECK: {sp_pct}% spread looks suspicious (likely a "
                         f"filing/data error, e.g. pack-size mismatch). Verify with sales rep")
        for w in slugs:
            p = prices.get(w, {})
            dw = p.get("deal_window") or {}
            rng = f" ({dw.get('from')}-{dw.get('to')})" if dw.get("from") else ""
            if p.get("net_excludes_qd"):
                notes.append(f"{_pretty_w(w)}: Best QD is a limited-time deal"
                             f"{rng}, excluded from Best Net")
            elif p.get("qd_time_sensitive"):
                notes.append(f"{_pretty_w(w)}: QD is limited-time{rng}")
        row = [r.get("product_name"), size, r.get("vintage"), _upc_txt(r.get("upc"))]
        for w in slugs:
            p = prices.get(w, {})
            row += [_money(p.get("frontline")), _money(p.get("after_qd")),
                    _money(p.get("effective"))]
        row += [_money(r.get("spread")), r.get("spread_pct"), win_txt, "; ".join(notes)]
        ws.append(row)

    # force the UPC column (D) to text so barcodes keep leading zeros
    for cell in ws.iter_rows(min_col=4, max_col=4):
        cell[0].number_format = "@"

    # column widths + a frozen header
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 16
    for c in ws[1]:
        c.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"compare_{'_'.join(slugs)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/tiers")
def compare_tiers(
    wholesalers: str = Query(...),
    upc_norm: str = Query(...),
    size_key: str = Query("", description="Physical-size bucket from /products rows"),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Full QD + RIP tier ladders, side by side, for one matched product.
    Ladders come from pricing.attach_tiers — the same canonical builder the
    catalog grid and product modal use."""
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        slugs = _parse_wholesalers(wholesalers, con)
        eds = _editions_for(con, src, slugs)
        ed_pred, ed_params = _edition_pred(slugs, eds)
        df = con.execute(f"""
            SELECT wholesaler, edition, upc, product_name, product_type,
                   unit_qty, unit_volume, vintage, rip_code,
                   frontline_case_price, frontline_unit_price,
                   best_case_price, effective_case_price,
                   discount_1_qty, discount_1_amt, discount_2_qty, discount_2_amt,
                   discount_3_qty, discount_3_amt, discount_4_qty, discount_4_amt,
                   discount_5_qty, discount_5_amt
            FROM {src}
            WHERE {ed_pred} AND LTRIM(upc,'0') = ?
        """, ed_params + [upc_norm]).df()
        records = [_nan_clean(r) for r in df.to_dict(orient="records")]
        if size_key:
            records = [r for r in records if _size_key(r.get("unit_volume")) == size_key]
        # best (cheapest effective) offer per wholesaler
        chosen: dict[str, dict] = {}
        for r in records:
            cur = chosen.get(r["wholesaler"])
            eff = r.get("effective_case_price")
            cur_eff = cur.get("effective_case_price") if cur else None
            if cur is None or (eff is not None and (cur_eff is None or eff < cur_eff)):
                chosen[r["wholesaler"]] = r
        records = list(chosen.values())
        _pricing.attach_tiers(con, records)
        # Distributor's own item number (Allied ABG / Fedway SKU) for the panel.
        try:
            _attach_sku_mapping(con, records)
        except Exception:
            pass
        # LIVE today: Best QD from raw cpl (every window row), Best Net layering
        # the active RIP — same rule as the grid so the panel never contradicts it.
        qd_live = _active_qd_from_raw(con, slugs, eds, None, [upc_norm])
        for rec in records:
            try:
                pack = float(rec.get("unit_qty"))
            except (TypeError, ValueError):
                pack = 0.0
            front = rec.get("frontline_case_price")
            key = (rec["wholesaler"], str(rec.get("upc") or "").lstrip("0"),
                   _size_key(rec.get("unit_volume")), _pack_norm(rec.get("unit_qty")))
            hit = qd_live.get(key)
            qd_amt = hit[0] if hit else 0.0
            best_qd = round(front - qd_amt, 2) if front is not None and qd_amt > 0 else front
            rip_amt = _active_rip_rebate(rec.get("tiers", []) or [], None, pack)
            best_net = best_qd
            if best_qd is not None and rip_amt > 0:
                best_net = max(round(best_qd - rip_amt, 2), 0.0)
            rec["_live_qd"], rec["_live_net"] = best_qd, best_net

    out = {}
    for rec in records:
        out[rec["wholesaler"]] = {
            "product_name": rec.get("product_name"),
            "upc": rec.get("upc"),
            "edition": rec.get("edition"),
            "unit_qty": rec.get("unit_qty"),
            "unit_volume": rec.get("unit_volume"),
            "vintage": rec.get("vintage"),
            "abg_sku": rec.get("abg_sku"),
            "frontline": rec.get("frontline_case_price"),
            "after_qd": rec.get("_live_qd"),
            "effective": rec.get("_live_net"),
            "tiers": rec.get("tiers", []),
        }
    return {"wholesalers": slugs, "ladders": out}


# ===========================================================================
# RIP comparison — RIP outcome is a landed-$/case curve as a function of cases
# bought; the same product can RIP completely differently across distributors.
# ===========================================================================

from backend.rip_utils import is_bottle_unit as _is_btl_unit  # canonical: any 'b…' -> bottle
# NOTE: do NOT re-implement bottle detection here. A local copy that only matched
# the substrings 'bottle'/'btl' missed single-letter 'B' (Fedway) while pricing.py's
# canonical detector matched it, so the SAME bottle tier rendered as '1 cs' at one
# distributor and '3 cs' at another. Identity is defined once, in rip_utils.


def _buy_label(tier: dict, pack: float) -> Optional[str]:
    """Human buy requirement for a tier row: '3 btl' for a bottle-bundle tier,
    '2 cs' for a case tier. Bottle tiers are NOT collapsed to cases — otherwise a
    3-bottle and a 6-bottle tier of the same RIP both read '1 cs' and look like a
    duplicate row (they are different tiers of the same program)."""
    q = tier.get("qty")
    if q is None:
        return None
    if _is_btl_unit(tier.get("unit")):
        return f"{int(round(float(q)))} btl"
    thr = _cases_threshold(tier, pack)
    return f"{math.ceil((thr if thr is not None else float(q)) - 1e-9)} cs"


def _norm_proof(v) -> Optional[float]:
    """Normalise a proof/ABV cell to a comparable proof number. '40%'/'40' ABV
    -> 80 proof; '80'/'80 proof' -> 80. None for blank/junk."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s or s in ("na", "n/a", "none"):
        return None
    is_abv = "%" in s
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    x = float(m.group(1))
    if is_abv or x <= 50:  # ABV given (or a low number that must be ABV)
        x *= 2
    return round(x, 1)


def _cases_threshold(tier: dict, pack: float) -> Optional[float]:
    """A tier's qty expressed in PHYSICAL cases (bottle-unit tiers / pack).

    Case-credit model (FOUNDATION): when a half-case rule gives this SKU's
    case a credit < 1.0, the printed case tier takes qty/credit physical
    cases to satisfy ("need 2 CS to qualify for the 1-CS RIP"). attach_tiers
    sets ``case_credit`` only when a rule matched; absence means 1.0.
    Bottle-unit tiers are explicit bottle counts and never scale."""
    q = tier.get("qty")
    if q is None:
        return None
    if _is_btl_unit(tier.get("unit")) and pack and pack > 0:
        return q / pack
    try:
        cc = float(tier.get("case_credit") or 1.0)
    except (TypeError, ValueError):
        cc = 1.0
    return float(q) / cc if cc > 0 else float(q)


def _landed_at(tiers: list, frontline: Optional[float], n_cases: float, pack: float) -> Optional[float]:
    """Best landed $/case at n_cases — min price_after over every QD/RIP tier
    whose threshold is cleared at n_cases (tiers already stack QD+RIP)."""
    best = frontline if frontline is not None else None
    for t in tiers:
        pa = t.get("price_after")
        thr = _cases_threshold(t, pack)
        if pa is None or thr is None:
            continue
        if n_cases + 1e-9 >= thr and (best is None or pa < best):
            best = pa
    return round(best, 2) if best is not None else None


def _rip_rebate_at(tiers: list, n_cases: float, pack: float) -> float:
    """Best RIP-only rebate $/case at n_cases (source == 'rip')."""
    best = 0.0
    for t in tiers:
        if t.get("source") != "rip":
            continue
        thr = _cases_threshold(t, pack)
        if thr is None or n_cases + 1e-9 < thr:
            continue
        save = t.get("rip_only_save_per_case")
        if save is None:
            save = t.get("save_per_case") or 0.0
        if save > best:
            best = save
    return round(best, 2)


def _min_cases_to_rip(tiers: list, pack: float) -> Optional[int]:
    """Fewest cases that unlock ANY RIP rebate (the 'less money required' metric)."""
    import math as _m
    mins = []
    for t in tiers:
        if t.get("source") != "rip":
            continue
        thr = _cases_threshold(t, pack)
        if thr is not None and thr > 0:
            mins.append(_m.ceil(thr - 1e-9))
    return min(mins) if mins else None


def _rip_tier_rows(tiers: list, pack: float) -> list[dict]:
    """Normalised RIP tiers for the side-by-side table."""
    import math as _m
    rows = []
    for t in tiers:
        if t.get("source") != "rip":
            continue
        thr = _cases_threshold(t, pack)
        rows.append({
            "cases_to_unlock": _m.ceil(thr - 1e-9) if thr is not None else None,
            "buy_label": _buy_label(t, pack),
            "code": t.get("code"),
            "raw_qty": t.get("qty"),
            "unit": t.get("unit"),
            "rebate_per_case": t.get("rip_only_save_per_case")
                if t.get("rip_only_save_per_case") is not None else t.get("save_per_case"),
            "price_after": t.get("price_after"),
            "window_status": t.get("window_status"),
            "is_time_sensitive": bool(t.get("is_time_sensitive")),
            "from_date": t.get("from_date"),
            "to_date": t.get("to_date"),
            # case-credit model: present only when a half-case rule matched
            "case_credit": t.get("case_credit"),
            "split_pack": t.get("split_pack"),
            "split_credit": t.get("split_credit"),
        })
    rows.sort(key=lambda r: (r["cases_to_unlock"] if r["cases_to_unlock"] is not None else 1e9))
    return rows


def _best_rip_tier_lines(tiers: list, pack: float) -> list[dict]:
    """One line per RIP tier for the Best RIPs board, carrying the
    Needed-for-Purchase / RIP-Profit economics the card shows:

      buy_label            '2 cs' / '3 btl' — the qualifying buy
      cases                physical cases to unlock (None for a pure bottle tier)
      rebate_per_case      RIP-only rebate $/case (the rebate that comes back)
      total_rebate         cases * rebate_per_case  — the '/$100' in '2C / $100'
      after_qd_per_case    list - QD at this tier's volume (NET OF QD, before the
                           RIP refund). Identity: a RIP tier's price_after already
                           = case_price - (RIP + QD), so price_after + RIP rebate
                           = the after-QD price (per FOUNDATION; same identity the
                           Compare RIPs unlock sticker uses).
      needed_for_purchase  cases * after_qd_per_case — the cash you put down
      rip_profit_pct       rebate / after-QD outlay * 100 (= total_rebate /
                           needed_for_purchase) — return on the cash down

    Bottle-unit and half-case (case_credit) tiers convert via the canonical
    _cases_threshold / _buy_label — no unit math re-implemented here."""
    import math as _m
    lines = []
    for t in tiers:
        if t.get("source") != "rip":
            continue
        thr = _cases_threshold(t, pack)
        cases = _m.ceil(thr - 1e-9) if thr is not None else None
        rpc = t.get("rip_only_save_per_case")
        if rpc is None:
            rpc = t.get("save_per_case") or 0.0
        pa = t.get("price_after")
        after_qd = (pa + rpc) if pa is not None else None
        # Sanity guard: a RIP rebate can NEVER exceed the cash you put down
        # (profit >= 100% is physically impossible). Such a tier is a source-data
        # error — e.g. an extra zero in a distributor's RIP amount ($4,200 vs
        # $420) — so drop it rather than show an absurd profit %.
        if after_qd is None or after_qd <= 0 or (rpc and rpc >= after_qd):
            continue
        needed = (round(cases * after_qd, 2)
                  if cases is not None and after_qd is not None else None)
        total_rebate = round(cases * rpc, 2) if cases is not None and rpc else None
        profit_pct = (round(rpc / after_qd * 100, 1)
                      if after_qd and after_qd > 0 and rpc else None)
        lines.append({
            "buy_label": _buy_label(t, pack),
            "cases": cases,
            "code": t.get("code"),
            "unit": t.get("unit"),
            "rebate_per_case": round(rpc, 2) if rpc else None,
            "total_rebate": total_rebate,
            "after_qd_per_case": round(after_qd, 2) if after_qd is not None else None,
            "price_after": pa,
            "needed_for_purchase": needed,
            "rip_profit_pct": profit_pct,
            "window_status": t.get("window_status"),
            "is_time_sensitive": bool(t.get("is_time_sensitive")),
            "from_date": t.get("from_date"),
            "to_date": t.get("to_date"),
        })
    lines.sort(key=lambda r: (r["cases"] if r["cases"] is not None else 1e9))
    return lines


def _rip_active_days(tiers: list, ref_date=None) -> int:
    """Distinct days in the edition month a RIP rebate is live for this SKU.
    A whole-month / evergreen RIP = the full month; dated windows contribute
    their day-span (clamped to the month). Higher = the rebate covers more days."""
    import calendar as _cal
    from datetime import date as _date, timedelta as _td
    ref = _pricing._to_date(ref_date) or _pricing.eastern_today()
    m_start = _date(ref.year, ref.month, 1)
    m_end = _date(ref.year, ref.month, _cal.monthrange(ref.year, ref.month)[1])
    full = (m_end - m_start).days + 1
    days = set()
    for t in tiers:
        if t.get("source") != "rip":
            continue
        st = t.get("window_status")
        if st in ("evergreen", "whole_month"):
            return full
        if st not in ("active", "upcoming"):
            continue
        f, to = _pricing._to_date(t.get("from_date")), _pricing._to_date(t.get("to_date"))
        if not f or not to:
            continue
        d, b = max(f, m_start), min(to, m_end)
        while d <= b:
            days.add(d); d += _td(days=1)
    return len(days)


def _rip_expires_in(tiers: list, ref_date=None) -> Optional[int]:
    """Days until the nearest LIVE dated RIP ends (urgency to buy). Looks only at
    dated windows that are live TODAY; a concurrent whole-month rebate no longer
    masks them, so a deal that is whole-month on one tier but ends Jun 17 on a
    deeper tier still reads as ending soon. None when no dated window is live."""
    ref = _pricing._to_date(ref_date) or _pricing.eastern_today()
    cands = []
    for t in tiers:
        if t.get("source") != "rip" or t.get("window_status") != "active":
            continue
        to = _pricing._to_date(t.get("to_date"))
        if to:
            cands.append((to - ref).days)
    return min(cands) if cands else None


def _rip_has_time_sensitive(tiers: list) -> bool:
    """True when this SKU has a dated/time-limited RIP window (live now or
    starting later this month), regardless of any concurrent whole-month rebate.
    This is what makes a rebate 'time-limited' rather than always-on."""
    return any(t.get("source") == "rip"
               and t.get("window_status") in ("active", "upcoming")
               for t in tiers)


def _rip_has_upcoming(tiers: list) -> bool:
    """A deeper RIP that hasn't started yet this month (plan ahead)."""
    return any(t.get("source") == "rip" and t.get("window_status") == "upcoming"
               for t in tiers)


def _rip_deepest(tiers: list, pack: float):
    """Deepest RIP rebate $/case reachable at ANY volume, and the cases to reach
    it. Returns (rebate, cases) or (0.0, None)."""
    import math as _m
    best, at = 0.0, None
    for t in tiers:
        if t.get("source") != "rip":
            continue
        save = t.get("rip_only_save_per_case")
        if save is None:
            save = t.get("save_per_case") or 0.0
        if save > best:
            best = save
            thr = _cases_threshold(t, pack)
            at = _m.ceil(thr - 1e-9) if thr is not None else None
    return (round(best, 2), at) if best > 0 else (0.0, None)


def _p360_tier_rows(tiers: list, pack: float, source: str) -> list[dict]:
    """Labeled tier ladder for the Price 360 breakdown — QD or RIP, each with
    the resulting case AND bottle price, sorted by cases-to-unlock."""
    import math as _m
    rows = []
    for t in tiers:
        if t.get("source") != source:
            continue
        thr = _cases_threshold(t, pack)
        pa = t.get("price_after")
        save = (t.get("rip_only_save_per_case")
                if source == "rip" and t.get("rip_only_save_per_case") is not None
                else t.get("save_per_case"))
        rows.append({
            "cases_to_unlock": _m.ceil(thr - 1e-9) if thr is not None else None,
            "buy_label": _buy_label(t, pack),
            "code": t.get("code"),
            "raw_qty": t.get("qty"),
            "unit": t.get("unit"),
            "save_per_case": save,
            "price_after": pa,
            "price_after_btl": round(pa / pack, 2) if pa is not None and pack else None,
            "window_status": t.get("window_status"),
            "is_time_sensitive": bool(t.get("is_time_sensitive")),
            "from_date": t.get("from_date"), "to_date": t.get("to_date"),
        })
    rows.sort(key=lambda r: (r["cases_to_unlock"] if r["cases_to_unlock"] is not None else 1e9))
    return rows


def _case_mix_sizes(con, src: str, slugs: list[str], eds: dict[str, str]) -> dict[tuple, int]:
    """(wholesaler, rip_code) -> number of distinct UPCs you can MIX under that
    rebate code to reach a tier.

    Counted from the RIP SHEET, the same authoritative source the RipMembersModal
    (/catalog/rip-siblings) uses, so the 'Mix to qualify' number always matches
    the product list that opens when you click the RIP code. Counting from the
    CPL's own rip_code column instead under-counts: a wholesaler can stack a SKU
    under several rebates yet reference only one on its CPL row."""
    ed_pred, ed_params = _edition_pred(slugs, eds)
    try:
        rip_src = read_parquet(con, "rip")
    except Exception:
        return {}
    # LEFT JOIN the rebate's distinct UPCs to the CPL: one output row per CPL
    # listing (a UPC carried in two sizes/vintages counts twice, exactly as the
    # modal lists it) plus one row per rebate UPC missing from the CPL (the modal
    # shows those as 'not in current CPL' stubs). COUNT(*) therefore equals the
    # modal's item count.
    sql = f"""
        WITH codes AS (
            SELECT DISTINCT wholesaler, edition,
                   CAST(rip_code AS VARCHAR) AS code,
                   LTRIM(CAST(upc AS VARCHAR), '0') AS un
            FROM {rip_src}
            WHERE {ed_pred}
              AND rip_code IS NOT NULL
              AND CAST(rip_code AS VARCHAR) NOT IN ('', '0', 'None', 'nan')
              AND upc IS NOT NULL
              AND CAST(upc AS VARCHAR) NOT IN ('', '0', 'None', 'nan')
              AND LTRIM(CAST(upc AS VARCHAR), '0') NOT IN ('', 'None', 'nan')
        )
        SELECT c.wholesaler, c.code, COUNT(*) AS n
        FROM codes c
        LEFT JOIN {src} e
          ON e.wholesaler = c.wholesaler AND e.edition = c.edition
         AND LTRIM(CAST(e.upc AS VARCHAR), '0') = c.un
        GROUP BY c.wholesaler, c.code
    """
    try:
        rows = con.execute(sql, ed_params).fetchall()
    except Exception:
        return {}
    return {(w, c): int(n) for w, c, n in rows}


def _product_case_mix(rec: dict, mix: dict, w: str) -> Optional[int]:
    codes = [c.strip() for c in re.split(r"\s+", str(rec.get("rip_code") or "")) if c.strip()]
    sizes = [mix.get((w, c)) for c in codes if (w, c) in mix]
    return max(sizes) if sizes else None


def _rip_verdict(row: dict, slugs: list[str], n: float) -> dict:
    """Plain-language recommendation over the structured break-even data — which
    distributor's RIP is the better buy and why. Deterministic (no LLM): every
    row gets one instantly, derived from spread, thresholds, flips and combos."""
    ni = int(n)
    d = row["dists"]
    w, sp = row["winner_at_n"], row["spread_at_n"]
    mins = {x: d[x]["min_cases"] for x in slugs if d[x]["min_cases"]}
    soonest = min(mins, key=mins.get) if mins else None
    parts, pick = [], w

    if w and w != "tie" and sp:
        parts.append(f"{w.title()} is the better RIP at {ni} case(s): "
                     f"${sp:.2f}/case lower landed cost.")
    elif w == "tie":
        parts.append(f"Landed cost ties at {ni} case(s).")
        pick = "tie"
    else:
        parts.append(f"No clear RIP edge at {ni} case(s).")

    # When the price is a tie/near-tie, the RIP TERMS are the tiebreaker: who needs
    # the least cash down to unlock the rebate.
    if (sp is None or sp < 1.0):
        invs = {x: d[x].get("unlock_investment") for x in slugs
                if d[x].get("unlock_investment") is not None}
        if len(invs) == len(slugs):
            lo_w = min(invs, key=invs.get)
            gap = max(invs.values()) - min(invs.values())
            if gap >= 1.0:
                parts.append(f"Prices are about the same, but {lo_w.title()} needs "
                             f"${gap:,.0f} less cash down to unlock its RIP "
                             f"(${invs[lo_w]:,.0f} vs ${max(invs.values()):,.0f}).")
                if pick in (None, "tie"):
                    pick = lo_w

    if soonest and len(set(mins.values())) > 1:
        c = mins[soonest]
        parts.append(f"{soonest.title()} unlocks its RIP soonest "
                     f"({c} case{'s' if c != 1 else ''} down).")
        if pick in (None, "tie"):
            pick = soonest

    if row["flips"]:
        be = "; ".join(
            f"{r['from']}{('-' + str(r['to'])) if r['to'] else '+'} cs to {r['winner']}"
            for r in row["breakeven"] if r["winner"])
        parts.append(f"Best choice shifts with volume: {be}.")

    combos = [x for x in slugs if d[x]["is_combination"]]
    if combos and len(combos) < len(slugs):
        cm = d[combos[0]]["case_mix"]
        parts.append(f"{combos[0].title()}'s RIP is a combination"
                     + (f" (mix across {cm} items)" if cm else "")
                     + ", easier to hit the tier.")

    return {"pick": pick, "text": " ".join(parts)}


@router.get("/rips")
def compare_rips(
    wholesalers: str = Query("allied,fedway", description="2-3 comma-separated slugs"),
    cases: float = Query(5, ge=1, description="Cases you plan to buy (drives winner@N)"),
    q: str = Query(""),
    product_type: str = Query(""),
    brand: str = Query("", description="Brand name contains"),
    only_differences: bool = Query(False),
    min_diff: float = Query(1.0, ge=0, description="Only show products whose best-vs-rest price gap at the chosen volume is at least this many $/case (0 = show all)"),
    include_anomalies: bool = Query(False, description="Include rows flagged as likely data issues (same UPC, very different list prices = probable pack mismatch). Hidden by default."),
    time_sensitive_only: bool = Query(False, description="Only products where some distributor's RIP is a dated/time-limited deal"),
    combo_only: bool = Query(False, description="Only products with a combination/case-mix RIP at some distributor"),
    expiring_only: bool = Query(False, description="Only products with a live RIP that ends this month at some distributor"),
    timing_diff_only: bool = Query(False, description="Only products where distributors differ on rebate TIMING (one dated/time-limited, the other all-month)"),
    qty_diff_only: bool = Query(False, description="Only products where distributors differ on the QUANTITY of cases needed to unlock the rebate"),
    better_terms_only: bool = Query(False, description="Only products where the per-case price is about the same (within $1) but the RIP TERMS differ (less cash to unlock, wider product mix, fewer cases)"),
    sort: str = Query("spread", description="spread | left_on_table | product | min_cases | best1 | deepest | active_days"),
    order: str = Query("desc"),
    limit: int = Query(2000, ge=1, le=50000),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Compare RIP OUTCOMES across 2-3 distributors for every product they ALL
    carry a RIP on. Each distributor's RIP is normalised to a landed-$/case
    ladder; we report the winner at the chosen volume, the full break-even
    map, the min cases to unlock a rebate, the best 1-case outcome, and the
    case-mix breadth."""
    import math as _m
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        slugs = _parse_wholesalers(wholesalers, con)
        eds = _editions_for(con, src, slugs)
        raw = _common_rows(con, src, slugs, eds)

        by_key: dict[str, dict[str, dict]] = {}
        for r in raw:
            by_key.setdefault(r["match_key"], {})[r["wholesaler"]] = r

        # Candidate products: carried by ALL selected distributors, each showing
        # a RIP SIGNAL (precomputed has_rip OR a real rip_code). The precise gate
        # — every distributor actually has RIP tiers from the canonical ladder —
        # is applied after attach_tiers below, so Opici text-coded RIPs the stale
        # has_rip flag misses still show without waiting for a cache rebuild.
        keys = [k for k, per in by_key.items()
                if len(per) == len(slugs) and all(_has_rip_signal(per[w]) for w in slugs)]

        # Apply the text/brand/type narrowing HERE, before the expensive tier
        # build, so a product-name search only runs attach_tiers on the handful of
        # matching products instead of the whole shared-RIP universe (~2000 rows).
        # The summary + counts below then reflect the searched context.
        if q or brand or product_type:
            qq, bb, pt = q.lower(), brand.lower(), product_type.lower()
            def _match(per):
                recs = per.values()
                if pt and not any((r.get("product_type") or "").lower() == pt for r in recs):
                    return False
                if bb and not any(bb in (r.get("brand") or "").lower() for r in recs):
                    return False
                if qq and not any(
                    qq in (r.get("product_name") or "").lower()
                    or qq in (r.get("brand") or "").lower()
                    or qq in str(r.get("upc") or "").lstrip("0")
                    for r in recs):
                    return False
                return True
            keys = [k for k in keys if _match(by_key[k])]

        flat = [by_key[k][w] for k in keys for w in slugs]
        _pricing.attach_tiers(con, flat)
        try:
            _pricing.attach_rip_gaps(con, flat)   # no-RIP day gaps between windows
        except Exception:
            pass
        mix = _case_mix_sizes(con, src, slugs, eds)

    n = float(cases)
    rows = []
    for key in keys:
        per = by_key[key]
        any_row = per[slugs[0]]
        dists, packs, tiers_by_w = {}, {}, {}
        for w in slugs:
            rec = per[w]
            pack = rec.get("uqd") or 1.0
            tiers = rec.get("tiers", []) or []
            packs[w] = pack
            tiers_by_w[w] = tiers
            case_mix = _product_case_mix(rec, mix, w)
            rip_n = _rip_rebate_at(tiers, n, pack)
            rip_1 = _rip_rebate_at(tiers, 1, pack)
            # Combination RIP: the qualifying quantity can be MIXED across more
            # than one product/size under the same code (statute's combination
            # logic) — true when the code spans >1 listing or a tier says so.
            is_combo = bool(case_mix and case_mix > 1) or any(
                "combo" in str(t.get("description") or "").lower()
                or "combination" in str(t.get("description") or "").lower()
                for t in tiers if t.get("source") == "rip")
            front = rec.get("frontline_case_price")
            deepest_rebate, deepest_at = _rip_deepest(tiers, pack)
            comp = _compliance_flags(tiers, pack)
            rtr = _rip_tier_rows(tiers, pack)
            # First rebate you can unlock: the fewest cases that turns on any RIP.
            # Investment = cases * the ACTUAL best price per case at that volume (so
            # it matches the headline, which already bakes in the quantity
            # discount); money back = cases * the RIP rebate at that volume. Using
            # the real landed price keeps the sticker, the headline and the
            # List/Discount/RIP breakdown all in agreement.
            uc = _min_cases_to_rip(tiers, pack)
            if uc:
                landed = _landed_at(tiers, front, uc, pack)   # after QD + RIP
                rpc = _rip_rebate_at(tiers, uc, pack)         # RIP rebate per case
                # You PAY the after-QD price up front (RIP not yet applied), then
                # the RIP comes back to you. So the cash you put down to unlock is
                # the price BEFORE the RIP = landed + the RIP that gets refunded.
                before_rip = (landed + rpc) if landed is not None else None
                unlock_cases = uc
                unlock_investment = round(uc * before_rip, 2) if before_rip is not None else None
                unlock_rebate_total = round(uc * rpc, 2)
            else:
                unlock_cases = unlock_investment = unlock_rebate_total = None
            dists[w] = {
                "frontline": front,
                "abv_proof": rec.get("abv_proof"),
                # this distributor's OWN vintage (normalised 4-digit), so the card
                # can show it per side and the buyer sees both years are the same.
                "vintage": rec.get("vintage_norm") or rec.get("vintage"),
                "landed_at_n": _landed_at(tiers, front, n, pack),
                "landed_at_1": _landed_at(tiers, front, 1, pack),
                "rip_at_1": rip_1,
                "rip_at_n": rip_n,
                # per-bottle normalisation (rebate $ spread over the pack)
                "rip_btl_at_1": round(rip_1 / pack, 2) if pack else None,
                "rip_btl_at_n": round(rip_n / pack, 2) if pack else None,
                "min_cases": _min_cases_to_rip(tiers, pack),
                "case_mix": case_mix,
                "is_combination": is_combo,
                # --- richer comparison metrics ---
                "deepest_rebate": deepest_rebate,           # best $/cs rebate at any volume
                "deepest_at_cases": deepest_at,             # cases to reach it
                "active_days": _rip_active_days(tiers),      # days this month a RIP is live
                "expires_in_days": _rip_expires_in(tiers),   # urgency (None = durable)
                "has_time_sensitive": _rip_has_time_sensitive(tiers),  # dated window exists
                "has_upcoming": _rip_has_upcoming(tiers),    # a deeper RIP starts later
                "total_rebate_at_n": round(rip_n * n, 2) if rip_n else 0.0,
                "effective_pct": (round(rip_n / front * 100, 1)
                                  if front and rip_n else 0.0),
                "pre_approval": comp["pre_approval"],        # NJ ABC statute flags
                "compliance_flags": comp["flags"],
                "rip_gaps": rec.get("rip_gaps") or [],       # no-RIP day gaps
                # first unlockable rebate: cash down + money back
                "unlock_cases": unlock_cases,
                "unlock_investment": unlock_investment,
                "unlock_rebate_total": unlock_rebate_total,
                "rip_tiers": rtr,
                "rip_code": rec.get("rip_code"),
                "product_name": rec.get("product_name"),
                "upc": rec.get("upc"),
                # each distributor's OWN listed size, so the buyer can confirm the
                # comparison is like-for-like (they always match: identity = UPC +
                # size bucket + bottles-per-case).
                "unit_qty": rec.get("unit_qty"),
                "unit_volume": rec.get("unit_volume"),
                "unit_type": rec.get("unit_type"),
            }

        # Precise RIP gate: this is a RIP-vs-RIP comparison, so every selected
        # distributor must ACTUALLY have a RIP tier from the canonical ladder.
        # Replaces the old precomputed-has_rip filter so text-coded RIPs (Opici)
        # the stale flag misses are included, and candidates with only a dangling
        # rip_code but no real tier are dropped.
        if not all(dists[w]["rip_tiers"] for w in slugs):
            continue

        # winner at N
        landed_n = {w: d["landed_at_n"] for w, d in dists.items() if d["landed_at_n"] is not None}
        winner_n = None
        spread_n = None
        if landed_n:
            lo, hi = min(landed_n.values()), max(landed_n.values())
            winners = [w for w, v in landed_n.items() if abs(v - lo) < _TIE_EPS]
            winner_n = "tie" if len(winners) > 1 else winners[0]
            spread_n = round(hi - lo, 2)

        # break-even map + curve: landed cost only changes at tier thresholds
        bpset = {1, int(_m.ceil(n))}
        for w in slugs:
            for t in tiers_by_w[w]:
                thr = _cases_threshold(t, packs[w])
                if thr is not None:
                    bpset.add(max(1, int(_m.ceil(thr - 1e-9))))
        breakpoints = sorted(bpset)[:24]
        curve = []
        for b in breakpoints:
            landed = {w: _landed_at(tiers_by_w[w], per[w].get("frontline_case_price"), b, packs[w])
                      for w in slugs}
            vals = {w: v for w, v in landed.items() if v is not None}
            win = None
            if vals:
                lo = min(vals.values())
                ws = [w for w, v in vals.items() if abs(v - lo) < _TIE_EPS]
                win = "tie" if len(ws) > 1 else ws[0]
            curve.append({"cases": b, "landed": landed, "winner": win})
        # collapse consecutive same-winner breakpoints into ranges
        ranges = []
        for pt in curve:
            if ranges and ranges[-1]["winner"] == pt["winner"]:
                ranges[-1]["to"] = pt["cases"]
            else:
                ranges.append({"from": pt["cases"], "to": pt["cases"], "winner": pt["winner"]})
        if ranges:
            ranges[-1]["to"] = None  # last range is open-ended (N+)

        # statute: a RIP comparison is only valid for like proof + size. Size
        # is already part of the match key; flag any proof disagreement so the
        # UI can warn (rare — same UPC usually means same proof).
        proofs = {_norm_proof(d["abv_proof"]) for d in dists.values()
                  if _norm_proof(d["abv_proof"]) is not None}

        # Price is only one axis. A rebate can also differ between distributors on
        # TIMING (one runs all month, the other is a dated deal that ends soon)
        # and on QUANTITY (one unlocks at 1 case, the other needs 5). Flag those
        # mismatches so the page can compare on dates and case counts, not just $.
        ts_set = {dists[x]["has_time_sensitive"] for x in slugs}
        timing_differs = len(ts_set) > 1            # some dated, some always-on
        qty_set = {dists[x]["min_cases"] for x in slugs}
        quantity_differs = len(qty_set) > 1         # different cases to unlock

        # "Same price, better RIP terms": the per-case landed cost is a tie (within
        # $1) yet the RIP TERMS differ. The buyer pays the same either way, so the
        # better choice is whoever needs less cash down to unlock the rebate, lets
        # you mix more products, or unlocks at fewer cases.
        price_tie = (spread_n is not None and spread_n < 1.0)
        invs = {x: dists[x]["unlock_investment"] for x in slugs
                if dists[x]["unlock_investment"] is not None}
        mix_set = {(dists[x]["case_mix"] or 0) for x in slugs}
        invest_differs = (len(invs) == len(slugs)
                          and (max(invs.values()) - min(invs.values())) >= 1.0)
        rip_terms_differ = invest_differs or len(mix_set) > 1 or quantity_differs
        # the standout "same price, better terms" case worth surfacing on its own
        better_terms_tie = price_tie and rip_terms_differ

        # Data-sanity check. The same barcode should be the same physical pack at
        # every distributor, so the list (frontline) prices should be in the same
        # ballpark. When they diverge wildly, the distributors are almost
        # certainly selling DIFFERENT packs under one shared UPC (e.g. a 120-count
        # master case vs a 12-count sleeve), and any rebate filed for the big pack
        # lands as an unbelievable net on the small one. Flag those so they don't
        # masquerade as a real "best deal".
        fronts = [d["frontline"] for d in dists.values() if d.get("frontline")]
        anomaly, reason = False, ""
        if len(fronts) >= 2:
            lo_f, hi_f = min(fronts), max(fronts)
            if lo_f > 0 and hi_f / lo_f > 2.5:
                anomaly = True
                reason = (f"List prices differ a lot for the same barcode "
                          f"(${lo_f:,.2f} vs ${hi_f:,.2f}). The distributors are "
                          f"likely selling different pack sizes under one UPC, so "
                          f"this comparison may not be like-for-like.")
        if not anomaly:
            for x in slugs:
                dd = dists[x]
                f, rb = dd.get("frontline"), dd.get("rip_at_n")
                if f and rb and rb / f > 0.6:
                    anomaly = True
                    reason = (f"The rebate wipes out over 60% of {x.title()}'s list "
                              f"price (${f:,.2f}). Double-check the pack size before "
                              f"trusting the net.")
                    break
        rows.append({
            "match_key": key,
            "upc_norm": key.split("|")[0],
            "size_key": key.split("|")[1] if "|" in key else "",
            "product_name": min((d["product_name"] for d in dists.values()), key=len),
            "product_type": any_row.get("product_type"),
            # vintage is part of the identity for wine, so both sides share it; the
            # UI shows it on the card for wine/sparkling/vermouth.
            "vintage": any_row.get("vintage"),
            "unit_type": any_row.get("unit_type"),   # container: bottle/can/keg label
            "proof_match": len(proofs) <= 1,
            "brand": any_row.get("brand"),
            "unit_qty": any_row.get("unit_qty"),
            "unit_volume": any_row.get("unit_volume"),
            "dists": dists,
            "winner_at_n": winner_n,
            "spread_at_n": spread_n,
            # total $ overpaid at N cases buying from anyone but the cheapest
            "left_on_table": round((spread_n or 0) * n, 2),
            "breakeven": ranges,
            "curve": curve,
            "flips": len({r["winner"] for r in ranges if r["winner"]}) > 1,
            # the landed CHOICE differs: one distributor is cheaper at the
            # chosen volume, or the winner flips as volume grows. (Structural
            # differences that don't change what you pay are not counted.)
            "has_difference": bool(
                (spread_n and spread_n > 0)
                or len({r["winner"] for r in ranges if r["winner"]}) > 1),
            "data_anomaly": anomaly,
            "anomaly_reason": reason,
            "timing_differs": timing_differs,
            "quantity_differs": quantity_differs,
            "rip_terms_differ": rip_terms_differ,
            "better_terms_tie": better_terms_tie,
        })

    # filters
    # q / brand / product_type are already applied above (before attach_tiers) so
    # search only pays for the matching products. Nothing to re-filter here.
    if time_sensitive_only:
        rows = [r for r in rows if any(
            d["has_time_sensitive"] for d in r["dists"].values())]
    if combo_only:
        rows = [r for r in rows if any(d["is_combination"] for d in r["dists"].values())]
    if expiring_only:
        rows = [r for r in rows if any(
            d["expires_in_days"] is not None for d in r["dists"].values())]
    if timing_diff_only:
        rows = [r for r in rows if r["timing_differs"]]
    if qty_diff_only:
        rows = [r for r in rows if r["quantity_differs"]]
    if better_terms_only:
        rows = [r for r in rows if r["better_terms_tie"]]

    # AI verdict per row (deterministic, over the break-even data)
    for r in rows:
        r["verdict"] = _rip_verdict(r, slugs, n)

    # summary
    wins = {w: 0 for w in slugs}
    ties = 0
    flips = 0
    least_money = {w: 0 for w in slugs}     # who needs fewest cases to unlock
    most_days = {w: 0 for w in slugs}       # who keeps a RIP live the most days
    most_mix = {w: 0 for w in slugs}        # who lets you mix across the most SKUs
    for r in rows:
        if r["winner_at_n"] == "tie":
            ties += 1
        elif r["winner_at_n"] in wins:
            wins[r["winner_at_n"]] += 1
        if r["flips"]:
            flips += 1
        mins = {w: r["dists"][w]["min_cases"] for w in slugs if r["dists"][w]["min_cases"]}
        if mins:
            lo = min(mins.values())
            for w, v in mins.items():
                if v == lo:
                    least_money[w] += 1
        days = {w: (r["dists"][w]["active_days"] or 0) for w in slugs}
        if any(days.values()):
            hi = max(days.values())
            for w, v in days.items():
                if v == hi and v > 0:
                    most_days[w] += 1
        mixes = {w: (r["dists"][w]["case_mix"] or 0) for w in slugs}
        if any(v > 1 for v in mixes.values()):
            hi = max(mixes.values())
            for w, v in mixes.items():
                if v == hi and v > 1:
                    most_mix[w] += 1

    total = len(rows)  # full common-RIP universe for this search context

    # display filter (grid only; summary above already computed over all)
    anomalies_hidden = sum(1 for r in rows if r["data_anomaly"])
    if not include_anomalies:
        rows = [r for r in rows if not r["data_anomaly"]]
    # The price-gap filters (only_differences, min_diff) make sense for the
    # default "who's cheaper" view, but they must NOT silently swallow rows when
    # the user is deliberately filtering on a NON-price axis (timing, quantity,
    # combination). Asking for "time-limited rebates" should surface them even if
    # the two distributors land within $1 of each other.
    attribute_filter = (time_sensitive_only or expiring_only or timing_diff_only
                        or qty_diff_only or combo_only or better_terms_only)
    if only_differences and not attribute_filter:
        rows = [r for r in rows if r["has_difference"]]
    if min_diff and min_diff > 0 and not attribute_filter:
        rows = [r for r in rows if (r["spread_at_n"] or 0) >= min_diff]

    keymap = {
        "spread": lambda r: r["spread_at_n"] or 0,
        "left_on_table": lambda r: r["left_on_table"] or 0,
        "product": lambda r: (r["product_name"] or "").lower(),
        "min_cases": lambda r: min((d["min_cases"] or 1e9 for d in r["dists"].values()), default=1e9),
        "best1": lambda r: max((d["rip_at_1"] or 0 for d in r["dists"].values()), default=0),
        "deepest": lambda r: max((d["deepest_rebate"] or 0 for d in r["dists"].values()), default=0),
        "active_days": lambda r: max((d["active_days"] or 0 for d in r["dists"].values()), default=0),
        # least cash to unlock the first RIP (ascending), widest product mix (desc)
        "least_investment": lambda r: min((d["unlock_investment"] for d in r["dists"].values()
                                           if d["unlock_investment"] is not None), default=1e12),
        "best_mix": lambda r: max((d["case_mix"] or 0 for d in r["dists"].values()), default=0),
    }
    # ascending sorts (smallest first): A-Z, fewest cases, least money down. The
    # UI never sends `order`, so these force ascending; the rest default desc.
    _ascending = ("product", "min_cases", "least_investment")
    rows.sort(key=keymap.get(sort, keymap["spread"]),
              reverse=False if sort in _ascending else (order != "asc"))
    rows = rows[:limit]

    insights = []
    if total:
        lead = max(wins, key=lambda w: wins[w])
        if wins[lead]:
            insights.append(
                f"At {int(n)} case(s), {lead} has the lowest price per case on "
                f"{wins[lead]} of {total} shared-RIP products.")
        lm = max(least_money, key=lambda w: least_money[w])
        if least_money[lm]:
            insights.append(
                f"{lm} requires the fewest cases to unlock a RIP on "
                f"{least_money[lm]} products (least money down).")
        md = max(most_days, key=lambda w: most_days[w])
        if most_days[md]:
            insights.append(
                f"{md} keeps a RIP live the most days on {most_days[md]} products "
                f"(its RIP is live more of the month).")
        mm = max(most_mix, key=lambda w: most_mix[w])
        if most_mix[mm]:
            insights.append(
                f"{mm} lets you mix across the most products to hit the tier on "
                f"{most_mix[mm]} RIPs (easier to qualify).")
        if flips:
            insights.append(
                f"{flips} product(s) change the best-RIP distributor as your "
                f"volume grows. Check the break-even before you commit.")

    return {
        "wholesalers": slugs,
        "editions": eds,
        "cases": n,
        "total_common": total,
        "rows": rows,
        "summary": {
            "common_rip_products": total,
            "wins_at_n": wins,
            "ties": ties,
            "flips": flips,
            "least_money": least_money,
            "most_active_days": most_days,
            "most_case_mix": most_mix,
            "anomalies_hidden": anomalies_hidden,
            "insights": insights,
        },
    }


# ===========================================================================
# Best RIPs board — discovery-first. One card per product carried by Allied,
# Fedway AND Opici, each distributor's full RIP ladder shown one line per tier
# with Needed-for-Purchase (net of QD) + RIP-Profit economics, plus flags for
# where the three distributors differ. Browse the standout rebates, then drill
# into Compare RIPs.
# ===========================================================================

# Fixed to the three NJ distributors this board compares (see project rules).
_BEST_RIP_SLUGS = ["allied", "fedway", "opici"]


@router.get("/best-rips")
def best_rips(
    q: str = Query(""),
    product_type: str = Query(""),
    brand: str = Query("", description="Brand name contains"),
    wholesalers: str = Query("", description="Comma-separated subset of Allied/Fedway/Opici to compare (empty = all three)."),
    months: str = Query("", description="Comma-separated editions (YYYY-MM) to show; empty = latest TWO editions present in the data."),
    only_differences: bool = Query(False, description="Only cards where the RIP is available at 2+ selected distributors AND they differ (one carries it without a RIP, different timing/quantity, or a profit-%% gap)."),
    min_profit: float = Query(0.0, ge=0, description="Hide cards whose best RIP profit %% is below this"),
    time_sensitive_only: bool = Query(False, description="Only products where some distributor's RIP is a dated/time-limited deal"),
    hide_expired: bool = Query(True, description="Drop tier lines whose window has already ended"),
    sort: str = Query("best_profit", description="best_profit | deepest | gap | expiring | product"),
    order: str = Query("desc"),
    limit: int = Query(2000, ge=1, le=50000),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Best RIPs board: EVERY RIP across Allied, Fedway and Opici. One card per
    product that ANY of the three files a RIP on — including products only one of
    them carries. Each distributor block shows its full RIP ladder, one line per
    tier, with the Needed-for-Purchase (net of QD) and RIP-Profit-%% economics.
    A distributor that carries the SKU but files no RIP shows as 'No RIP'; one
    that doesn't stock it at all shows as 'Not carried'.

    Scope is the three NJ distributors only. All tier/unit/rebate math comes from
    the canonical pricing.attach_tiers + rip_utils helpers — nothing
    re-implemented here. Product images come from product_enrichment (Go-UPC)."""
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        known = {r[0] for r in con.execute(
            f"SELECT DISTINCT wholesaler FROM {src}").fetchall()}
        avail = [s for s in _BEST_RIP_SLUGS if s in known]
        if not avail:
            raise HTTPException(400, "Need Allied/Fedway/Opici data loaded for this board")
        # Distributor filter: a subset of the board's three may be selected.
        # Keep the canonical order and constrain to the three (board scope).
        if wholesalers.strip():
            req = {w.strip() for w in wholesalers.split(",") if w.strip()}
            slugs = [s for s in avail if s in req]
            if not slugs:
                raise HTTPException(400, "Pick at least one of Allied, Fedway, Opici")
        else:
            slugs = avail
        # Available editions for the selected distributors, newest first.
        ph = ",".join("?" * len(slugs))
        all_eds = [r[0] for r in con.execute(
            f"SELECT DISTINCT edition FROM {src} WHERE wholesaler IN ({ph}) "
            f"ORDER BY edition DESC", slugs).fetchall()]
        if not all_eds:
            raise HTTPException(400, "No editions loaded for the selected distributors")
        # Month filter: default to the latest TWO editions PRESENT in the data
        # (not the calendar month) — so a freshly loaded July edition shows up
        # next to June immediately, before the calendar rolls over.
        if months.strip():
            want = {m.strip() for m in months.split(",") if m.strip()}
            sel_months = [m for m in all_eds if m in want]
        else:
            sel_months = all_eds[:2]
        if not sel_months:
            sel_months = all_eds[:1]

        # Which distributors carry each selected edition.
        eph = ",".join("?" * len(sel_months))
        month_ws: dict[str, list] = {}
        for w, e in con.execute(
                f"SELECT DISTINCT wholesaler, edition FROM {src} "
                f"WHERE wholesaler IN ({ph}) AND edition IN ({eph})",
                slugs + sel_months).fetchall():
            month_ws.setdefault(e, []).append(w)
        cur_ym = _pricing.current_yyyy_mm()
        # Only the top candidates can rank into the visible page, so build the
        # expensive tier ladders for a generous multiple of `limit`, not the
        # whole universe. Big headroom keeps display filters from starving.
        cand_k = max(limit * 4, 1500)

        def _absent_dist():
            return {
                "carried": False, "has_rip": False, "rip_code": None, "frontline": None,
                "case_mix": None, "deepest_rebate": None, "deepest_at_cases": None,
                "min_cases": None, "best_profit_pct": None, "active_days": None,
                "expires_in_days": None, "has_time_sensitive": False, "tiers": [],
                "unit_qty": None, "unit_volume": None,
            }

        def _cards_for_month(month: str) -> list[dict]:
            wm = [w for w in slugs if w in month_ws.get(month, [])]
            if not wm:
                return []
            eds_m = {w: month for w in wm}
            # require_all=False: keep every distributor that CARRIES each SKU.
            raw = _common_rows(con, src, wm, eds_m, require_all=False)
            bk: dict[str, dict[str, dict]] = {}
            for r in raw:
                bk.setdefault(r["match_key"], {})[r["wholesaler"]] = r
            mkeys = [k for k, per in bk.items() if any(per[w].get("has_rip") for w in per)]
            # Narrow on search BEFORE the tier build (same as compare_rips).
            if q or brand or product_type:
                qq, bb, pt = q.lower(), brand.lower(), product_type.lower()
                def _match(per):
                    recs = per.values()
                    if pt and not any((r.get("product_type") or "").lower() == pt for r in recs):
                        return False
                    if bb and not any(bb in (r.get("brand") or "").lower() for r in recs):
                        return False
                    if qq and not any(
                        qq in (r.get("product_name") or "").lower()
                        or qq in (r.get("brand") or "").lower()
                        or qq in str(r.get("upc") or "").lstrip("0")
                        for r in recs):
                        return False
                    return True
                mkeys = [k for k in mkeys if _match(bk[k])]

            n_cand = len(mkeys)   # full candidate universe for this month
            # PERF: preselect the top candidates by a CHEAP proxy from precomputed
            # columns before the expensive attach_tiers (which would otherwise run
            # over the whole ~8k-product universe just to show ~400). Proxy uses
            # rip_savings (deepest full-month rebate) and rip_savings/frontline
            # (profit), guarded against impossible amounts (source typos).
            def _proxy(k):
                best_sav, profs = 0.0, []
                for _w2, rec in bk[k].items():
                    sav = rec.get("rip_savings") or 0
                    front = rec.get("frontline_case_price") or 0
                    if sav and sav > 0 and front and sav < front:
                        best_sav = max(best_sav, sav)
                        profs.append(sav / front)
                if sort in ("deepest", "expiring"):
                    return best_sav
                if sort == "gap":
                    return (max(profs) - min(profs)) if len(profs) > 1 else 0.0
                return max(profs) if profs else 0.0   # best_profit (default)
            if sort == "product":
                mkeys.sort(key=lambda k: min(
                    (bk[k][w].get("product_name") or "" for w in bk[k]), key=len).lower())
            else:
                mkeys.sort(key=_proxy, reverse=True)
            sel = mkeys[:cand_k]

            flat = [bk[k][w] for k in sel for w in bk[k]]
            # Classify each edition's windows from ITS month: today for the
            # current edition, mid-month otherwise — so a future July edition
            # loaded in June reads as a real (not "upcoming") RIP.
            ref = None if month == cur_ym else f"{month}-15"
            _pricing.attach_tiers(con, flat, ref_date=ref)
            try:
                _attach_image(con, flat)          # rec["image_url"] from Go-UPC
            except Exception:
                pass
            mmix = _case_mix_sizes(con, src, wm, eds_m)

            out = []
            for key in sel:
                per = bk[key]
                present = [w for w in slugs if w in per]
                any_row = per[present[0]]
                dists = {}
                for w in slugs:
                    if w not in per:
                        dists[w] = _absent_dist()  # distributor lacks this SKU/edition
                        continue
                    rec = per[w]
                    pack = rec.get("uqd") or 1.0
                    tiers = rec.get("tiers", []) or []
                    lines = _best_rip_tier_lines(tiers, pack)
                    if hide_expired:
                        lines = [ln for ln in lines if ln["window_status"] != "expired"]
                    # Trust the canonical tier builder, not the precomputed has_rip
                    # flag, which can lag (Opici text-coded RIPs).
                    has_rip = bool(lines)
                    deepest_rebate, deepest_at = _rip_deepest(tiers, pack)
                    best_profit = max((ln["rip_profit_pct"] or 0 for ln in lines), default=0.0)
                    dists[w] = {
                        "carried": True,
                        "has_rip": has_rip,
                        "rip_code": rec.get("rip_code"),
                        "frontline": rec.get("frontline_case_price"),
                        "case_mix": _product_case_mix(rec, mmix, w),
                        "deepest_rebate": deepest_rebate,
                        "deepest_at_cases": deepest_at,
                        "min_cases": _min_cases_to_rip(tiers, pack),
                        "best_profit_pct": round(best_profit, 1) if best_profit else None,
                        "active_days": _rip_active_days(tiers, ref),
                        "expires_in_days": _rip_expires_in(tiers, ref),
                        "has_time_sensitive": _rip_has_time_sensitive(tiers),
                        "tiers": lines,
                        "unit_qty": rec.get("unit_qty"),
                        "unit_volume": rec.get("unit_volume"),
                    }

                ripping = [w for w in present if dists[w]["has_rip"]]
                missing = [w for w in present if not dists[w]["has_rip"]]
                not_carried = [w for w in slugs if w not in per]
                profits = {w: dists[w]["best_profit_pct"] for w in ripping
                           if dists[w]["best_profit_pct"]}
                best_w = best_profit_pct = profit_delta = None
                if profits:
                    best_w = max(profits, key=profits.get)
                    best_profit_pct = profits[best_w]
                    if len(profits) > 1:
                        second = sorted(profits.values(), reverse=True)[1]
                        profit_delta = round(best_profit_pct - second, 1)
                profit_gap = (round(max(profits.values()) - min(profits.values()), 1)
                              if len(profits) > 1 else 0.0)

                # "Differs" needs the RIP at 2+ selected distributors (a single
                # offer is not a difference); then diverging profit/timing/quantity
                # or a distributor carrying it without a RIP. Not-carried is absence.
                timing_differs = len({dists[w]["has_time_sensitive"] for w in ripping}) > 1
                quantity_differs = len({dists[w]["min_cases"] for w in ripping}) > 1
                differs = len(ripping) >= 2 and (
                    profit_gap >= 1.0 or timing_differs or quantity_differs or bool(missing))

                image_url = next((per[w].get("image_url") for w in present
                                  if per[w].get("image_url")), None)

                out.append({
                    "match_key": f"{month}|{key}",
                    "edition": month,
                    "upc_norm": key.split("|")[0],
                    "size_key": key.split("|")[1] if "|" in key else "",
                    "product_name": min((per[w].get("product_name") for w in present),
                                        key=lambda s: len(s or "")),
                    "product_type": any_row.get("product_type"),
                    "brand": any_row.get("brand"),
                    "vintage": any_row.get("vintage"),
                    "unit_qty": any_row.get("unit_qty"),
                    "unit_volume": any_row.get("unit_volume"),
                    "unit_type": any_row.get("unit_type"),
                    "upc": any_row.get("upc"),
                    "image_url": image_url,
                    "dists": dists,
                    "ripping": ripping,
                    "missing": missing,
                    "not_carried": not_carried,
                    "best_distributor": best_w,
                    "best_profit_pct": best_profit_pct,
                    "profit_delta": profit_delta,
                    "profit_gap": profit_gap,
                    "deepest_rebate": max((dists[w]["deepest_rebate"] or 0 for w in slugs), default=0.0),
                    "timing_differs": timing_differs,
                    "quantity_differs": quantity_differs,
                    "differs": differs,
                    "soonest_expiry": min((dists[w]["expires_in_days"] for w in slugs
                                           if dists[w]["expires_in_days"] is not None), default=None),
                })
            return out, n_cand

        rows = []
        universe = 0
        for m in sel_months:
            cards, n_cand = _cards_for_month(m)
            rows.extend(cards)
            universe += n_cand

        # ---- Month-over-month RIP trend sticker ------------------------------
        # RIP codes are recycled each edition, so DON'T match by code across
        # months — track the AMOUNT (rip_savings = deepest full-month rebate per
        # case, i.e. the deepest tier) keyed by product identity (UPC + size).
        # CALENDAR-relative: last / this / next month around the current edition.
        # Only meaningful on the current edition's cards (a past-month card isn't
        # "this month"). A slot with no loaded edition stays null (so an unloaded
        # next month never shows a value).
        def _shift_ym(ym: str, delta: int) -> str:
            i = int(ym[:4]) * 12 + (int(ym[5:7]) - 1) + delta
            return f"{i // 12:04d}-{i % 12 + 1:02d}"
        slot_ed = {"last": _shift_ym(cur_ym, -1), "this": cur_ym, "next": _shift_ym(cur_ym, 1)}
        t_eds = [e for e in slot_ed.values() if e in all_eds]   # loaded slots only
        trend_amt: dict = {}   # (upc_norm, size_key, edition) -> max rip_savings
        if t_eds:
            tph = ",".join("?" * len(t_eds))
            for _w, e, upc, uv, rs in con.execute(
                    f"SELECT wholesaler, edition, upc, unit_volume, rip_savings FROM {src} "
                    f"WHERE wholesaler IN ({ph}) AND edition IN ({tph}) "
                    f"AND rip_savings IS NOT NULL AND rip_savings > 0 "
                    # drop impossible rebates (>= case price) — source typos
                    f"AND frontline_case_price IS NOT NULL "
                    f"AND rip_savings < frontline_case_price",
                    slugs + t_eds).fetchall():
                k = (str(upc or "").lstrip("0"), _size_key(uv), e)
                if float(rs) > trend_amt.get(k, 0):
                    trend_amt[k] = round(float(rs), 2)
        for r in rows:
            ident = (r["upc_norm"], r["size_key"])
            amts = {s: (trend_amt.get((*ident, e)) if e in all_eds else None)
                    for s, e in slot_ed.items()}
            present = {s: v for s, v in amts.items() if v is not None}
            best = None
            # only the current edition's cards carry the calendar this/last/next
            if r["edition"] == cur_ym and len(present) >= 2:
                top = max(present, key=present.get)
                best = "this" if ("this" in present and present[top] - present["this"] < 1.0) else top
            r["rip_trend"] = {
                "this": amts["this"], "last": amts["last"], "next": amts["next"],
                "this_ed": slot_ed["this"], "last_ed": slot_ed["last"], "next_ed": slot_ed["next"],
                "best": best,
            }

    if time_sensitive_only:
        rows = [r for r in rows
                if any(r["dists"][w]["has_time_sensitive"] for w in r["ripping"])]
    total = universe  # full candidate universe (pre-display-filter) for this context
    if only_differences:
        rows = [r for r in rows if r["differs"]]
    if min_profit > 0:
        rows = [r for r in rows if (r["best_profit_pct"] or 0) >= min_profit]

    keymap = {
        "best_profit": lambda r: r["best_profit_pct"] or 0,
        "deepest": lambda r: r["deepest_rebate"] or 0,
        "gap": lambda r: r["profit_gap"] or 0,
        # soonest-expiring first (smallest days); None = no dated window -> last
        "expiring": lambda r: r["soonest_expiry"] if r["soonest_expiry"] is not None else 1e9,
        "product": lambda r: (r["product_name"] or "").lower(),
    }
    _ascending = ("product", "expiring")
    rows.sort(key=keymap.get(sort, keymap["best_profit"]),
              reverse=False if sort in _ascending else (order != "asc"))
    rows = rows[:limit]

    return {
        "wholesalers": slugs,
        "months": sel_months,
        "available_months": all_eds,
        "total": total,
        "rows": rows,
    }


def assistant_rip_comparison(con, match: str, wholesalers: Optional[list[str]] = None,
                             cases: float = 5) -> dict:
    """Single-product RIP-outcome comparison for the AI assistant. Resolves
    `match` (product name or UPC) to the SKU that the most of the selected
    distributors put a RIP on, then returns each distributor's landed $/case
    at `cases`, best 1-case rebate, min cases to unlock, case-mix breadth,
    combination flag, full break-even map and a verdict — the same logic as
    the Compare RIPs page, scoped to one product."""
    import math as _m
    src = read_parquet(con, "cpl_enriched")
    known = {r[0] for r in con.execute(f"SELECT DISTINCT wholesaler FROM {src}").fetchall()}
    slugs = [s for s in (wholesalers or ["allied", "fedway", "opici"]) if s in known]
    if len(slugs) < 2:
        return {"found": False, "note": "Need at least two valid distributors to compare."}
    eds = _editions_for(con, src, slugs)
    raw = _common_rows(con, src, slugs, eds)

    # Group ALL distributors per identity first; the same UPC is often named
    # differently across distributors, so match the text to find candidate
    # KEYS, then compare every distributor that carries that key.
    by_key: dict[str, dict] = {}
    for r in raw:
        by_key.setdefault(r["match_key"], {})[r["wholesaler"]] = r

    m = (match or "").strip().lower()
    digits = re.sub(r"\D", "", m)
    is_upc = len(digits) >= 8
    cand_keys = []
    for k, per in by_key.items():
        hit = (k.split("|")[0] == digits.lstrip("0")) if is_upc else any(
            m in (rec.get("product_name") or "").lower() for rec in per.values())
        if hit:
            cand_keys.append(k)
    best_key, best_n = None, 0
    for k in cand_keys:
        per = by_key[k]
        n_rip = sum(1 for w in slugs if per.get(w) and _has_rip_signal(per[w]))
        if n_rip >= 2 and n_rip > best_n:
            best_n, best_key = n_rip, k
    if not best_key:
        return {"found": False, "match": match,
                "note": "No product matched with a RIP at 2+ of the selected distributors."}

    per = by_key[best_key]
    present = [w for w in slugs if per.get(w) and _has_rip_signal(per[w])]
    flat = [per[w] for w in present]
    _pricing.attach_tiers(con, flat)
    # Precise gate on the canonical ladder: keep only distributors that actually
    # have a RIP tier (drops a dangling rip_code with no real tier; includes
    # Opici text codes the stale has_rip flag misses).
    present = [w for w in present
               if any(t.get("source") == "rip" for t in (per[w].get("tiers") or []))]
    if len(present) < 2:
        return {"found": False, "match": match,
                "note": "No product matched with a RIP at 2+ of the selected distributors."}
    mix = _case_mix_sizes(con, src, present, eds)
    n = float(cases)

    dists, packs, tiers_by_w = {}, {}, {}
    for w in present:
        rec = per[w]
        pack = rec.get("uqd") or 1.0
        tiers = rec.get("tiers", []) or []
        packs[w], tiers_by_w[w] = pack, tiers
        case_mix = _product_case_mix(rec, mix, w)
        dists[w] = {
            "landed_at_n": _landed_at(tiers, rec.get("frontline_case_price"), n, pack),
            "frontline": rec.get("frontline_case_price"),
            "rip_at_1": _rip_rebate_at(tiers, 1, pack),
            "rip_at_n": _rip_rebate_at(tiers, n, pack),
            "min_cases": _min_cases_to_rip(tiers, pack),
            "case_mix": case_mix,
            "is_combination": bool(case_mix and case_mix > 1),
            "rip_code": rec.get("rip_code"),
            "rip_tiers": _rip_tier_rows(tiers, pack),
        }

    landed_n = {w: d["landed_at_n"] for w, d in dists.items() if d["landed_at_n"] is not None}
    winner, spread = None, None
    if landed_n:
        lo, hi = min(landed_n.values()), max(landed_n.values())
        ws = [w for w, v in landed_n.items() if abs(v - lo) < _TIE_EPS]
        winner = "tie" if len(ws) > 1 else ws[0]
        spread = round(hi - lo, 2)

    bpset = {1, int(_m.ceil(n))}
    for w in present:
        for t in tiers_by_w[w]:
            thr = _cases_threshold(t, packs[w])
            if thr is not None:
                bpset.add(max(1, int(_m.ceil(thr - 1e-9))))
    curve = []
    for b in sorted(bpset)[:24]:
        landed = {w: _landed_at(tiers_by_w[w], per[w].get("frontline_case_price"), b, packs[w])
                  for w in present}
        vals = {w: v for w, v in landed.items() if v is not None}
        win = None
        if vals:
            lo = min(vals.values())
            ws2 = [w for w, v in vals.items() if abs(v - lo) < _TIE_EPS]
            win = "tie" if len(ws2) > 1 else ws2[0]
        curve.append({"cases": b, "winner": win})
    ranges = []
    for pt in curve:
        if ranges and ranges[-1]["winner"] == pt["winner"]:
            ranges[-1]["to"] = pt["cases"]
        else:
            ranges.append({"from": pt["cases"], "to": pt["cases"], "winner": pt["winner"]})
    if ranges:
        ranges[-1]["to"] = None

    row = {
        "winner_at_n": winner, "spread_at_n": spread,
        "breakeven": ranges, "flips": len({r["winner"] for r in ranges if r["winner"]}) > 1,
        "dists": dists,
    }
    return {
        "found": True,
        "product_name": min((per[w].get("product_name") for w in present), key=len),
        "unit_volume": per[present[0]].get("unit_volume"),
        "unit_qty": per[present[0]].get("unit_qty"),
        "distributors": present,
        "cases": n,
        "winner_at_n": winner,
        "spread_at_n": spread,
        "breakeven": ranges,
        "dists": dists,
        "verdict": _rip_verdict(row, present, n)["text"],
    }


# ===========================================================================
# Price 360 — one holistic label per product comparing every wholesaler offer
# across all price layers, reduced to a reachability-adjusted effective net
# cost. Invoice cost (legal, discounts only) is kept separate from economic
# net cost (rebates incl. RIP). Score weights are FIXED + PUBLISHED.
# ===========================================================================

# Published, fixed value-score weights (sum 100). Identical for every offer.
PRICE360_WEIGHTS = {"net_cost": 70, "savings": 15, "stability": 10, "compliance": 5}

# NJ ABC RIP statutory limits (for the pre-approval flag).
_RIP_MAX_CASES = 50
_RIP_MAX_REBATE = 1000.0
_RIP_SMALL_QTY_CASES = 5


def _compliance_flags(tiers: list, pack: float) -> dict:
    """NJ ABC RIP limit checks → pre-approval flags (#17)."""
    rip = [t for t in tiers if t.get("source") == "rip"]
    flags = []
    if any((_cases_threshold(t, pack) or 0) > _RIP_MAX_CASES for t in rip):
        flags.append(f"RIP tier over {_RIP_MAX_CASES} cases")
    if any((t.get("amount") or 0) > _RIP_MAX_REBATE for t in rip):
        flags.append(f"single rebate over ${int(_RIP_MAX_REBATE):,}")
    if rip and not any((_cases_threshold(t, pack) or 1e9) <= _RIP_SMALL_QTY_CASES for t in rip):
        flags.append(f"no small-quantity tier (≤{_RIP_SMALL_QTY_CASES} cases)")
    return {"flags": flags, "pre_approval": bool(flags)}


def _reachability(rip_rebate_case: float, qualifying: Optional[int],
                  typical_cases: Optional[float], mode: str) -> dict:
    """Soft/hard/off reachability (#13/14). Returns the likelihood (0..1) the
    retailer hits the qualifying quantity and the rebate value to actually
    credit. Soft scales the rebate by likelihood; hard zeroes it if unreachable;
    off always credits full value (badge only). No history -> full value."""
    if not rip_rebate_case or not qualifying:
        return {"status": "no_rip", "likelihood": 1.0, "credited_rebate": 0.0,
                "qualifying": qualifying, "typical": typical_cases}
    if typical_cases is None:
        return {"status": "unknown", "likelihood": 1.0,
                "credited_rebate": round(rip_rebate_case, 2),
                "qualifying": qualifying, "typical": None}
    ratio = typical_cases / qualifying if qualifying else 1.0
    if mode == "hard":
        like = 1.0 if ratio >= 1 else 0.0
    elif mode == "off":
        like = 1.0
    else:  # soft
        like = max(0.0, min(1.0, ratio))
    status = ("likely" if ratio >= 1 else ("partial" if ratio > 0 else "unreachable"))
    return {"status": status, "likelihood": round(like, 2),
            "credited_rebate": round(rip_rebate_case * like, 2),
            "qualifying": qualifying, "typical": round(typical_cases, 1)}


def _value_score(net_case: Optional[float], min_net: Optional[float],
                 adj_savings_pct: float, full_month: bool, compliant: bool) -> dict:
    """Net-cost-dominant 0-100 composite (#8/20). Net cost dominates so the
    score never contradicts the net-cost ranking. Weights are PRICE360_WEIGHTS."""
    w = PRICE360_WEIGHTS
    net = w["net_cost"] * (min_net / net_case) if (net_case and min_net) else 0.0
    sav = w["savings"] * max(0.0, min(1.0, adj_savings_pct / 25.0))
    stab = w["stability"] * (1.0 if full_month else 0.5)
    comp = w["compliance"] * (1.0 if compliant else 0.0)
    total = round(net + sav + stab + comp, 1)
    return {"score": min(100.0, total), "breakdown": {
        "net_cost": round(net, 1), "savings": round(sav, 1),
        "stability": round(stab, 1), "compliance": round(comp, 1),
        "weights": w}}


def _fetch_product_offers(con, src: str, match: str, size_key: Optional[str] = None):
    """Resolve `match` to the product identity carried by the MOST distributors,
    return (identity_key, one best row per distributor) — every offer is the
    same UPC + size, so directly comparable."""
    all_slugs = [r[0] for r in con.execute(f"SELECT DISTINCT wholesaler FROM {src}").fetchall()]
    eds = _editions_for(con, src, all_slugs)
    ed_pred, ed_params = _edition_pred(all_slugs, eds)
    m = (match or "").strip()
    digits = re.sub(r"\D", "", m)
    is_upc = len(digits) >= 8
    match_pred = "LTRIM(upc,'0') = ?" if is_upc else "lower(product_name) LIKE ?"
    match_param = digits.lstrip("0") if is_upc else f"%{m.lower()}%"
    # Resolve the text to UPCs first, then pull EVERY distributor carrying those
    # UPCs — the same barcode is often named differently per distributor (Allied
    # 'CAMPARI APERITIVO' vs Fedway 'CAMPARI BITTERS'), so a name-only filter
    # would silently drop offers. (No combo-code exclusion: a product's
    # standalone CPL price is a valid offer regardless of combo linkage, and a
    # blanket combo_code filter wrongly drops Allied/Shore-Point internal codes.)
    base = f"{ed_pred} AND {_VALID_UPC}"
    params = list(ed_params) + list(ed_params) + [match_param]
    vn = _pricing.vintage_norm_sql("vintage")
    df = con.execute(f"""
        SELECT wholesaler, edition, upc, product_name, product_type, brand,
               unit_qty, unit_volume, vintage, abv_proof,
               frontline_case_price, frontline_unit_price,
               best_case_price, best_unit_price, effective_case_price,
               next_effective_case_price,
               rip_savings, has_discount, has_rip, rip_code, rip_windows,
               discount_1_qty, discount_1_amt, discount_2_qty, discount_2_amt,
               discount_3_qty, discount_3_amt, discount_4_qty, discount_4_amt,
               discount_5_qty, discount_5_amt,
               LTRIM(upc,'0') AS upc_norm, TRY_CAST(unit_qty AS DOUBLE) AS uqd,
               {vn} AS vintage_norm,
               UPPER(product_type) IN ('WINE','SPARKLING','VERMOUTH') AS vintage_sensitive
        FROM {src} e WHERE {base}
          AND LTRIM(upc,'0') IN (
              SELECT DISTINCT LTRIM(upc,'0') FROM {src}
              WHERE {base} AND {match_pred})
    """, params).df()
    recs = [_nan_clean(r) for r in df.to_dict(orient="records")]
    if not recs:
        return None, []
    for r in recs:
        r["match_key"] = "|".join([
            r["upc_norm"], _size_key(r.get("unit_volume")), _pack_norm(r.get("unit_qty")),
            (r.get("vintage_norm") or "") if r.get("vintage_sensitive") else ""])
    # identity carried by the most distributors
    by_key: dict[str, dict] = {}
    for r in recs:
        by_key.setdefault(r["match_key"], {})
        cur = by_key[r["match_key"]].get(r["wholesaler"])
        if cur is None or (r.get("effective_case_price") or 1e9) < (cur.get("effective_case_price") or 1e9):
            by_key[r["match_key"]][r["wholesaler"]] = r
    # every distinct size found for this product (for the size selector),
    # deduped by size bucket — keep the listing carried by the most distributors.
    seen: dict[str, dict] = {}
    for k, per in by_key.items():
        rep = next(iter(per.values()))
        sk = k.split("|")[1] if "|" in k else ""
        vstr = str(rep.get("vintage") or "").strip()
        if vstr.lower() in ("", "0", "0.0", "nan", "none"):
            vstr = ""
        entry = {"match_key": k, "size_key": sk,
                 "unit_volume": rep.get("unit_volume"), "unit_qty": rep.get("unit_qty"),
                 "vintage": vstr or None, "n_distributors": len(per)}
        cur = seen.get(sk)
        if cur is None or entry["n_distributors"] > cur["n_distributors"]:
            seen[sk] = entry
    available = sorted(seen.values(), key=lambda a: -a["n_distributors"])
    chosen = None
    if size_key:
        chosen = next((k for k in by_key
                       if (k.split("|")[1] if "|" in k else "") == size_key), None)
    if not chosen:
        chosen = max(by_key, key=lambda k: len(by_key[k]))
    return chosen, list(by_key[chosen].values()), available


def price360_offers(con, match: str, typical_map: Optional[dict] = None,
                    reach_mode: str = "soft", size_key: Optional[str] = None) -> dict:
    """The Price 360 label data: every wholesaler offer for one product, each
    reduced to a reachability-adjusted effective net cost, ranked cheapest
    first. Reuses pricing.attach_tiers; no pricing math re-implemented."""
    src = read_parquet(con, "cpl_enriched")
    key, recs, available = _fetch_product_offers(con, src, match, size_key)
    if not recs:
        return {"found": False, "match": match, "note": "No product matched a valid barcode."}
    _pricing.attach_tiers(con, recs)
    all_slugs = list({r["wholesaler"] for r in recs})
    eds = _editions_for(con, src, all_slugs)
    mix = _case_mix_sizes(con, src, all_slugs, eds)
    typical_map = typical_map or {}

    offers = []
    for rec in recs:
        w = rec["wholesaler"]
        pack = rec.get("uqd") or 1.0
        tiers = rec.get("tiers", []) or []
        invoice_case = rec.get("best_case_price")            # legal cost basis (discounts only)
        economic_case = rec.get("effective_case_price")      # after RIP rebates
        frontline_case = rec.get("frontline_case_price")
        rip_rebate = round(max(0.0, (invoice_case or 0) - (economic_case or 0)), 2) \
            if invoice_case is not None and economic_case is not None else 0.0
        qualifying = _min_cases_to_rip(tiers, pack)
        reach = _reachability(rip_rebate, qualifying,
                              typical_map.get(rec["upc_norm"]), reach_mode)
        # reachability-adjusted net: invoice minus the credited (reachable) rebate
        net_case = round((invoice_case if invoice_case is not None else frontline_case or 0)
                         - reach["credited_rebate"], 2)
        case_mix = _product_case_mix(rec, mix, w)
        comp = _compliance_flags(tiers, pack)
        full_month = any(t.get("source") == "rip" and t.get("window_status") in (None, "whole_month", "evergreen", "active")
                         and not t.get("is_time_sensitive") for t in tiers) or rip_rebate == 0
        sav_case = round((frontline_case or 0) - net_case, 2) if frontline_case else 0.0
        sav_pct = round(sav_case / frontline_case * 100, 1) if frontline_case else 0.0
        offers.append({
            "wholesaler": w, "edition": rec.get("edition"),
            "product_name": rec.get("product_name"), "upc": rec.get("upc"),
            "frontline_case": frontline_case,
            "frontline_btl": rec.get("frontline_unit_price"),
            "invoice_case": invoice_case,
            "invoice_btl": rec.get("best_unit_price"),
            "net_case": net_case, "net_btl": round(net_case / pack, 2) if pack else None,
            "rip_rebate_full": rip_rebate,
            "rip_rebate_credited": reach["credited_rebate"],
            "savings_case": sav_case, "savings_pct": sav_pct,
            "reachability": reach,
            "divergence": bool(economic_case is not None and invoice_case is not None
                               and economic_case < invoice_case - 0.005),
            "compliance": comp,
            "case_mix": case_mix, "single_sku": not (case_mix and case_mix > 1),
            "abv_proof": rec.get("abv_proof"),
            "unit_volume": rec.get("unit_volume"), "unit_qty": rec.get("unit_qty"),
            "qd_tiers": _p360_tier_rows(tiers, pack, "discount"),
            "rip_tiers": _p360_tier_rows(tiers, pack, "rip"),
            "full_month": bool(full_month),
            "_pack": pack,
        })

    # rank by reachability-adjusted net cost (authoritative)
    ranked = [o for o in offers if o["net_case"] is not None]
    min_net = min((o["net_case"] for o in ranked), default=None)
    max_rebate = max((o["rip_rebate_full"] for o in offers), default=0)
    for o in offers:
        sc = _value_score(o["net_case"], min_net, o["savings_pct"],
                          o["full_month"], not o["compliance"]["pre_approval"])
        o["value_score"] = sc["score"]
        o["score_breakdown"] = sc["breakdown"]
    offers.sort(key=lambda o: (o["net_case"] if o["net_case"] is not None else 1e9))
    for o in offers:
        nc = o["net_case"]
        # competition ranking: every offer tied at the lowest net cost is a
        # co-winner (rank 1) — identical offers are never arbitrarily ordered.
        o["rank"] = 1 + sum(1 for x in offers if x["net_case"] is not None
                            and nc is not None and x["net_case"] < nc - 0.005)
        o["is_winner"] = nc is not None and min_net is not None and abs(nc - min_net) < 0.005
        # "bigger rebate, costs more" (#11): biggest headline rebate yet a
        # STRICTLY higher net cost than the winner — the rebate misleads.
        o["rebate_misleads"] = bool(
            not o["is_winner"] and max_rebate > 0
            and o["rip_rebate_full"] >= max_rebate - 0.005
            and min_net is not None and nc is not None
            and nc > min_net + 0.005)
        o.pop("_pack", None)
    n_winners = sum(1 for o in offers if o["is_winner"])

    meta = recs[0]
    # Same UPC + size => directly comparable. A differing filed proof is a data
    # warning, not a different product, so it never blocks ranking (#4/5).
    proofs = {_norm_proof(o["abv_proof"]) for o in offers if _norm_proof(o["abv_proof"]) is not None}
    return {
        "found": True,
        "product": {
            "product_name": min((o["product_name"] for o in offers), key=len),
            "upc": meta.get("upc"), "unit_volume": meta.get("unit_volume"),
            "unit_qty": meta.get("unit_qty"), "unit_type": meta.get("unit_type"),
            "abv_proof": meta.get("abv_proof"),
            "product_type": meta.get("product_type"), "brand": meta.get("brand"),
        },
        "comparability": "direct",
        "proof_warning": len(proofs) > 1,
        "reach_mode": reach_mode,
        "weights": PRICE360_WEIGHTS,
        "tie": n_winners > 1,
        "n_winners": n_winners,
        "size_key": (key.split("|")[1] if key and "|" in key else ""),
        "available_sizes": available,
        "offers": offers,
    }


def _typical_cases_map(user_id: int, upc_norms: list[str]) -> dict:
    """Median cases this retailer historically orders per UPC (for soft
    reachability). Empty when there's no order history."""
    if not user_id or not upc_norms:
        return {}
    try:
        from backend.pg import get_pg
        with get_pg() as pg:
            ph = ", ".join(["%s"] * len(upc_norms))
            rows = pg.execute(
                f"""SELECT LTRIM(ol.upc, '0') AS u, AVG(ol.qty_cases) AS avg_cs
                    FROM order_lines ol JOIN orders o ON o.id = ol.order_id
                    WHERE o.user_id = %s AND ol.qty_cases > 0
                      AND LTRIM(ol.upc, '0') IN ({ph})
                    GROUP BY 1""",
                (user_id, *upc_norms)).fetchall()
        return {dict(r)["u"]: float(dict(r)["avg_cs"]) for r in rows if dict(r)["avg_cs"]}
    except Exception:
        return {}


@router.get("/price360")
def price360(
    match: str = Query(..., description="Product name or UPC"),
    reach_mode: str = Query("soft", description="soft | hard | off (reachability)"),
    size: str = Query("", description="size_key to scope to one size (else most-carried)"),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Price 360 label: every wholesaler's offer for ONE product, each reduced
    to a reachability-adjusted effective net cost (case + bottle), ranked
    cheapest first, with invoice vs economic cost kept separate, a fixed-weight
    value score, NJ-ABC pre-approval flags, and full layer traceability."""
    if reach_mode not in ("soft", "hard", "off"):
        reach_mode = "soft"
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        key, recs, _av = _fetch_product_offers(con, src, match, size or None)
        upc_norms = list({r["upc_norm"] for r in recs}) if recs else []
    typical = _typical_cases_map(user.get("id") if user else 0, upc_norms) if recs else {}
    with get_duckdb() as con:
        return price360_offers(con, match, typical, reach_mode, size or None)


# ===========================================================================
# Edition comparison — one distributor across two CPL periods, every diff
# expressed in Price 360 net-cost terms (not raw frontline).
# ===========================================================================

def _wholesaler_editions(con, src: str, w: str) -> list:
    rows = con.execute(
        f"SELECT DISTINCT edition FROM {src} WHERE wholesaler = ? ORDER BY edition DESC",
        [w]).fetchall()
    return [r[0] for r in rows]


def _edition_rows(con, src: str, w: str, ed: str) -> dict:
    """One best row per SKU identity for a (wholesaler, edition). Same-distributor
    spelling is stable, so the identity is built in SQL."""
    vn = _pricing.vintage_norm_sql("vintage")
    ident_sql = ("LTRIM(upc,'0') || '|' || COALESCE(unit_volume,'') || '|' "
                 "|| COALESCE(CAST(TRY_CAST(unit_qty AS DOUBLE) AS VARCHAR),'') || '|' "
                 f"|| COALESCE({vn},'')")
    df = con.execute(f"""
        SELECT upc, product_name, product_type, brand, unit_qty, unit_volume, unit_type,
               vintage, abv_proof,
               frontline_case_price, frontline_unit_price,
               best_case_price, best_unit_price, effective_case_price,
               rip_savings, has_rip, has_discount, rip_code,
               LTRIM(upc,'0') AS upc_norm, TRY_CAST(unit_qty AS DOUBLE) AS uqd,
               {vn} AS vintage_norm, {ident_sql} AS ident
        FROM {src}
        WHERE wholesaler = ? AND edition = ? AND {_VALID_UPC}
        QUALIFY ROW_NUMBER() OVER (PARTITION BY {ident_sql}
            ORDER BY effective_case_price NULLS LAST) = 1
    """, [w, ed]).df()
    return {r["ident"]: _nan_clean(r) for r in df.to_dict(orient="records")}


def _layer_changes(a: dict, b: dict) -> list:
    """Which underlying layers moved between editions (B11)."""
    out = []

    def chg(k):
        return abs((a.get(k) or 0) - (b.get(k) or 0)) > 0.005

    if chg("frontline_case_price"):
        out.append("frontline")
    if chg("best_case_price"):
        out.append("discount")   # invoice (single-case / QD) moved
    if bool(a.get("has_rip")) != bool(b.get("has_rip")):
        out.append("rip_gained" if b.get("has_rip") else "rip_lost")
    elif chg("rip_savings"):
        out.append("rip_modified")
    return out


@router.get("/editions/options")
def edition_options(wholesaler: str = Query(...),
                    user: Optional[dict] = Depends(get_optional_user)):
    """Available CPL periods for a distributor + the default (latest two)."""
    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        eds = _wholesaler_editions(con, src, wholesaler)
    if not eds:
        raise HTTPException(404, f"No editions for {wholesaler}")
    return {
        "wholesaler": wholesaler, "editions": eds,
        "default_newer": eds[0],
        "default_older": eds[1] if len(eds) > 1 else None,
        "single_edition": len(eds) < 2,
    }


def edition_comparison(con, wholesaler: str, older: str = "", newer: str = "",
                       scope: str = "catalog", match: str = "", change: str = "",
                       sort: str = "net_delta", order: str = "desc",
                       limit: int = 3000) -> dict:
    """Core edition comparison (shared by the HTTP endpoint and the assistant
    tool). Compares one distributor across two editions; every diff is the
    change in effective NET cost, with the layer that moved + add/remove
    classification + a delta summary."""
    src = read_parquet(con, "cpl_enriched")
    eds = _wholesaler_editions(con, src, wholesaler)
    if not eds:
        return {"wholesaler": wholesaler, "error": f"No editions for {wholesaler}"}
    if len(eds) < 2:
        return {"wholesaler": wholesaler, "single_edition": True, "editions": eds,
                "note": f"{wholesaler} has only one edition ({eds[0]}), nothing to compare."}
    nw = newer if newer in eds else eds[0]
    ol = older if older in eds else eds[1]
    if nw == ol:
        ol = next((e for e in eds if e != nw), ol)
    a = _edition_rows(con, src, wholesaler, ol)   # older
    b = _edition_rows(con, src, wholesaler, nw)   # newer

    m = (match or "").strip().lower()
    digits = re.sub(r"\D", "", m)
    is_upc = len(digits) >= 8

    def matches(rec):
        if not m:
            return True
        return (rec.get("upc_norm") == digits.lstrip("0")) if is_upc \
            else (m in (rec.get("product_name") or "").lower())

    rows = []
    summary = {"rose": 0, "fell": 0, "unchanged": 0, "added": 0, "removed": 0,
               "rip_changed": 0, "not_comparable": 0}
    for k in (set(a) | set(b)):
        ra, rb = a.get(k), b.get(k)
        ref = rb or ra
        if m and not matches(ref):
            continue
        pack = (ref.get("uqd") or 1.0)
        if ra and rb:
            comparable = (_norm_proof(ra.get("abv_proof")) == _norm_proof(rb.get("abv_proof"))
                          or _norm_proof(ra.get("abv_proof")) is None
                          or _norm_proof(rb.get("abv_proof")) is None)
            net_a, net_b = ra.get("effective_case_price"), rb.get("effective_case_price")
            delta = round((net_b or 0) - (net_a or 0), 2) if net_a is not None and net_b is not None else None
            pct = round(delta / net_a * 100, 1) if (delta is not None and net_a) else None
            layers = _layer_changes(ra, rb)
            if not comparable:
                summary["not_comparable"] += 1
            elif delta is None or abs(delta) < 0.005:
                summary["unchanged"] += 1
            elif delta > 0:
                summary["rose"] += 1
            else:
                summary["fell"] += 1
            if any(str(l).startswith("rip") for l in layers):
                summary["rip_changed"] += 1
            rows.append({
                "ident": k, "status": "both", "comparable": comparable,
                "product_name": ref.get("product_name"), "unit_volume": ref.get("unit_volume"),
                "unit_qty": ref.get("unit_qty"), "unit_type": ref.get("unit_type"),
                "product_type": ref.get("product_type"),
                "upc": ref.get("upc"),
                "net_a_case": net_a, "net_b_case": net_b,
                "net_a_btl": round(net_a / pack, 2) if net_a and pack else None,
                "net_b_btl": round(net_b / pack, 2) if net_b and pack else None,
                "net_delta_case": delta, "net_delta_pct": pct,
                "net_delta_btl": round(delta / pack, 2) if delta is not None and pack else None,
                "frontline_a": ra.get("frontline_case_price"), "frontline_b": rb.get("frontline_case_price"),
                "invoice_a": ra.get("best_case_price"), "invoice_b": rb.get("best_case_price"),
                "rip_a": ra.get("rip_savings"), "rip_b": rb.get("rip_savings"),
                "layers": layers,
            })
        elif rb:
            summary["added"] += 1
            rows.append({"ident": k, "status": "added", "comparable": True,
                         "product_name": rb.get("product_name"), "unit_volume": rb.get("unit_volume"),
                         "unit_qty": rb.get("unit_qty"), "unit_type": rb.get("unit_type"), "product_type": rb.get("product_type"),
                         "upc": rb.get("upc"), "net_b_case": rb.get("effective_case_price"),
                         "net_b_btl": round((rb.get("effective_case_price") or 0) / pack, 2) if pack else None,
                         "net_delta_case": None, "layers": []})
        else:
            summary["removed"] += 1
            rows.append({"ident": k, "status": "removed", "comparable": True,
                         "product_name": ra.get("product_name"), "unit_volume": ra.get("unit_volume"),
                         "unit_qty": ra.get("unit_qty"), "unit_type": ra.get("unit_type"), "product_type": ra.get("product_type"),
                         "upc": ra.get("upc"), "net_a_case": ra.get("effective_case_price"),
                         "net_a_btl": round((ra.get("effective_case_price") or 0) / pack, 2) if pack else None,
                         "net_delta_case": None, "layers": []})

    total = len(rows)
    if change == "increase":
        rows = [r for r in rows if (r.get("net_delta_case") or 0) > 0]
    elif change == "decrease":
        rows = [r for r in rows if (r.get("net_delta_case") or 0) < 0]
    elif change == "added":
        rows = [r for r in rows if r["status"] == "added"]
    elif change == "removed":
        rows = [r for r in rows if r["status"] == "removed"]
    elif change == "rip":
        rows = [r for r in rows if any(str(l).startswith("rip") for l in r.get("layers", []))]
    elif change == "changed":
        rows = [r for r in rows if r["status"] != "both"
                or (r.get("net_delta_case") and abs(r["net_delta_case"]) >= 0.005)]

    keyf = {
        "net_delta": lambda r: r.get("net_delta_case") if r.get("net_delta_case") is not None else 0,
        "net_delta_pct": lambda r: r.get("net_delta_pct") or 0,
        "product": lambda r: (r.get("product_name") or "").lower(),
    }
    rows.sort(key=keyf.get(sort, keyf["net_delta"]),
              reverse=(order != "asc") if sort != "product" else (order == "desc"))
    rows = rows[:limit]

    return {
        "wholesaler": wholesaler, "single_edition": False,
        "older": ol, "newer": nw, "editions": eds, "scope": scope,
        "total": total, "summary": summary, "rows": rows,
    }


@router.get("/editions")
def compare_editions(
    wholesaler: str = Query(...),
    older: str = Query(""),
    newer: str = Query(""),
    scope: str = Query("catalog", description="catalog | product"),
    match: str = Query(""),
    change: str = Query("", description="all|increase|decrease|added|removed|rip|changed"),
    sort: str = Query("net_delta", description="net_delta | net_delta_pct | product"),
    order: str = Query("desc"),
    limit: int = Query(3000, ge=1, le=50000),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Compare one distributor across two editions; every diff is the change in
    effective NET cost (case + bottle, $ + %), which layer moved, with
    added/removed classification and a delta summary."""
    with get_duckdb() as con:
        return edition_comparison(con, wholesaler, older, newer, scope, match,
                                  change, sort, order, limit)


# ===========================================================================
# Rate Shop — the clarity-first flagship: one product, MY quantity, one ranked
# answer with net-at-volume, the conditions to capture each price, a break-even
# band across volume, and a stretch-to-next-tier nudge. Reuses every engine
# helper above; no new pricing math.
# ===========================================================================

def _applied_tier_at(tiers: list, frontline: Optional[float], n: float, pack: float):
    """Best landed price at n cases and the tier that produces it (or None =
    base/frontline)."""
    best = frontline if frontline is not None else None
    best_t = None
    for t in tiers:
        thr = _cases_threshold(t, pack)
        pa = t.get("price_after")
        if thr is None or pa is None:
            continue
        if n + 1e-9 >= thr and (best is None or pa < best - 1e-9):
            best, best_t = pa, t
    return (round(best, 2) if best is not None else None), best_t


def _next_tier_above(tiers: list, n: float, pack: float, current: Optional[float]):
    """The nearest tier above n cases that beats the current price (for the
    'stretch to unlock' nudge)."""
    cand = None
    for t in tiers:
        thr = _cases_threshold(t, pack)
        pa = t.get("price_after")
        if thr is None or pa is None or current is None:
            continue
        if thr > n + 1e-9 and pa < current - 0.005:
            if cand is None or thr < cand[0]:
                cand = (thr, t, pa)
    return cand


def _rateshop_conditions(applied_t, pack: float, case_mix, is_combination, compliance) -> list:
    """What the buyer must DO to capture this offer's price at the chosen volume."""
    import math as _m
    conds = []
    if applied_t is not None:
        thr = _cases_threshold(applied_t, pack)
        is_rip = applied_t.get("source") == "rip"
        if thr is not None:
            conds.append({"type": "qty", "text": f"buy ≥{_m.ceil(thr - 1e-9)} cs"})
        if is_rip:
            if applied_t.get("is_time_sensitive") and applied_t.get("window_status") != "expired":
                fd, td = applied_t.get("from_date"), applied_t.get("to_date")
                conds.append({"type": "window",
                              "text": f"valid {str(fd)[5:] if fd else '?'} to {str(td)[5:] if td else '?'}"})
            conds.append({"type": "invoice", "text": "single invoice"})
            if is_combination and case_mix:
                conds.append({"type": "combo", "text": f"mix across {case_mix} items"})
    if compliance.get("pre_approval"):
        conds.append({"type": "preapproval", "text": "needs pre-approval"})
    return conds


def rateshop_data(con, match: str, cases: float = 5, size_key: Optional[str] = None) -> dict:
    """Rate Shop core (shared by the HTTP endpoint and the assistant tool):
    every distributor's offer for ONE product at the quantity the buyer plans
    to purchase — true landed net cost, conditions, break-even, stretch nudge,
    timing. Ranks by net-at-volume."""
    import math as _m
    n = float(cases)
    src = read_parquet(con, "cpl_enriched")
    key, recs, available = _fetch_product_offers(con, src, match, size_key)
    if not recs:
        return {"found": False, "match": match, "note": "No product matched a valid barcode."}
    _pricing.attach_tiers(con, recs)
    try:
        from backend.enrichment_join import attach_sku_mapping
        attach_sku_mapping(con, recs)
    except Exception:
        pass
    all_slugs = list({r["wholesaler"] for r in recs})
    eds = _editions_for(con, src, all_slugs)
    mix = _case_mix_sizes(con, src, all_slugs, eds)

    offers, tiers_by_w, packs = [], {}, {}
    for rec in recs:
        w = rec["wholesaler"]
        pack = rec.get("uqd") or 1.0
        tiers = rec.get("tiers", []) or []
        packs[w], tiers_by_w[w] = pack, tiers
        frontline = rec.get("frontline_case_price")
        net_n, applied = _applied_tier_at(tiers, frontline, n, pack)
        case_mix = _product_case_mix(rec, mix, w)
        is_combo = bool(case_mix and case_mix > 1)
        comp = _compliance_flags(tiers, pack)
        conds = _rateshop_conditions(applied, pack, case_mix, is_combo, comp)
        nxt = _next_tier_above(tiers, n, pack, net_n)
        stretch = None
        if nxt:
            thr_n, t_n, pa_n = nxt
            stretch = {"to_cases": _m.ceil(thr_n - 1e-9),
                       "extra_per_case": round((net_n or 0) - pa_n, 2),
                       "price_after": pa_n}
        sav = round((frontline or 0) - (net_n or 0), 2) if frontline and net_n is not None else 0.0
        # timing: is this product cheaper (or pricier) next CPL period?
        next_eff = rec.get("next_effective_case_price")
        cur_eff = rec.get("effective_case_price")
        timing = None
        if next_eff is not None and cur_eff is not None:
            delta = round(next_eff - cur_eff, 2)
            if abs(delta) >= 0.50:
                timing = {"dir": "drop" if delta < 0 else "rise",
                          "next_case": round(next_eff, 2), "delta": delta}
        offers.append({
            "wholesaler": w, "edition": rec.get("edition"),
            "product_name": rec.get("product_name"), "upc": rec.get("upc"),
            "sku": rec.get("abg_sku"),
            "frontline_case": frontline, "frontline_btl": rec.get("frontline_unit_price"),
            "net_case": net_n, "net_btl": round(net_n / pack, 2) if (net_n is not None and pack) else None,
            "savings_case": sav, "savings_pct": round(sav / frontline * 100, 1) if frontline else 0.0,
            "timing": timing,
            "applied_kind": (None if applied is None else ("RIP" if applied.get("source") == "rip" else "QD")),
            "applied_code": (applied.get("code") if applied else None),
            "conditions": conds,
            "stretch": stretch,
            "case_mix": case_mix, "single_sku": not is_combo,
            "compliance": comp, "abv_proof": rec.get("abv_proof"),
            "qd_tiers": _p360_tier_rows(tiers, pack, "discount"),
            "rip_tiers": _p360_tier_rows(tiers, pack, "rip"),
        })

    # rank by net-at-volume (authoritative); co-winners on ties
    nets = [o["net_case"] for o in offers if o["net_case"] is not None]
    min_net = min(nets) if nets else None
    offers.sort(key=lambda o: o["net_case"] if o["net_case"] is not None else 1e9)
    for o in offers:
        nc = o["net_case"]
        o["rank"] = 1 + sum(1 for x in offers if x["net_case"] is not None and nc is not None and x["net_case"] < nc - 0.005)
        o["is_winner"] = nc is not None and min_net is not None and abs(nc - min_net) < 0.005
    n_winners = sum(1 for o in offers if o["is_winner"])

    # break-even across volume — net cost changes only at tier thresholds
    bpset = {1, int(_m.ceil(n))}
    for w in all_slugs:
        for t in tiers_by_w[w]:
            thr = _cases_threshold(t, packs[w])
            if thr is not None:
                bpset.add(max(1, int(_m.ceil(thr - 1e-9))))
    fronts = {o["wholesaler"]: o["frontline_case"] for o in offers}
    curve = []
    for b in sorted(bpset)[:24]:
        landed = {w: _applied_tier_at(tiers_by_w[w], fronts.get(w), b, packs[w])[0] for w in all_slugs}
        vals = {w: v for w, v in landed.items() if v is not None}
        win = None
        if vals:
            lo = min(vals.values())
            ws = [w for w, v in vals.items() if abs(v - lo) < _TIE_EPS]
            win = "tie" if len(ws) > 1 else ws[0]
        curve.append({"cases": b, "net": landed, "winner": win})
    ranges = []
    for pt in curve:
        if ranges and ranges[-1]["winner"] == pt["winner"]:
            ranges[-1]["to"] = pt["cases"]
        else:
            ranges.append({"from": pt["cases"], "to": pt["cases"], "winner": pt["winner"]})
    if ranges:
        ranges[-1]["to"] = None

    # plain-language verdict at the chosen volume
    win = offers[0] if offers else None
    verdict = ""
    if win and win["net_case"] is not None:
        if n_winners > 1:
            verdict = (f"At {int(n)} case(s), {n_winners} distributors tie at "
                       f"${win['net_case']:.2f}/case. Pick on service or delivery.")
        else:
            runner = next((o for o in offers if not o["is_winner"] and o["net_case"] is not None), None)
            gap = f", ${runner['net_case'] - win['net_case']:.2f}/cs cheaper than {distributorName_safe(runner['wholesaler'])}" if runner else ""
            verdict = f"At {int(n)} case(s), buy from {distributorName_safe(win['wholesaler'])}: ${win['net_case']:.2f}/case{gap}."
        flips = len({r['winner'] for r in ranges if r['winner'] and r['winner'] != 'tie'}) > 1
        if flips:
            be = "; ".join(f"{r['from']}{('-'+str(r['to'])) if r['to'] else '+'} cs to "
                           f"{('tie' if r['winner']=='tie' else distributorName_safe(r['winner']))}"
                           for r in ranges if r["winner"])
            verdict += f" Best choice shifts with volume: {be}."

    meta = recs[0]
    # Same UPC + size => directly comparable. A differing FILED proof across
    # distributors is a data-quality warning (not a different product), so the UI
    # must not assert an unqualified "verified match" — mirror Price 360 (#E audit).
    proofs = {_norm_proof(o.get("abv_proof")) for o in offers if _norm_proof(o.get("abv_proof")) is not None}
    return {
        "found": True, "cases": n, "size_key": (key.split("|")[1] if key and "|" in key else ""),
        "available_sizes": available,
        "proof_warning": len(proofs) > 1,
        "product": {
            "product_name": min((o["product_name"] for o in offers), key=len),
            "upc": meta.get("upc"), "unit_volume": meta.get("unit_volume"),
            "unit_qty": meta.get("unit_qty"), "unit_type": meta.get("unit_type"),
            "abv_proof": meta.get("abv_proof"),
            "product_type": meta.get("product_type"), "brand": meta.get("brand"),
        },
        "tie": n_winners > 1, "verdict": verdict,
        "breakeven": ranges, "curve": curve, "offers": offers,
    }


@router.get("/rateshop")
def rateshop(
    match: str = Query(..., description="Product name or UPC"),
    cases: float = Query(5, ge=1, description="How many cases you plan to buy"),
    size: str = Query(""),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Rate Shop: best distributor for ONE product at the quantity you plan to
    buy — true landed net cost, the conditions to capture it, a break-even map,
    a stretch nudge and next-month timing. Ranks by net-at-volume."""
    with get_duckdb() as con:
        return rateshop_data(con, match, float(cases), size or None)


def distributorName_safe(slug: str) -> str:
    try:
        from backend.routers.catalog import DISTRIBUTOR_NAMES
        return DISTRIBUTOR_NAMES.get(slug, slug)
    except Exception:
        return slug


# ===========================================================================
# Basket rate shopping — for the buyer's whole order (cart / favorites): the
# optimal split (each line from its cheapest distributor) vs single-sourcing.
# ===========================================================================

def _read_basket_lines(user_id: int, source: str) -> list[dict]:
    from backend.pg import get_pg
    with get_pg() as pg:
        if source == "favorites":
            rows = pg.execute(
                "SELECT product_name, wholesaler, upc, unit_volume, 1 AS qty "
                "FROM watchlist WHERE user_id = %s", (user_id,)).fetchall()
        else:  # cart
            rows = pg.execute(
                "SELECT product_name, wholesaler, upc, unit_volume, "
                "GREATEST(COALESCE(qty_cases,0),1) AS qty "
                "FROM cart_items WHERE user_id = %s AND COALESCE(saved_for_later,0)=0",
                (user_id,)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        un = str(d.get("upc") or "").lstrip("0")
        if not un or len(un) < 8:
            continue
        out.append({"product_name": d.get("product_name"), "wholesaler": d.get("wholesaler"),
                    "upc": d.get("upc"), "upc_norm": un,
                    "unit_volume": d.get("unit_volume"), "qty": float(d.get("qty") or 1)})
    return out


@router.get("/basket")
def basket(source: str = Query("cart", description="cart | favorites"),
           user: dict = Depends(get_current_user)):
    """Basket rate shopping: price every line of the buyer's cart (or favorites)
    across all distributors at the line's quantity, then compute the OPTIMAL
    SPLIT (each line from its cheapest distributor) vs single-sourcing the whole
    order — and what splitting saves."""
    lines = _read_basket_lines(user["id"], source if source in ("cart", "favorites") else "cart")
    if not lines:
        return {"found": False, "source": source,
                "note": f"Your {source} is empty (or has no valid-barcode lines)."}

    with get_duckdb() as con:
        src = read_parquet(con, "cpl_enriched")
        all_slugs = [r[0] for r in con.execute(f"SELECT DISTINCT wholesaler FROM {src}").fetchall()]
        eds = _editions_for(con, src, all_slugs)
        ed_pred, ed_params = _edition_pred(all_slugs, eds)
        upcs = sorted({l["upc_norm"] for l in lines})
        vn = _pricing.vintage_norm_sql("vintage")
        ph = ", ".join("?" * len(upcs))
        df = con.execute(f"""
            SELECT wholesaler, edition, upc, product_name, unit_qty, unit_volume, vintage,
                   frontline_case_price, frontline_unit_price, best_case_price,
                   effective_case_price, rip_code,
                   discount_1_qty, discount_1_amt, discount_2_qty, discount_2_amt,
                   discount_3_qty, discount_3_amt, discount_4_qty, discount_4_amt,
                   discount_5_qty, discount_5_amt,
                   LTRIM(upc,'0') AS upc_norm, TRY_CAST(unit_qty AS DOUBLE) AS uqd
            FROM {src} e
            WHERE {ed_pred} AND {_VALID_UPC} AND LTRIM(upc,'0') IN ({ph})
        """, ed_params + upcs).df()
        recs = [_nan_clean(r) for r in df.to_dict(orient="records")]
        _pricing.attach_tiers(con, recs)

    # index: (upc_norm, size_key) -> {wholesaler: best rec}
    idx: dict = {}
    for r in recs:
        k = (r["upc_norm"], _size_key(r.get("unit_volume")))
        idx.setdefault(k, {})
        cur = idx[k].get(r["wholesaler"])
        if cur is None or (r.get("effective_case_price") or 1e9) < (cur.get("effective_case_price") or 1e9):
            idx[k][r["wholesaler"]] = r

    out_lines = []
    dist_totals: dict = {}   # wholesaler -> {total, covered, lines}
    split_total = 0.0
    current_total = 0.0
    for ln in lines:
        k = (ln["upc_norm"], _size_key(ln.get("unit_volume")))
        per = idx.get(k) or {}
        n = ln["qty"]
        prices = {}
        for w, rec in per.items():
            pack = rec.get("uqd") or 1.0
            net, _t = _applied_tier_at(rec.get("tiers", []) or [], rec.get("frontline_case_price"), n, pack)
            if net is not None:
                prices[w] = round(net, 2)
        if not prices:
            out_lines.append({"product_name": ln["product_name"], "unit_volume": ln["unit_volume"],
                              "qty": n, "upc": ln["upc"], "prices": {}, "best_w": None,
                              "best_net": None, "current_w": ln["wholesaler"], "no_match": True})
            continue
        best_w = min(prices, key=prices.get)
        best_net = prices[best_w]
        line_cost = best_net * n
        split_total += line_cost
        cur_w = ln["wholesaler"]
        cur_net = prices.get(cur_w, best_net)
        current_total += cur_net * n
        for w, p in prices.items():
            dt = dist_totals.setdefault(w, {"total": 0.0, "covered": 0, "lines": 0})
            dt["total"] += p * n
            dt["covered"] += 1
        out_lines.append({
            "product_name": ln["product_name"], "unit_volume": ln["unit_volume"], "qty": n,
            "upc": ln["upc"], "prices": prices, "best_w": best_w, "best_net": best_net,
            "current_w": cur_w, "no_match": False,
            "saving_vs_current": round((cur_net - best_net) * n, 2),
        })

    n_lines = sum(1 for l in out_lines if not l["no_match"])
    # single-source candidates: distributors that carry EVERY matched line
    single = []
    for w, dt in dist_totals.items():
        single.append({"wholesaler": w, "total": round(dt["total"], 2),
                       "covered": dt["covered"], "covers_all": dt["covered"] == n_lines})
    single.sort(key=lambda s: (not s["covers_all"], s["total"]))
    best_single = next((s for s in single if s["covers_all"]), None)

    split_distributors = sorted({l["best_w"] for l in out_lines if l["best_w"]})
    return {
        "found": True, "source": source, "line_count": n_lines,
        "split_total": round(split_total, 2),
        "split_distributors": split_distributors,
        "current_total": round(current_total, 2),
        "saving_vs_current": round(current_total - split_total, 2),
        "best_single": best_single,
        "saving_vs_single": round(best_single["total"] - split_total, 2) if best_single else None,
        "single_source": single[:8],
        "lines": out_lines,
    }
