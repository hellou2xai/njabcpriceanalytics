"""
Base parser for NJ ABC eCPL files.

This handles the common logic shared by ALL wholesalers:
  - Finding the header row in each sheet
  - Mapping raw headers to canonical column names
  - Extracting data rows
  - Type coercion and cleaning

Wholesaler-specific configs override only what differs (header row hints,
extra discount tiers, product type normalization, etc.).
"""

import re
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

from nj_abc_parser.template import (
    CPL_COLUMNS, CPL_HEADER_MAP,
    RIP_COLUMNS, RIP_HEADER_MAP,
    COMBO_COLUMNS, COMBO_HEADER_MAP,
    BEER_MM_COLUMNS, BEER_MM_HEADER_MAP,
    SHEET_ALIASES,
)

logger = logging.getLogger("nj_abc_parser")


def _normalize_header(raw: str) -> str:
    """Strip whitespace, newlines, special chars from header for matching."""
    if not isinstance(raw, str):
        return ""
    # Collapse whitespace/newlines
    s = re.sub(r"[\n\r]+", " ", raw)
    s = re.sub(r"\s+", " ", s).strip()
    # Remove quotes, dollar signs, hashes for matching
    s = re.sub(r'[\"\'#]', "", s)
    return s.lower()


def _match_header(raw_header: str, header_map: dict) -> Optional[str]:
    """Find the best matching canonical column name for a raw header."""
    norm = _normalize_header(raw_header)
    if not norm:
        return None

    # Exact match first
    if norm in header_map:
        return header_map[norm]

    # Substring match — find the longest key that appears in the header
    best_key = None
    best_len = 0
    for key, canonical in header_map.items():
        if key in norm and len(key) > best_len:
            best_key = key
            best_len = len(key)

    if best_key:
        return header_map[best_key]

    return None


def _find_sheet(wb: openpyxl.Workbook, sheet_type: str) -> Optional[str]:
    """Find the actual sheet name matching a sheet type alias."""
    aliases = SHEET_ALIASES.get(sheet_type, [])
    for alias in aliases:
        if alias in wb.sheetnames:
            return alias
    # Fuzzy fallback — case-insensitive contains
    for name in wb.sheetnames:
        if sheet_type.replace("_", " ") in name.lower():
            return name
    # Single-sheet workbook: the only sheet IS the CPL (some distributors submit
    # a bare price list with no RIP/COMBO/TERMS tabs, sometimes named "Sheet1").
    if sheet_type == "cpl" and len(wb.sheetnames) == 1:
        return wb.sheetnames[0]
    return None


class NJABCParser:
    """
    Base parser for NJ ABC eCPL Excel files.

    Subclass or configure via WholesalerConfig to handle per-wholesaler differences.
    Most wholesalers need ZERO code — just a config dict.
    """

    def __init__(self, config: dict):
        """
        config keys:
            slug: str               — short identifier (e.g., "allied")
            name: str               — display name
            header_row_hint: int    — 1-indexed row where headers likely are (default: 5)
            max_header_search: int  — how many rows to search for headers (default: 10)
            discount_tiers: int     — number of discount tiers in CPL (default: 3)
            rip_tiers: int          — number of RIP tiers (default: 2)
            product_type_map: dict  — normalize product type labels
            skip_sheets: list       — sheets to skip entirely
            cpl_header_map_overrides: dict — extra header mappings
            rip_header_map_overrides: dict
        """
        self.config = config
        self.slug = config["slug"]
        self.name = config["name"]
        self.header_row_hint = config.get("header_row_hint", 5)
        self.max_header_search = config.get("max_header_search", 10)
        self.discount_tiers = config.get("discount_tiers", 3)
        self.rip_tiers = config.get("rip_tiers", 2)
        self.product_type_map = config.get("product_type_map", {})
        self.skip_sheets = config.get("skip_sheets", [])

        # Build effective header maps with overrides
        self.cpl_header_map = {**CPL_HEADER_MAP, **config.get("cpl_header_map_overrides", {})}
        self.rip_header_map = {**RIP_HEADER_MAP, **config.get("rip_header_map_overrides", {})}
        self.combo_header_map = {**COMBO_HEADER_MAP, **config.get("combo_header_map_overrides", {})}
        self.beer_mm_header_map = {**BEER_MM_HEADER_MAP, **config.get("beer_mm_header_map_overrides", {})}

        # Optional post-process hook. Called after every sheet has parsed with
        # the (parser, sheets_dict) so a wholesaler-specific config can
        # cross-join sheets — e.g. enrich COMBO using CPL when the source file
        # records component item codes instead of product names. The hook may
        # mutate the dict in place and/or return a replacement dict.
        self.post_process = config.get("post_process")

    def parse_file(self, filepath: Path) -> dict[str, pd.DataFrame]:
        """
        Parse an NJ ABC Excel file into DataFrames.

        Returns dict with keys: "cpl", "rip", "combo", "beer_mm"
        (only present if sheet exists and has data).
        """
        import shutil, tempfile
        # Copy to temp file to avoid OneDrive lock issues
        tmp = Path(tempfile.mkdtemp()) / filepath.name
        shutil.copy2(filepath, tmp)

        try:
            wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)
            result = {}

            for sheet_type, parser_method in [
                ("cpl", self._parse_cpl),
                ("rip", self._parse_rip),
                ("combo", self._parse_combo),
                ("beer_mm", self._parse_beer_mm),
            ]:
                if sheet_type in self.skip_sheets:
                    continue
                sheet_name = _find_sheet(wb, sheet_type)
                if sheet_name is None:
                    logger.debug(f"[{self.slug}] Sheet '{sheet_type}' not found, skipping")
                    continue
                ws = wb[sheet_name]
                df = parser_method(ws)
                if df is not None and len(df) > 0:
                    result[sheet_type] = df
                    logger.info(f"[{self.slug}] {sheet_type}: {len(df)} rows parsed")
                else:
                    logger.debug(f"[{self.slug}] {sheet_type}: no data rows")

            # Extra CPL-format sheets (config: extra_cpl_sheets). Some wholesalers
            # publish a block of priced products on a sheet whose NAME isn't "CPL"
            # — e.g. Fedway lists ~270 Value-Added Packs (gift packs) on the
            # "TERMS and CONDITIONS" sheet, fully priced, in the SAME column layout
            # as the CPL sheet but with NO header row of their own. Those rows were
            # silently dropped, so the product (and its distributor) never showed.
            # Parse them with the MAIN CPL sheet's positional column map and union
            # them into the CPL frame. Done before wb.close() so the workbook is
            # still open; before post_process so hooks (combo repair) see them.
            extra_sheets = self.config.get("extra_cpl_sheets") or []
            if extra_sheets and "cpl" in result:
                main_name = _find_sheet(wb, "cpl")
                hr = self._find_header_row(wb[main_name], self.cpl_header_map, min_matches=5) if main_name else None
                if hr:
                    col_map = self._build_column_mapping(wb[main_name], hr, self.cpl_header_map)
                    if col_map and self.config.get("cpl_dist_item_after_headers"):
                        col_map.setdefault(max(col_map.keys()) + 1, "dist_item_no")
                    extra_frames = []
                    for want in extra_sheets:
                        real = (want if want in wb.sheetnames
                                else next((n for n in wb.sheetnames if want.lower() in n.lower()), None))
                        if real is None:
                            continue
                        edf = self._extract_cpl_like(wb[real], col_map)
                        if edf is not None and len(edf) > 0:
                            logger.info(f"[{self.slug}] extra CPL sheet '{real}': {len(edf)} rows parsed")
                            extra_frames.append(edf)
                    if extra_frames:
                        result["cpl"] = pd.concat([result["cpl"], *extra_frames], ignore_index=True)

            wb.close()
            if self.post_process and result:
                replaced = self.post_process(self, result)
                if replaced is not None:
                    result = replaced
            return result
        finally:
            try:
                tmp.unlink()
                tmp.parent.rmdir()
            except Exception:
                pass

    def _find_header_row(self, ws, header_map: dict, min_matches: int = 3) -> Optional[int]:
        """
        Scan the first N rows to find the header row.
        Returns 1-indexed row number, or None.
        """
        for row_idx in range(1, self.max_header_search + 1):
            row_values = []
            for cell in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
                row_values = list(cell)
                break

            matches = 0
            for val in row_values:
                if val is not None and _match_header(str(val), header_map):
                    matches += 1

            if matches >= min_matches:
                return row_idx

        return None

    def _build_column_mapping(self, ws, header_row: int, header_map: dict) -> dict[int, str]:
        """
        Build {column_index: canonical_name} mapping from the header row.
        Returns dict mapping 0-based column index to canonical column name.
        """
        col_map = {}
        for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
            for col_idx, val in enumerate(row):
                if val is None:
                    continue
                canonical = _match_header(str(val), header_map)
                if canonical and canonical not in col_map.values():
                    col_map[col_idx] = canonical
            break
        return col_map

    def _extract_rows(self, ws, header_row: int, col_map: dict[int, str],
                      canonical_columns: list[str]) -> pd.DataFrame:
        """Extract data rows starting after header_row using the column mapping."""
        records = []
        max_col_idx = max(col_map.keys()) if col_map else 0

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Skip completely empty rows
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue

            record = {}
            for col_idx, canonical in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                record[canonical] = val
            records.append(record)

        if not records:
            return None

        df = pd.DataFrame(records)

        # Ensure all canonical columns exist
        for col in canonical_columns:
            if col not in df.columns:
                df[col] = None

        # Reorder to canonical order, keeping only canonical columns
        df = df[[c for c in canonical_columns if c in df.columns]]
        return df

    def _parse_cpl(self, ws) -> Optional[pd.DataFrame]:
        """Parse CPL sheet."""
        header_row = self._find_header_row(ws, self.cpl_header_map, min_matches=5)
        if header_row is None:
            # Some files ship the standard ABC column ORDER but omit the header
            # row (e.g. Banville's July export). When the config opts in, map
            # columns positionally to the canonical order (CPL_COLUMNS) and treat
            # header_row_hint as the blank/absent header line, so data extraction
            # starts on the row after it.
            if self.config.get("cpl_assume_standard_order"):
                # Find the first DATA row (column A is a UPC-like long digit
                # string); the absent header sits just above it. Column indices
                # are 0-based to match _build_column_mapping / _extract_rows.
                data_row = None
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                    c0 = row[0] if row else None
                    if c0 is not None and re.fullmatch(r"0*\d{8,}", str(c0).strip()):
                        data_row = r_idx
                        break
                header_row = (data_row - 1) if data_row else self.header_row_hint
                col_map = {i: CPL_COLUMNS[i] for i in range(min(len(CPL_COLUMNS), ws.max_column))}
                logger.info(f"[{self.slug}] CPL: no header row — assuming standard "
                            f"ABC column order, data from row {header_row + 1}")
                df = self._extract_rows(ws, header_row, col_map, CPL_COLUMNS)
                if df is None:
                    return None
                return self._clean_cpl(df)
            logger.warning(f"[{self.slug}] CPL: could not find header row")
            return None

        col_map = self._build_column_mapping(ws, header_row, self.cpl_header_map)
        logger.debug(f"[{self.slug}] CPL header at row {header_row}, mapped {len(col_map)} columns")

        # Distributor item number: Fedway appends an UNNAMED column right of the
        # last labelled CPL header (column Z) carrying its internal item number,
        # one per product row (same key as the RIP sheet's dist_item_no). The
        # header cell is blank there, so map the first unmapped column after the
        # rightmost mapped one. Gated by config so other wholesalers, whose
        # trailing column may be empty or something else, are untouched.
        if col_map and self.config.get("cpl_dist_item_after_headers"):
            extra_idx = max(col_map.keys()) + 1
            if extra_idx not in col_map:
                col_map[extra_idx] = "dist_item_no"

        df = self._extract_rows(ws, header_row, col_map, CPL_COLUMNS)
        if df is None:
            return None

        df = self._clean_cpl(df)
        return df

    def _extract_cpl_like(self, ws, col_map: dict[int, str]) -> Optional[pd.DataFrame]:
        """Extract product rows from a sheet that shares the CPL column LAYOUT but
        has no header row of its own (e.g. Fedway's VAP block on the TERMS sheet).
        Uses the supplied CPL column map (derived from the real CPL sheet) and
        keeps only rows whose UPC column holds a real >=8-digit barcode, so the
        sheet's terms/notes text rows are skipped. Cleaned via _clean_cpl (which
        also drops nameless rows), so the output is identical in shape to a normal
        CPL parse."""
        upc_idx = next((i for i, c in col_map.items() if c == "upc"), 0)
        records = []
        for row in ws.iter_rows(values_only=True):
            raw = row[upc_idx] if upc_idx < len(row) else None
            upc = _to_upc_string(raw)
            if not (upc and upc.isdigit() and len(upc) >= 8):
                continue
            rec = {}
            for ci, canonical in col_map.items():
                v = row[ci] if ci < len(row) else None
                # Number-formatted code cells (rip_code, brand_reg_no, unit_qty)
                # arrive as int/float here; a numeric column that also has None
                # is promoted by pandas to float64, so an int 10168 surfaces as
                # "10168.0" and breaks the RIP join. Stringify integer values so
                # the assembled column stays object dtype (matching the main CPL
                # sheet, whose code cells are text). True decimal prices like
                # 294.54 stay numeric; _clean_cpl re-coerces price columns anyway.
                if isinstance(v, bool):
                    pass
                elif isinstance(v, int):
                    v = str(v)
                elif isinstance(v, float) and v.is_integer():
                    v = str(int(v))
                rec[canonical] = v
            records.append(rec)
        if not records:
            return None
        df = pd.DataFrame(records)
        for col in CPL_COLUMNS:
            if col not in df.columns:
                df[col] = None
        df = df[[c for c in CPL_COLUMNS if c in df.columns]]
        return self._clean_cpl(df)

    def _parse_rip(self, ws) -> Optional[pd.DataFrame]:
        """Parse RIP sheet."""
        header_row = self._find_header_row(ws, self.rip_header_map, min_matches=3)
        if header_row is None:
            logger.warning(f"[{self.slug}] RIP: could not find header row")
            return None

        col_map = self._build_column_mapping(ws, header_row, self.rip_header_map)
        # Distributor item number: Fedway appends an UNNAMED column right of
        # the last template column carrying its internal item number (the key
        # its half-case comments reference, and the only product key on rows
        # filed with UPC=0). The header row has no text there, so map the
        # first unmapped column after the rightmost mapped one.
        if col_map:
            extra_idx = max(col_map.keys()) + 1
            if extra_idx not in col_map:
                col_map[extra_idx] = "dist_item_no"
        df = self._extract_rows(ws, header_row, col_map, RIP_COLUMNS)
        if df is None:
            return None

        df = self._clean_rip(df)
        return df

    def _parse_combo(self, ws) -> Optional[pd.DataFrame]:
        """Parse COMBO sheet."""
        header_row = self._find_header_row(ws, self.combo_header_map, min_matches=3)
        if header_row is None:
            logger.warning(f"[{self.slug}] COMBO: could not find header row")
            return None

        col_map = self._build_column_mapping(ws, header_row, self.combo_header_map)
        df = self._extract_rows(ws, header_row, col_map, COMBO_COLUMNS)
        if df is None:
            return None

        df = self._clean_combo(df)
        return df

    def _parse_beer_mm(self, ws) -> Optional[pd.DataFrame]:
        """Parse BEER MIX and MATCH sheet."""
        header_row = self._find_header_row(ws, self.beer_mm_header_map, min_matches=3)
        if header_row is None:
            logger.debug(f"[{self.slug}] BEER_MM: could not find header row (often empty)")
            return None

        col_map = self._build_column_mapping(ws, header_row, self.beer_mm_header_map)
        df = self._extract_rows(ws, header_row, col_map, BEER_MM_COLUMNS)
        if df is None:
            return None

        df = self._clean_beer_mm(df)
        return df

    # -------------------------------------------------------------------
    # Cleaning methods — override in subclass if needed
    # -------------------------------------------------------------------

    def _clean_cpl(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and normalize CPL data."""
        # UPC: normalize to clean digit string (handles Excel float coercion)
        df["upc"] = df["upc"].apply(_to_upc_string)

        # Dates
        for col in ["from_date", "to_date"]:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        # Numeric price columns — strip currency formatting first
        # (e.g., Jersey Beverage submits best_unit_price as "$29.00")
        price_cols = [
            "frontline_case_price", "frontline_unit_price",
            "best_case_price", "best_unit_price", "split_case_surcharge",
        ]
        for col in price_cols:
            df[col] = pd.to_numeric(
                df[col].apply(
                    lambda x: re.sub(r"[$,\s]", "", x) if isinstance(x, str) else x
                ),
                errors="coerce",
            )

        # Discount tiers — qty as string (may contain "5 Cases"), amount as numeric
        for i in range(1, 6):
            qty_col = f"discount_{i}_qty"
            amt_col = f"discount_{i}_amt"
            if qty_col in df.columns:
                df[qty_col] = df[qty_col].apply(_to_str_stripped)
            if amt_col in df.columns:
                df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce")

        # Normalize product type
        if self.product_type_map:
            df["product_type"] = df["product_type"].apply(
                lambda x: self.product_type_map.get(
                    str(x).strip().upper(), str(x).strip() if pd.notna(x) else None
                )
            )

        # String columns
        for col in ["product_name", "unit_type", "brand_reg_no",
                     "rip_code", "combo_code", "closeout_permit", "vintage"]:
            if col in df.columns:
                df[col] = df[col].apply(_to_str_stripped)
        if "unit_volume" in df.columns:
            df["unit_volume"] = df["unit_volume"].apply(_normalize_volume)

        df["unit_qty"] = df["unit_qty"].apply(_to_str_stripped)
        df["abv_proof"] = df["abv_proof"].apply(_to_str_stripped)

        # Distributor item number (Fedway column Z): Excel floats ("10040.0") ->
        # clean digit strings; anything non-numeric -> None. Mirrors _clean_rip.
        if "dist_item_no" in df.columns:
            df["dist_item_no"] = df["dist_item_no"].apply(
                lambda v: str(v).split(".")[0].strip()
                if v is not None and str(v).split(".")[0].strip().isdigit()
                else None
            )

        # Drop rows where product_name is empty (junk rows)
        df = df[df["product_name"].notna() & (df["product_name"] != "")]

        return df.reset_index(drop=True)

    def _clean_rip(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean RIP data."""
        df["rip_code"] = df["rip_code"].apply(_to_str_stripped)
        df["upc"] = df["upc"].apply(_to_upc_string)
        df["brand_reg_no"] = df["brand_reg_no"].apply(_to_str_stripped)
        df["rip_description"] = df["rip_description"].apply(_to_str_stripped)
        df["comments"] = df["comments"].apply(_to_str_stripped)
        if "dist_item_no" in df.columns:
            # Excel floats ("130360.0") -> clean digit strings; junk -> None
            df["dist_item_no"] = df["dist_item_no"].apply(
                lambda v: str(v).split(".")[0].strip()
                if v is not None and str(v).split(".")[0].strip().isdigit()
                else None
            )

        for col in ["from_date", "to_date"]:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        for i in range(1, 5):
            unit_col = f"rip_unit_{i}"
            qty_col = f"rip_qty_{i}"
            amt_col = f"rip_amt_{i}"
            if unit_col in df.columns:
                df[unit_col] = df[unit_col].apply(_to_str_stripped)
            if qty_col in df.columns:
                df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce")
            if amt_col in df.columns:
                df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce")

        df = df[df["rip_description"].notna() & (df["rip_description"] != "")]
        return df.reset_index(drop=True)

    def _clean_combo(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean COMBO data."""
        df["combo_code"] = df["combo_code"].apply(_to_str_stripped)
        df["upc"] = df["upc"].apply(_to_upc_string)
        df["product_name"] = df["product_name"].apply(_to_str_stripped)
        df["brand_reg_no"] = df["brand_reg_no"].apply(_to_str_stripped)
        df["comments"] = df["comments"].apply(_to_str_stripped)
        df["qty_per_pack"] = df["qty_per_pack"].apply(_to_str_stripped)

        for col in ["from_date", "to_date"]:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        for col in ["combo_pack_price", "frontline_price_each",
                     "combo_price_each", "total_savings"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        df = df[df["product_name"].notna() & (df["product_name"] != "")]
        return df.reset_index(drop=True)

    def _clean_beer_mm(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean BEER MIX AND MATCH data."""
        for col in ["beer_mm_code", "description", "brand_reg_no", "rolling_keg"]:
            if col in df.columns:
                df[col] = df[col].apply(_to_str_stripped)
        if "upc" in df.columns:
            df["upc"] = df["upc"].apply(_to_upc_string)

        for col in ["from_date", "to_date"]:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        for col in ["frontline_case_keg_price", "min_qty", "discount_pct",
                     "price_each", "per_case_keg_discount"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        df = df[df["description"].notna() & (df["description"] != "")]
        return df.reset_index(drop=True)

    # -------------------------------------------------------------------
    # Metadata extraction
    # -------------------------------------------------------------------

    def extract_metadata(self, filepath: Path) -> dict:
        """Extract wholesaler name, license, submission date from the file header."""
        import shutil, tempfile
        tmp = Path(tempfile.mkdtemp()) / filepath.name
        shutil.copy2(filepath, tmp)

        try:
            wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)
            sheet_name = _find_sheet(wb, "cpl")
            if sheet_name is None:
                wb.close()
                return {}

            ws = wb[sheet_name]
            meta = {}
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                vals = [v for v in row if v is not None]
                if len(vals) >= 2:
                    key = str(vals[0]).strip().upper()
                    if "WHOLESALER" in key:
                        meta["wholesaler_name"] = str(vals[1]).strip()
                    elif "LICENSE" in key or "PERMIT" in key:
                        meta["license_no"] = str(vals[1]).strip()
                    elif "SUBMISSION" in key:
                        meta["submission_date"] = vals[1]

            wb.close()
            return meta
        finally:
            try:
                tmp.unlink()
                tmp.parent.rmdir()
            except Exception:
                pass


# -----------------------------------------------------------------------
# Utility functions
# -----------------------------------------------------------------------

def _to_str_stripped(val) -> Optional[str]:
    """Convert value to stripped string, return None for empty/whitespace."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.lower() == "nan" or s.lower() == "none":
        return None
    return s


def _to_upc_string(val) -> Optional[str]:
    """Normalize a UPC value to a clean digit string.

    Excel often coerces long numeric UPCs to floats, so str() produces
    '812066021598.0'. Drop a trailing '.0' (and a few other noisy patterns)
    so the UPC stays a clean digit string and search/joins work.
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        try:
            return str(int(val))
        except (OverflowError, ValueError):
            return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    if "." in s and s.replace(".", "", 1).isdigit():
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except (TypeError, ValueError):
            pass
    return s


_VOLUME_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*([A-Za-z]+)\s*$")


def _normalize_volume(val) -> Optional[str]:
    """Normalize a unit_volume string to a canonical form.

    Different wholesalers write the same size differently:
      "750 ML", "750ML", "750ml"  -> "750ML"
      "1.75 L", "1.75LT", "1.75L" -> "1.75L"
      "12 OZ"                     -> "12OZ"

    Returns the canonical string, or the trimmed original if it doesn't fit
    the "<number> <unit>" pattern.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    m = _VOLUME_RE.match(s.replace("\xa0", " "))
    if not m:
        return s
    num = m.group(1)
    unit = m.group(2).upper()
    if unit in ("LT", "LTR", "LITER", "LITERS"):
        unit = "L"
    elif unit in ("ML", "MILLILITER", "MILLILITERS"):
        unit = "ML"
    elif unit in ("OZ", "OUNCE", "OUNCES"):
        unit = "OZ"
    # Drop trailing ".0" on whole-number decimals
    if "." in num:
        try:
            f = float(num)
            if f.is_integer():
                num = str(int(f))
        except ValueError:
            pass
    return f"{num}{unit}"
