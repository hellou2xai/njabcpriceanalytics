"""
Wholesaler registry — maps file patterns to parser configs.

To add a new wholesaler:
  1. Create a config file in wholesalers/ with a CONFIG dict
  2. Add an import + entry to REGISTRY below
  3. Run ETL — the base parser handles the rest

No code changes needed unless the new wholesaler has a truly unique layout.
"""

import re
from pathlib import Path
from fnmatch import fnmatch

from nj_abc_parser.wholesalers.allied import CONFIG as ALLIED
from nj_abc_parser.wholesalers.fedway import CONFIG as FEDWAY
from nj_abc_parser.wholesalers.opici import CONFIG as OPICI
from nj_abc_parser.wholesalers.peerless import CONFIG as PEERLESS
from nj_abc_parser.wholesalers.high_grade import CONFIG as HIGH_GRADE
from nj_abc_parser.wholesalers.kramer import CONFIG as KRAMER
from nj_abc_parser.wholesalers.shore_point import CONFIG as SHORE_POINT
from nj_abc_parser.wholesalers.jersey_beverage import CONFIG as JERSEY_BEVERAGE
from nj_abc_parser.wholesalers.other_brothers import CONFIG as OTHER_BROTHERS
from nj_abc_parser.wholesalers.winebow import CONFIG as WINEBOW
from nj_abc_parser.wholesalers.gallo import CONFIG as GALLO
from nj_abc_parser.wholesalers.regal_wine import CONFIG as REGAL_WINE
from nj_abc_parser.wholesalers.wine_enterprises import CONFIG as WINE_ENTERPRISES
from nj_abc_parser.wholesalers.trivin import CONFIG as TRIVIN
from nj_abc_parser.wholesalers.monsieur import CONFIG as MONSIEUR
from nj_abc_parser.wholesalers.wilson_daniels import CONFIG as WILSON_DANIELS
from nj_abc_parser.wholesalers.banville import CONFIG as BANVILLE
from nj_abc_parser.wholesalers.david_bowler import CONFIG as DAVID_BOWLER

REGISTRY = [ALLIED, FEDWAY, OPICI, PEERLESS, HIGH_GRADE, KRAMER, SHORE_POINT,
            JERSEY_BEVERAGE, OTHER_BROTHERS, WINEBOW, GALLO, REGAL_WINE,
            WINE_ENTERPRISES, TRIVIN, MONSIEUR, WILSON_DANIELS, BANVILLE,
            DAVID_BOWLER]


def list_wholesalers() -> list[dict]:
    """Return all registered wholesaler configs."""
    return REGISTRY


def get_wholesaler_config(slug: str) -> dict | None:
    """Look up a wholesaler config by slug."""
    for cfg in REGISTRY:
        if cfg["slug"] == slug:
            return cfg
    return None


def detect_wholesaler(filepath: Path) -> dict | None:
    """
    Auto-detect which wholesaler a file belongs to based on filename patterns.
    Returns the matching config, or None.
    """
    name = filepath.name
    for cfg in REGISTRY:
        patterns = cfg.get("file_pattern", [])
        if isinstance(patterns, str):
            patterns = [patterns]
        for pattern in patterns:
            if fnmatch(name, pattern):
                return cfg
    return None


def edition_year_from_submission(filepath: Path) -> int | None:
    """Fallback when a filename gives a MONTH but no 4-digit YEAR (e.g. Wine
    Enterprises' "Wine Enterprises_June.xlsx"): read the workbook's SUBMISSION
    DATE and return its year. Cheap header-only scan; failures return None."""
    import datetime
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        for sn in wb.sheetnames[:3]:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                cells = list(row or [])
                for i, cell in enumerate(cells):
                    if isinstance(cell, str) and "SUBMISSION" in cell.upper():
                        for v in cells[i + 1:]:
                            if isinstance(v, datetime.datetime):
                                return v.year
                            if isinstance(v, str):
                                m = re.search(r"\b(20\d{2})\b", v)
                                if m:
                                    return int(m.group(1))
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def parse_edition_from_filename(filepath: Path) -> dict:
    """
    Extract year and month from the filename.

    Handles patterns like:
      - "Allied Beverage Group April CPL 2026.xlsx"
      - "Fedway Associates 2026-04 CPL.xlsx"
      - "2026 April Price File.xlsx"
      - "Peerless Beverage Co. April 2026 CPL.xlsx"
      - "ECPL Randolph April 2026.xlsx"
      - "eCPL May 2026.xlsx"
    """
    name = filepath.stem  # without extension

    MONTH_MAP = {
        "january": "01", "february": "02", "march": "03", "april": "04",
        "may": "05", "june": "06", "july": "07", "august": "08",
        "september": "09", "october": "10", "november": "11", "december": "12",
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "jun": "06", "jul": "07", "aug": "08",
        "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    }

    year = None
    month = None

    # Try YYYY-MM pattern first (Fedway style)
    m = re.search(r"(\d{4})-(\d{2})", name)
    if m:
        year = m.group(1)
        month = m.group(2)
    else:
        # Extract 4-digit year, or 2-digit year (e.g., "26" → "2026")
        m_year = re.search(r"(20\d{2})", name)
        if m_year:
            year = m_year.group(1)
        else:
            # Try 2-digit year at word boundary (e.g., "June 26.xlsx")
            m_year2 = re.search(r"\b(\d{2})\b", name)
            if m_year2:
                short = int(m_year2.group(1))
                if 20 <= short <= 39:  # plausible range 2020-2039
                    year = str(2000 + short)

        # Extract month name
        name_lower = name.lower()
        for month_name, month_num in MONTH_MAP.items():
            if month_name in name_lower:
                month = month_num
                break

        # Numeric edition tokens when there's no month NAME: "MM-YY" (Royal
        # "6-26") or "MMYY" (Monsieur "0626"). Guarded so a real year (2026) or
        # a day-month pair isn't misread: first part must be a month (1-12) and
        # the 2-digit year in the plausible 2020-2039 range. The filename month
        # wins over the submission date (which can be the PRIOR month).
        if month is None:
            m_dash = re.search(r"(?<!\d)(\d{1,2})-(\d{2})(?!\d)", name)
            m_mmyy = re.search(r"(?<!\d)(\d{2})(\d{2})(?!\d)", name)
            if m_dash and 1 <= int(m_dash.group(1)) <= 12 and 20 <= int(m_dash.group(2)) <= 39:
                month = f"{int(m_dash.group(1)):02d}"
                year = year or str(2000 + int(m_dash.group(2)))
            elif m_mmyy and 1 <= int(m_mmyy.group(1)) <= 12 and 20 <= int(m_mmyy.group(2)) <= 39:
                month = m_mmyy.group(1)
                year = year or str(2000 + int(m_mmyy.group(2)))

    return {
        "year": int(year) if year else None,
        "month": int(month) if month else None,
        "edition": f"{year}-{month}" if year and month else None,
    }
