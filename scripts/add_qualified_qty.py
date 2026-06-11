"""Qualified Quantity enrichment (user spec 2026-06-11).

Some RIP descriptions declare that a pack counts as a FRACTION of a RIP case,
e.g. "DON Q RUM 3PK 1.75L= 1/2 CASE", "(6PK=1/2CS)", "VAPS= 1/2 CASE",
"COPPOLA DIRECTOR'S CUT 6PK=1/2 CASE ...". For those products the RIP sheet's
"case" tiers are counted in a bigger unit than the CPL case, so a buyer needs
MORE CPL cases/bottles to earn one RIP case. This script:

  1. scans the RIP sheet for fraction-of-case descriptions (1/2, 1/3, 1/4, 1/6
     CASE or CS, or HALF/QUARTER CASE),
  2. matches each flagged RIP ROW to CPL rows by UPC *and* RIP code (never by
     product name — descriptions abbreviate, e.g. DC = Director's Cut),
  3. when the description names a pack ("3PK", "6PK"), only CPL rows with that
     unit quantity are flagged (full-case sizes sharing the RIP stay untouched),
  4. writes a new "Qualified Quantity" column on the CPL sheet: the number of
     BOTTLES that make up ONE RIP case for that row
     (unit_qty / fraction — pack 12 at 1/2 case -> 24 bottles = 2 CPL cases).

Nothing else in the workbook is modified. Idempotent: reruns overwrite the
same column.

Usage:
    python scripts/add_qualified_qty.py            # both Enhancement files
    python scripts/add_qualified_qty.py <file.xlsx> [...]
"""
from __future__ import annotations

import re
import sys
from fractions import Fraction

import openpyxl

DEFAULT_FILES = [
    r"Data\Enhancement\Allied Beverage Group June CPL 2026.xlsx",
    r"Data\Enhancement\Fedway Associates 2026-06 CPL.xlsx",
]

QUAL_HEADER = "Qualified Quantity"

# "1/2 CASE", "1/2CS", "= 1/4 CS", and the spelled-out variants.
FRACTION_RX = re.compile(r"\b1\s*/\s*([2346])\s*(?:CASE|CS)\b", re.I)
WORD_RX = re.compile(r"\b(HALF|QUARTER)\s+(?:CASE|CS)\b", re.I)
# Pack hint like "3PK", "6 PK", "6-PK" anywhere in the description.
PACK_RX = re.compile(r"\b(\d{1,2})\s*-?\s*PK\b", re.I)


def norm_upc(v) -> str:
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s.lstrip("0")


def norm_code(v) -> str:
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper()


def parse_fraction(desc: str) -> Fraction | None:
    m = FRACTION_RX.search(desc)
    if m:
        return Fraction(1, int(m.group(1)))
    w = WORD_RX.search(desc)
    if w:
        return Fraction(1, 2) if w.group(1).upper() == "HALF" else Fraction(1, 4)
    return None


def header_index(ws, header_row: int, needle: str) -> int | None:
    for cell in ws[header_row]:
        v = re.sub(r"\s+", " ", str(cell.value or "")).strip().lower()
        if needle in v:
            return cell.column  # 1-based
    return None


def find_header_row(ws) -> int | None:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
        for cell in row:
            if "upc" in str(cell.value or "").lower():
                return i
    return None


def process(path: str) -> None:
    print("=" * 70)
    print(path)
    wb = openpyxl.load_workbook(path)
    cpl, rip = wb["CPL"], wb["RIP"]
    cpl_hr, rip_hr = find_header_row(cpl), find_header_row(rip)
    if not cpl_hr or not rip_hr:
        print("  !! header row not found, skipping file")
        return

    # --- column positions ---
    c_upc = header_index(cpl, cpl_hr, "upc code")
    c_qty = header_index(cpl, cpl_hr, "unit quantity")
    c_rip = header_index(cpl, cpl_hr, "rip code")
    r_code = header_index(rip, rip_hr, "rip code")
    r_upc = header_index(rip, rip_hr, "upc code")
    r_desc = header_index(rip, rip_hr, "rip description")
    if not all([c_upc, c_qty, c_rip, r_code, r_upc, r_desc]):
        print("  !! required columns missing, skipping file")
        return

    # --- 1) flagged RIP rows: (rip_code, upc) -> (fraction, pack_hint, desc) ---
    flagged: dict[tuple[str, str], tuple[Fraction, int | None, str]] = {}
    for row in rip.iter_rows(min_row=rip_hr + 1, values_only=True):
        desc = str(row[r_desc - 1] or "")
        frac = parse_fraction(desc)
        if not frac:
            continue
        upc = norm_upc(row[r_upc - 1])
        code = norm_code(row[r_code - 1])
        if not upc or not code:
            continue  # placeholder UPC rows can't be matched safely
        pm = PACK_RX.search(desc)
        pack_hint = int(pm.group(1)) if pm else None
        prev = flagged.get((code, upc))
        # On conflicting descriptions keep the SMALLER fraction (more bottles
        # needed), so the annotation never overstates eligibility.
        if prev is None or frac < prev[0]:
            flagged[(code, upc)] = (frac, pack_hint, desc.strip())
    print(f"  RIP rows with a fraction-case qualifier: {len(flagged)} distinct (code, upc) pairs")

    # --- 2) walk CPL rows, match by UPC + RIP code, write the column ---
    qual_col = header_index(cpl, cpl_hr, QUAL_HEADER.lower())
    if not qual_col:
        qual_col = cpl.max_column + 1
    cpl.cell(row=cpl_hr, column=qual_col, value=QUAL_HEADER)

    written = skipped_pack = 0
    for r in range(cpl_hr + 1, cpl.max_row + 1):
        upc = norm_upc(cpl.cell(row=r, column=c_upc).value)
        code = norm_code(cpl.cell(row=r, column=c_rip).value)
        if not upc or not code:
            continue
        hit = flagged.get((code, upc))
        if not hit:
            continue
        frac, pack_hint, _desc = hit
        try:
            unit_qty = float(cpl.cell(row=r, column=c_qty).value)
        except (TypeError, ValueError):
            continue
        if unit_qty <= 0:
            continue
        if pack_hint is not None and round(unit_qty) != pack_hint:
            skipped_pack += 1   # e.g. the 12-pack sharing a 6PK=1/2CASE RIP
            continue
        bottles = unit_qty / float(frac)   # 12 @ 1/2 -> 24 bottles (= 2 cases)
        cpl.cell(row=r, column=qual_col,
                 value=int(bottles) if float(bottles).is_integer() else round(bottles, 2))
        written += 1

    print(f"  CPL rows annotated: {written}"
          f"  (pack-hint mismatches skipped: {skipped_pack})")
    wb.save(path)
    print("  saved.")


if __name__ == "__main__":
    files = sys.argv[1:] or DEFAULT_FILES
    for f in files:
        process(f)
