"""RIP membership mismatch analysis (user request 2026-06-11).

For every Case Mix RIP code in the Allied + Fedway June 2026 workbooks,
compare the RIP sheet's distinct UPC list against what the app's
UPC + rip_code matching finds on the CPL, and classify every UPC:

  matched_code        - >=1 CPL row for this UPC carries this rip_code
  kept_single_listing - in CPL with ONE product name but a different
                        rip_code (sheet presence keeps it, stacked RIPs)
  dropped_multi_sku   - in CPL with SEVERAL product names on the barcode
                        and NO row referencing this code (strict rule
                        drops these from the member list)
  not_in_cpl          - UPC absent from the CPL entirely (UI shows an
                        "unavailable" stub)

Output: Data/Enhancement/RIP membership mismatch analysis 2026-06.xlsx
  - Summary sheet: one row per (wholesaler, rip_code) with counts and the
    mismatch (sheet count minus what the member list shows as buyable),
    sorted worst-first.
  - One detail sheet per wholesaler: one row per (rip_code, upc) that is
    NOT matched_code, with the CPL names + rip_codes the barcode carries.

Pure analysis: reads the workbooks, writes the analysis file, changes
nothing else.
"""
from __future__ import annotations

import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font

FILES = {
    "allied": r"Data\Enhancement\Allied Beverage Group June CPL 2026.xlsx",
    "fedway": r"Data\Enhancement\Fedway Associates 2026-06 CPL.xlsx",
}
OUT = r"Data\Enhancement\RIP membership mismatch analysis 2026-06.xlsx"


def norm_upc(v) -> str:
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s.lstrip("0")


def norm_code(v) -> str:
    s = str(v if v is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper()


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if any("upc" in str(v or "").lower() for v in row):
            return i, [re.sub(r"\s+", " ", str(v or "")).strip().lower() for v in row]
    return None, []


def col(headers, needle):
    for i, h in enumerate(headers):
        if needle in h:
            return i
    return None


def analyze(name: str, path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cpl, rip = wb["CPL"], wb["RIP"]

    chr_, ch = find_header_row(cpl)
    rhr, rh = find_header_row(rip)
    c_upc, c_name, c_rip = col(ch, "upc code"), col(ch, "product name"), col(ch, "rip code")
    r_code, r_upc, r_desc = col(rh, "rip code"), col(rh, "upc code"), col(rh, "rip description")

    # CPL index: upc -> list of (product_name, rip_code)
    cpl_idx: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for row in cpl.iter_rows(min_row=chr_ + 1, values_only=True):
        u = norm_upc(row[c_upc])
        if not u:
            continue
        cpl_idx[u].append((str(row[c_name] or "").strip(), norm_code(row[c_rip])))

    # RIP sheet: code -> set of upcs (+ first description)
    sheet: dict[str, set[str]] = defaultdict(set)
    desc: dict[str, str] = {}
    for row in rip.iter_rows(min_row=rhr + 1, values_only=True):
        code = norm_code(row[r_code])
        if not code or code in ("0", "NONE", "NAN"):
            continue
        u = norm_upc(row[r_upc])
        if u:
            sheet[code].add(u)
        if code not in desc and row[r_desc]:
            desc[code] = str(row[r_desc]).strip()
    wb.close()

    summary, detail = [], []
    for code, upcs in sheet.items():
        counts = dict(matched_code=0, kept_single_listing=0, dropped_multi_sku=0, not_in_cpl=0)
        for u in sorted(upcs):
            rows = cpl_idx.get(u)
            if not rows:
                status = "not_in_cpl"
                names, codes = "", ""
            else:
                names_set = {n.upper() for n, _ in rows if n}
                codes_set = {c for _, c in rows if c and c not in ("0", "NONE", "NAN")}
                if any(c == code for _, c in rows):
                    status = "matched_code"
                elif len(names_set) <= 1:
                    status = "kept_single_listing"
                else:
                    status = "dropped_multi_sku"
                names = " | ".join(sorted(names_set))[:120]
                codes = ", ".join(sorted(codes_set))[:80]
            counts[status] += 1
            if status != "matched_code":
                detail.append([code, desc.get(code, ""), u, status, names, codes])
        shown = counts["matched_code"] + counts["kept_single_listing"]
        summary.append([
            name, code, desc.get(code, "")[:90], len(upcs),
            counts["matched_code"], counts["kept_single_listing"],
            counts["dropped_multi_sku"], counts["not_in_cpl"],
            len(upcs) - shown,
        ])
    return summary, detail


def main():
    out = openpyxl.Workbook()
    s = out.active
    s.title = "Summary"
    s.append(["wholesaler", "rip_code", "rip_description (first)", "sheet_upc_count",
              "matched_code", "kept_single_listing", "dropped_multi_sku",
              "not_in_cpl", "mismatch (sheet - shown)"])
    all_summary = []
    for name, path in FILES.items():
        print(f"analyzing {name}...")
        summary, detail = analyze(name, path)
        all_summary.extend(summary)
        ws = out.create_sheet(f"{name} detail")
        ws.append(["rip_code", "rip_description (first)", "upc", "status",
                   "cpl_product_names", "cpl_rip_codes_on_this_upc"])
        for c in ws[1]:
            c.font = Font(bold=True)
        detail.sort(key=lambda r: (r[0], r[3]))
        for row in detail:
            ws.append(row)
        n_mis = sum(1 for r in summary if r[8] > 0)
        print(f"  {name}: {len(summary)} RIP codes, {n_mis} with a mismatch, "
              f"{len(detail)} non-matched (code, upc) pairs")
    # worst mismatches first, then by sheet size
    all_summary.sort(key=lambda r: (-r[8], -r[3]))
    for row in all_summary:
        s.append(row)
    for c in s[1]:
        c.font = Font(bold=True)
    s.freeze_panes = "A2"
    out.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
