"""RIP membership mismatch analysis v3 (2026-06, after commit e295740).

Same comparison as v2 (sheet membership vs what the product-page combo
shows), re-run against the rules NOW live, with one refinement: a junk-UPC
sheet row counts as RESOLVED when the code has ANY code-claiming CPL row
that the sheet couldn't address by barcode (junk-UPC claim rows AND
real-UPC claim rows outside the sheet's barcode set) — that is exactly the
member set /rip-siblings serves since the code-claim join shipped.

mismatch_v1 = sheet - (matched_code + kept_single_listing)   [old rules]
mismatch_v3 = stubs (real barcode, not in CPL)
            + junk sheet UPCs with no code-claim coverage

Output: Data/Enhancement/RIP membership mismatch analysis 2026-06 v3.xlsx
Pure analysis; changes nothing else.
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
P = (ROOT / "parquet_output").as_posix()
WHOLESALERS = ("allied", "fedway")

# Edition from --edition (monthly_load.py passes the new month); default June.
import argparse

_ap = argparse.ArgumentParser()
_ap.add_argument("--edition", default="2026-06")
EDITION = _ap.parse_args().edition
OUT = ROOT / "Data" / "Enhancement" / f"RIP membership mismatch analysis {EDITION} v3.xlsx"


def is_clean_upc(u: str) -> bool:
    s = (u or "").strip()
    if s in ("", "0"):
        return False
    if re.fullmatch(r"(0+|9+|1+)", s):
        return False
    if s.startswith("999999"):
        return False
    if re.match(r"^(\d)\1{8,}", s):
        return False
    return len(s.lstrip("0")) >= 8


def main() -> None:
    con = duckdb.connect()
    rip = f"read_parquet('{P}/rip/*/*/*.parquet', hive_partitioning=1)"
    cpl = f"read_parquet('{P}/derived/cpl_enriched.parquet')"

    sheet: dict[tuple, set] = defaultdict(set)
    desc: dict[tuple, str] = {}
    for w, rc, u, d in con.execute(f"""
        SELECT wholesaler, CAST(rip_code AS VARCHAR),
               LTRIM(CAST(upc AS VARCHAR), '0'), ANY_VALUE(rip_description)
        FROM {rip}
        WHERE edition = '{EDITION}' AND wholesaler IN {WHOLESALERS}
          AND rip_code IS NOT NULL AND upc IS NOT NULL
        GROUP BY 1, 2, 3
    """).fetchall():
        rc = (rc or "").strip()
        if not rc or rc in ("0", "None", "nan"):
            continue
        if u and u.strip() and u.strip() not in ("None", "nan"):
            sheet[(w, rc)].add(u.strip())
        desc.setdefault((w, rc), str(d or "").strip())

    cpl_idx: dict[tuple, list] = defaultdict(list)
    claims: dict[tuple, list] = defaultdict(list)
    for w, un, pn, rc in con.execute(f"""
        SELECT wholesaler, LTRIM(CAST(upc AS VARCHAR), '0'), product_name,
               CAST(rip_code AS VARCHAR)
        FROM {cpl}
        WHERE edition = '{EDITION}' AND wholesaler IN {WHOLESALERS}
    """).fetchall():
        un = (un or "").strip()
        rcs = (rc or "").strip()
        if un and un not in ("None", "nan"):
            cpl_idx[(w, un)].append((pn or "", rcs))
        if rcs and rcs not in ("0", "None", "nan"):
            claims[(w, rcs)].append((un, pn or ""))

    summary, detail, stub_rows = [], [], []
    tot = dict(codes=0, mis1_codes=0, mis3_codes=0, mis1=0, mis3=0)
    for (w, rc), upcs in sorted(sheet.items()):
        c1 = dict(matched=0, kept_single=0, dropped_multi=0, not_in_cpl=0)
        c3 = dict(shown=0, stubs=0, junk=0, junk_unresolved=0)
        claim_rows = claims.get((w, rc), [])
        # Live rule: every code-claiming CPL row is a member. Junk sheet rows
        # are covered when the claim set reaches beyond the sheet's barcodes.
        extra_claims = [(u, n) for u, n in claim_rows
                        if not is_clean_upc(u) or u not in upcs]
        junk_resolved = len(extra_claims) > 0
        for u in sorted(upcs):
            rows = cpl_idx.get((w, u), [])
            names = {n.strip().upper() for n, _ in rows if n}
            code_match = any(c == rc for _, c in rows)
            if not rows:
                c1["not_in_cpl"] += 1
            elif code_match:
                c1["matched"] += 1
            elif len(names) <= 1:
                c1["kept_single"] += 1
            else:
                c1["dropped_multi"] += 1
            if not is_clean_upc(u):
                c3["junk"] += 1
                if not junk_resolved:
                    c3["junk_unresolved"] += 1
                    detail.append([w, rc, u, "junk_unresolved",
                                   "; ".join(sorted(names))[:120]])
            elif not rows:
                c3["stubs"] += 1
                detail.append([w, rc, u, "stub_not_in_cpl", ""])
                stub_rows.append([w, rc, desc.get((w, rc), "")[:90], u])
            else:
                c3["shown"] += 1
        mis1 = len(upcs) - (c1["matched"] + c1["kept_single"])
        mis3 = c3["stubs"] + c3["junk_unresolved"]
        summary.append([w, rc, desc.get((w, rc), "")[:90], len(upcs),
                        c1["matched"], c1["kept_single"], c1["dropped_multi"],
                        c1["not_in_cpl"], mis1,
                        c3["shown"], c3["junk"], len(extra_claims),
                        c3["stubs"], c3["junk_unresolved"], mis3])
        tot["codes"] += 1
        tot["mis1"] += mis1
        tot["mis3"] += mis3
        tot["mis1_codes"] += 1 if mis1 > 0 else 0
        tot["mis3_codes"] += 1 if mis3 > 0 else 0

    out = openpyxl.Workbook()
    s = out.active
    s.title = "Summary"
    s.append(["wholesaler", "rip_code", "rip_description (first)",
              "sheet_upc_count",
              "v1 matched_code", "v1 kept_single", "v1 dropped_multi",
              "v1 not_in_cpl", "v1 MISMATCH",
              "v3 shown_buyable", "v3 junk_upcs", "v3 code-claim members",
              "v3 stubs (not in CPL)", "v3 junk_unresolved", "v3 MISMATCH"])
    for row in sorted(summary, key=lambda r: (-r[14], -r[8], -r[3])):
        s.append(row)
    for c in s[1]:
        c.font = Font(bold=True)
    s.freeze_panes = "A2"
    # Mark the "not in CPL" counts so they jump out: amber fill + red bold on
    # every Summary row where the RIP sheet grants barcodes the CPL lacks.
    amber = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_bold = Font(bold=True, color="C00000")
    stub_col = 13   # "v3 stubs (not in CPL)"
    for r in range(2, s.max_row + 1):
        cell = s.cell(row=r, column=stub_col)
        if isinstance(cell.value, (int, float)) and cell.value > 0:
            cell.fill = amber
            cell.font = red_bold

    # Dedicated sheet: every UPC the RIP sheet grants that is NOT on the
    # June CPL, with headline counts on top.
    nc = out.create_sheet("UPCs not in CPL", 1)
    by_w: dict = {}
    for w, *_ in stub_rows:
        by_w[w] = by_w.get(w, 0) + 1
    nc.append(["UPCs granted by the RIP sheet but NOT FOUND in the CPL tab "
               f"({EDITION})"])
    nc.append([f"TOTAL: {len(stub_rows)}"]
              + [f"{w}: {n}" for w, n in sorted(by_w.items())])
    nc.append([])
    nc.append(["wholesaler", "rip_code", "rip_description (first)", "upc"])
    for c in nc[1] + nc[2] + nc[4]:
        c.font = Font(bold=True)
    for row in sorted(stub_rows):
        nc.append(row)
    nc.freeze_panes = "A5"

    ws = out.create_sheet("v3 remaining issues")
    ws.append(["wholesaler", "rip_code", "upc", "v3_status", "cpl_names_on_upc"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in sorted(detail):
        ws.append(row)

    t = out.create_sheet("Totals")
    t.append(["metric", "v1 (old rules)", "v3 (live rules)"])
    t.append(["RIP codes analysed", tot["codes"], tot["codes"]])
    t.append(["codes with a mismatch", tot["mis1_codes"], tot["mis3_codes"]])
    t.append(["mismatched (code, upc) pairs", tot["mis1"], tot["mis3"]])
    t.append(["UPCs not found in CPL (stubs)", "",
              f"{len(stub_rows)} ("
              + ", ".join(f"{w} {n}" for w, n in sorted(by_w.items())) + ")"])
    for c in t[1]:
        c.font = Font(bold=True)

    try:
        out.save(OUT)
        saved = OUT
    except PermissionError:
        # The workbook is open in Excel; save beside it instead of failing.
        saved = OUT.with_name(OUT.stem + " (updated).xlsx")
        out.save(saved)
    print(f"UPCs not in CPL: {len(stub_rows)} "
          + str({w: n for w, n in sorted(by_w.items())}))
    print(f"codes: {tot['codes']}")
    print(f"codes with mismatch: v1 {tot['mis1_codes']} -> v3 {tot['mis3_codes']}")
    print(f"mismatched pairs:    v1 {tot['mis1']} -> v3 {tot['mis3']}")
    print(f"wrote {saved}")


if __name__ == "__main__":
    main()
