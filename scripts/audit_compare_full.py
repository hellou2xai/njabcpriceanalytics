"""Compare Prices audit, FULL DETAIL (Kramer vs Shore Point) — every row the
page shows, with both distributors' own catalogue names, every number on
screen, the same numbers from the API payload and from cpl_enriched, the
difference shown on screen and the difference computed from the backend.
Notes are factual value-vs-value statements only.

Inputs: scripts/_audit_compare_out.json (from frontend/audit_compare.mjs).
Output: Data/Enhancement/Compare Prices audit Kramer vs ShorePoint <ed> FULL.xlsx
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
IN = ROOT / "scripts" / "_audit_compare_out.json"
PARQUET = ROOT / "parquet_output" / "derived" / "cpl_enriched.parquet"
WS = ["kramer", "shore_point"]
DISPLAY = {"kramer": "Kramer", "shore_point": "Shore Point"}
TOL = 0.011


def parse_money(t):
    if not t:
        return None
    m = re.search(r"-?\$?\s*([\d,]+(?:\.\d+)?)", str(t))
    return float(m.group(1).replace(",", "")) if m else None


def norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def n_upc(v) -> str:
    return re.sub(r"\D", "", str(v or "")).lstrip("0")


def r2(v):
    return None if v is None else round(float(v), 2)


def main() -> None:
    payload = json.loads(IN.read_text(encoding="utf-8"))
    dom, api = payload["dom"], payload["api"]
    api_rows = api["rows"]
    editions = api.get("editions", {})
    ed_label = next(iter(editions.values()), "2026-06")
    out_path = (ROOT / "Data" / "Enhancement" /
                f"Compare Prices audit Kramer vs ShorePoint {ed_label} FULL.xlsx")

    def api_key(r) -> str:
        size = f"{r.get('unit_qty')} × {r.get('unit_volume')}"
        if r.get("vintage"):
            size += f" · {r['vintage']}"
        return f"{norm(r.get('product_name'))}|{norm(size)}"

    api_idx: dict[str, dict] = {}
    for r in api_rows:
        api_idx.setdefault(api_key(r), r)

    con = duckdb.connect()
    pq = con.execute(
        f"""SELECT wholesaler, edition, CAST(upc AS VARCHAR) AS upc, product_name,
                   unit_qty, unit_volume, unit_type, vintage,
                   frontline_case_price, best_case_price, effective_case_price,
                   frontline_unit_price, total_savings_per_case,
                   has_discount, has_rip, CAST(rip_code AS VARCHAR) AS rip_code
            FROM '{PARQUET.as_posix()}'
            WHERE wholesaler IN ('kramer', 'shore_point')""",
    ).fetchdf()
    pq_idx: dict[tuple, list[dict]] = {}
    for rec in pq.to_dict("records"):
        pq_idx.setdefault((rec["wholesaler"], rec["edition"], n_upc(rec["upc"])), []).append(rec)

    def month_row(w: str, p: dict) -> dict | None:
        rows = pq_idx.get((w, p.get("edition"), n_upc(p.get("upc"))))
        if not rows:
            return None
        if len(rows) > 1:
            named = [r for r in rows if norm(r["product_name"]) == norm(p.get("product_name"))]
            if named:
                rows = named
        return rows[0]

    all_rows: list[list] = []
    issue_count = 0
    for row in dom:
        ar = api_idx.get(f"{norm(row['name'])}|{norm(row['size'])}")
        cells = [parse_money(c["text"]) for c in row["prices"]]
        rec: dict = {
            "display_name": row["name"], "size": row["size"],
            "screen_spread": parse_money(row["spread_text"]),
            "winner": row["winner_text"],
        }
        issues: list[str] = []
        nets_screen, nets_api, nets_month = {}, {}, {}
        for wi, w in enumerate(WS):
            sl, sq, sn = (cells[wi * 3 + j] if wi * 3 + j < len(cells) else None
                          for j in range(3))
            p = (ar or {}).get("prices", {}).get(w) or {}
            mr = month_row(w, p) if p else None
            d = DISPLAY[w]
            rec[f"{d} name"] = (mr or {}).get("product_name") or p.get("product_name")
            rec[f"{d} upc"] = p.get("upc")
            rec[f"{d} rip_code"] = (mr or {}).get("rip_code")
            rec[f"{d} screen List"] = sl
            rec[f"{d} screen BestQD"] = sq
            rec[f"{d} screen Net"] = sn
            rec[f"{d} api List"] = r2(p.get("frontline"))
            rec[f"{d} api BestQD"] = r2(p.get("after_qd"))
            rec[f"{d} api Net"] = r2(p.get("effective"))
            rec[f"{d} month List"] = r2((mr or {}).get("frontline_case_price"))
            rec[f"{d} month BestQD"] = r2((mr or {}).get("best_case_price"))
            rec[f"{d} month Net"] = r2((mr or {}).get("effective_case_price"))
            nets_screen[w], nets_api[w] = sn, r2(p.get("effective"))
            nets_month[w] = r2((mr or {}).get("effective_case_price"))
            for label, dv, av in (("List", sl, r2(p.get("frontline"))),
                                  ("BestQD", sq, r2(p.get("after_qd"))),
                                  ("Net", sn, r2(p.get("effective")))):
                if dv is not None and av is not None and abs(dv - av) > TOL:
                    issues.append(f"{d} {label}: screen {dv} vs api {av}")
            mv = r2((mr or {}).get("effective_case_price"))
            av = r2(p.get("effective"))
            if mv is not None and av is not None and abs(mv - av) > TOL:
                issues.append(f"{d} Net: api {av} vs month {mv}")
            fv = r2((mr or {}).get("frontline_case_price"))
            al = r2(p.get("frontline"))
            if fv is not None and al is not None and abs(fv - al) > TOL:
                issues.append(f"{d} List: api {al} vs month {fv}")

        sv = [v for v in nets_screen.values() if v is not None and v > 0]
        av = [v for v in nets_api.values() if v is not None and v > 0]
        mv = [v for v in nets_month.values() if v is not None and v > 0]
        rec["backend diff (api nets)"] = r2(abs(av[0] - av[1])) if len(av) == 2 else None
        rec["month diff (cpl_enriched nets)"] = r2(abs(mv[0] - mv[1])) if len(mv) == 2 else None
        if rec["screen_spread"] is not None and len(sv) == 2 \
                and abs(rec["screen_spread"] - abs(sv[0] - sv[1])) > TOL:
            issues.append(f"Spread: shown {rec['screen_spread']} vs nets imply "
                          f"{r2(abs(sv[0] - sv[1]))}")
        if rec["screen_spread"] is not None and rec["backend diff (api nets)"] is not None:
            rec["spread vs backend delta"] = r2(abs(rec["screen_spread"]
                                                    - rec["backend diff (api nets)"]))
            if rec["spread vs backend delta"] > TOL:
                issues.append(f"Spread {rec['screen_spread']} vs backend diff "
                              f"{rec['backend diff (api nets)']}")
        else:
            rec["spread vs backend delta"] = None
        if len(av) == 2 and abs(av[0] - av[1]) > 0.005:
            cheaper = DISPLAY[WS[0]] if (nets_api[WS[0]] or 9e9) < (nets_api[WS[1]] or 9e9) \
                else DISPLAY[WS[1]]
            wt = norm(row["winner_text"])
            if wt and "TIE" not in wt and norm(cheaper) not in wt:
                issues.append(f"Winner shows {row['winner_text']} but backend "
                              f"cheaper net is {cheaper}")
        if ar is None:
            issues.append("DOM row not matched to an API row")
        rec["status"] = "OK" if not issues else "CHECK"
        rec["issues"] = " | ".join(issues)
        if issues:
            issue_count += 1
        all_rows.append(rec)

    cols = (["display_name", "size", "Kramer name", "Shore Point name",
             "Kramer upc", "Shore Point upc",
             "Kramer screen List", "Kramer screen BestQD", "Kramer screen Net",
             "Kramer api List", "Kramer api BestQD", "Kramer api Net",
             "Kramer month List", "Kramer month BestQD", "Kramer month Net",
             "Shore Point screen List", "Shore Point screen BestQD", "Shore Point screen Net",
             "Shore Point api List", "Shore Point api BestQD", "Shore Point api Net",
             "Shore Point month List", "Shore Point month BestQD", "Shore Point month Net",
             "screen_spread", "backend diff (api nets)",
             "month diff (cpl_enriched nets)", "spread vs backend delta",
             "winner", "Kramer rip_code", "Shore Point rip_code",
             "status", "issues"])

    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Summary"
    s.append(["Compare Prices FULL audit", "Kramer vs Shore Point",
              "route", payload.get("route", "")])
    s.append(["rows on screen", len(dom)])
    s.append(["rows with a flagged difference", issue_count])
    s.append(["rows fully consistent (screen = api = month, spread correct)",
              len(dom) - issue_count])
    s.append(["editions"] + [f"{k}: {v}" for k, v in editions.items()])

    d = wb.create_sheet("All rows")
    d.append(cols)
    for c in d[1]:
        c.font = Font(bold=True)
    amber = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    rows_sorted = sorted(all_rows, key=lambda r: (r["status"] == "OK", r["display_name"]))
    for rec in rows_sorted:
        d.append([rec.get(c) for c in cols])
        if rec["status"] != "OK":
            for cell in d[d.max_row]:
                cell.fill = amber
    d.freeze_panes = "C2"
    wb.save(out_path)
    print(f"rows: {len(all_rows)}  flagged: {issue_count}  ok: {len(all_rows) - issue_count}")
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
