"""Dump the SKUs that lost their RIP attribution after commit b52fd41
(the strict (rip_code, UPC) rule). The previous derive.py applied a
RIP whenever a CPL row's rip_code appeared anywhere in the RIP sheet
for that wholesaler+edition, even if the RIP sheet didn't explicitly
list that product's UPC. The new rule requires an explicit (code, UPC)
pair, matching the wholesaler's own intent.

Output: to_be_tested_after_code_change/rip_attribution_changes_<ts>.xlsx
with a cover sheet plus one sheet per current edition listing every
affected SKU. Designed for manual spot-checking against the source
Excel files before declaring the change correct on production.

Run AFTER rebuilding cpl_enriched on the new derive.py:
    python -c "from nj_abc_parser.derive import build_all; build_all()"
    PRICING_SOURCE=parquet python -c "from backend.pricing_cache import build_pricing_cache; build_pricing_cache()"
    python scripts/dump_rip_test_set.py
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.db import get_duckdb


OUT_DIR = Path(__file__).parent.parent / "to_be_tested_after_code_change"
DISTRIBUTOR_NAMES = {
    "allied": "Allied",
    "fedway": "Fedway",
    "high_grade": "Highgrade",
    "opici": "Opici",
    "peerless": "Peerless",
}


def fetch_affected(con, edition: str) -> list[dict]:
    """Rows whose cpl.rip_code is non-empty but has_rip is now false (= the
    strict pair rule excluded what the old code-level fallback used to grant).
    For each row we ALSO compute the per-case rebate the old code-level rule
    WOULD have applied, by scanning the rip parquet for any row matching one
    of the split codes (regardless of UPC). That number is the per-case rebate
    the buyer used to see and now does not."""
    return con.execute(
        f"""
        WITH affected AS (
            SELECT wholesaler, edition, product_name, brand, upc, vintage,
                   unit_volume, unit_qty, rip_code,
                   frontline_case_price, best_case_price, effective_case_price
            FROM cpl_enriched
            WHERE has_rip = false
              AND rip_code IS NOT NULL AND rip_code != '' AND rip_code != '0'
              AND edition = $ed
        ),
        codes AS (
            SELECT a.*,
                   UNNEST(string_split(REGEXP_REPLACE(a.rip_code, '\\s+', ' '), ' ')) AS code
            FROM affected a
        ),
        per_code_old AS (
            SELECT wholesaler, edition, rip_code AS code,
                   MAX(GREATEST(
                       COALESCE(CASE WHEN rip_qty_1 > 0 AND LOWER(rip_unit_1) NOT LIKE 'b%' THEN rip_amt_1 / rip_qty_1 END, 0),
                       COALESCE(CASE WHEN rip_qty_2 > 0 AND LOWER(rip_unit_2) NOT LIKE 'b%' THEN rip_amt_2 / rip_qty_2 END, 0),
                       COALESCE(CASE WHEN rip_qty_3 > 0 AND LOWER(rip_unit_3) NOT LIKE 'b%' THEN rip_amt_3 / rip_qty_3 END, 0),
                       COALESCE(CASE WHEN rip_qty_4 > 0 AND LOWER(rip_unit_4) NOT LIKE 'b%' THEN rip_amt_4 / rip_qty_4 END, 0)
                   )) AS best_case_per_case,
                   MAX(GREATEST(
                       COALESCE(CASE WHEN rip_qty_1 > 0 AND LOWER(rip_unit_1) LIKE 'b%' THEN rip_amt_1 / rip_qty_1 END, 0),
                       COALESCE(CASE WHEN rip_qty_2 > 0 AND LOWER(rip_unit_2) LIKE 'b%' THEN rip_amt_2 / rip_qty_2 END, 0),
                       COALESCE(CASE WHEN rip_qty_3 > 0 AND LOWER(rip_unit_3) LIKE 'b%' THEN rip_amt_3 / rip_qty_3 END, 0),
                       COALESCE(CASE WHEN rip_qty_4 > 0 AND LOWER(rip_unit_4) LIKE 'b%' THEN rip_amt_4 / rip_qty_4 END, 0)
                   )) AS best_bottle_per_bottle
            FROM rip
            WHERE rip_code IS NOT NULL
            GROUP BY wholesaler, edition, rip_code
        ),
        joined AS (
            SELECT c.*,
                   GREATEST(
                       COALESCE(p.best_case_per_case, 0),
                       COALESCE(p.best_bottle_per_bottle, 0)
                           * COALESCE(TRY_CAST(c.unit_qty AS DOUBLE), 1)
                   ) AS old_per_case_rebate
            FROM codes c
            LEFT JOIN per_code_old p
              ON p.wholesaler = c.wholesaler
             AND p.edition = c.edition
             AND p.code = c.code
        )
        SELECT wholesaler, edition, product_name, brand, upc, vintage,
               unit_volume, unit_qty, rip_code,
               frontline_case_price, best_case_price, effective_case_price,
               MAX(old_per_case_rebate) AS old_per_case_rebate_was
        FROM joined
        GROUP BY wholesaler, edition, product_name, brand, upc, vintage,
                 unit_volume, unit_qty, rip_code,
                 frontline_case_price, best_case_price, effective_case_price
        ORDER BY wholesaler, product_name
        """,
        {"ed": edition},
    ).fetchdf().to_dict(orient="records")


def _header(ws, headers):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="left", vertical="center")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = fill; c.font = font; c.alignment = align
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _autosize(ws, max_w=44):
    for col in ws.columns:
        # `col[0].column_letter` is unreliable on merged ranges; iterate to find max.
        col_letter = get_column_letter(col[0].column)
        m = 8
        for cell in col:
            v = cell.value
            if v is None:
                continue
            m = max(m, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[col_letter].width = m


def add_cover(wb: Workbook, summary: dict[str, int]) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 100
    ws["A1"] = "RIP-attribution test set — strict (rip_code, UPC) rule"
    ws["A1"].font = Font(bold=True, size=16)
    ws.row_dimensions[1].height = 28

    body = [
        "",
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        f"Trigger commit: b52fd41 (Drop code-level RIP fallback - only (rip_code, UPC) pairs count)",
        "",
        "What changed",
        "-------------",
        "The previous derive.py used a TWO-step RIP join: first by (rip_code, UPC) and then,",
        "if that missed, a CODE-LEVEL fallback that matched ANY rip-sheet row carrying the same",
        "code regardless of which UPCs it listed. That fallback was originally added because some",
        "wholesalers (notably Fedway) anchor a RIP to a stub UPC like 812066000000 for an entire",
        "product line. The side effect was that any product carrying that code 'inherited' a RIP",
        "even when the wholesaler's RIP sheet didn't explicitly enrol that UPC.",
        "",
        "Per the user's canonical rule: the RIP sheet is the primary decision for RIP applicability.",
        "A RIP applies to a product ONLY when the sheet has a row explicitly pairing this product's",
        "UPC with the code. The code-level fallback was removed in both derive.py (the rip_per_code",
        "CTE) and backend/pricing.py (the rip_by_code dict in attach_tiers).",
        "",
        "Effect",
        "------",
        "Every SKU listed in the per-edition sheets used to receive a code-level-fallback RIP. After",
        "the change, these SKUs no longer have a RIP attached because the RIP sheet does not list",
        "their UPC under any of their rip codes. Their effective_case_price is now equal to their",
        "best_case_price (CPL discount only). Their rip_savings is 0. has_rip is false.",
        "",
        "Counts (this edition):",
        "",
    ]
    for ed, n in summary.items():
        body.append(f"   {ed}: {n} SKUs")
    body += [
        "",
        "How to verify",
        "-------------",
        "1. Open the wholesaler's source Excel file for the month in question.",
        "2. On the RIP sheet, search for the rip_code (or each space-separated code).",
        "3. Confirm the UPC column does NOT list this product's UPC for that code.",
        "4. If the UPC truly is not listed, the new rule is correct: no RIP applies.",
        "5. If the UPC IS listed but we still excluded it, surface the row to engineering.",
        "",
        "Column definitions",
        "------------------",
        "  Distributor                — wholesaler (Allied / Fedway / Highgrade / Opici / Peerless)",
        "  Edition (YYYY-MM)          — source-file month",
        "  Product Name               — as on the CPL sheet",
        "  Brand                      — extracted brand name",
        "  UPC                        — primary identifier on the CPL row",
        "  Vintage                    — wine vintage (blank for non-wine)",
        "  Unit Volume                — bottle size (e.g. 750ML, 1L)",
        "  Unit Qty                   — bottles per case",
        "  RIP Code(s)                — value of CPL.rip_code (may be space-separated)",
        "  Frontline /cs              — sticker price per case",
        "  Best /cs (CPL only)        — after CPL discount; now equals effective",
        "  Effective /cs (NEW)        — current value under the strict rule (= Best /cs)",
        "  Old per-case RIP $         — what the code-level fallback used to grant per case",
        "                               (max across all rip-sheet rows under any matching code,",
        "                               irrespective of UPC). 0 means the row had no code-level",
        "                               rebate either; it appears here only because rip_code was",
        "                               non-empty but no rebate was ever active.",
        "",
        "Filename convention",
        "-------------------",
        "rip_attribution_changes_<UTC timestamp>.xlsx",
        "",
    ]
    for i, line in enumerate(body, 2):
        c = ws.cell(row=i, column=1, value=line)
        if line and not line.startswith(("---", "   ")) and (line.endswith("changed") or line.endswith("Effect") or line.endswith("verify") or line.endswith("definitions") or line.endswith("convention") or line == "Counts (this edition):"):
            c.font = Font(bold=True)
    # Keep the cover narrow-ish for screen readability.
    ws.sheet_view.showGridLines = False


def add_edition_sheet(wb: Workbook, edition: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=edition)
    headers = [
        "Distributor", "Edition (YYYY-MM)", "Product Name", "Brand", "UPC",
        "Vintage", "Unit Volume", "Unit Qty", "RIP Code(s)",
        "Frontline /cs", "Best /cs (CPL only)", "Effective /cs (NEW)",
        "Old per-case RIP $ (was applied)",
    ]
    _header(ws, headers)
    money_fill = PatternFill("solid", fgColor="FEF3C7")
    money_cols = (10, 11, 12, 13)
    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=DISTRIBUTOR_NAMES.get(row["wholesaler"], row["wholesaler"]))
        ws.cell(row=r_idx, column=2, value=row["edition"])
        ws.cell(row=r_idx, column=3, value=row.get("product_name"))
        ws.cell(row=r_idx, column=4, value=row.get("brand"))
        # UPC as text so Excel doesn't strip leading zeros / convert to float.
        upc = row.get("upc")
        upc_cell = ws.cell(row=r_idx, column=5, value=("" if upc is None else str(upc)))
        upc_cell.number_format = "@"
        ws.cell(row=r_idx, column=6, value=row.get("vintage"))
        ws.cell(row=r_idx, column=7, value=row.get("unit_volume"))
        ws.cell(row=r_idx, column=8, value=row.get("unit_qty"))
        ws.cell(row=r_idx, column=9, value=row.get("rip_code"))
        for col, key in ((10, "frontline_case_price"), (11, "best_case_price"),
                         (12, "effective_case_price"), (13, "old_per_case_rebate_was")):
            v = row.get(key)
            c = ws.cell(row=r_idx, column=col, value=(float(v) if v is not None else None))
            c.number_format = '"$"#,##0.00'
        # Soft-highlight the "Old per-case RIP $ (was applied)" column for fast
        # scanning. Bigger numbers = bigger user-visible price change.
        ws.cell(row=r_idx, column=13).fill = money_fill
    _autosize(ws)
    # Wider for product name + rip codes
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 36)
    ws.column_dimensions["I"].width = max(ws.column_dimensions["I"].width, 18)


def main() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with get_duckdb() as con:
        editions = [r[0] for r in con.execute(
            "SELECT DISTINCT edition FROM cpl_enriched WHERE edition >= '2026-04' ORDER BY edition"
        ).fetchall()]
        per_edition: dict[str, list[dict]] = {ed: fetch_affected(con, ed) for ed in editions}

    summary = {ed: len(rows) for ed, rows in per_edition.items()}
    wb = Workbook()
    add_cover(wb, summary)
    for ed, rows in per_edition.items():
        add_edition_sheet(wb, ed, rows)

    ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    out = OUT_DIR / f"rip_attribution_changes_{ts}.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    for ed, n in summary.items():
        print(f"  {ed}: {n} SKUs")
    return out


if __name__ == "__main__":
    main()
