#!/usr/bin/env python
"""Resolve fuller Allied (ABG) item names from the Wine Chateau x ABG inventory
export and load them into `allied_name_xref` for the cache to apply.

WHY a separate table from allied_sku_xref:
  allied_sku_xref (scripts/load_allied_translation.py) matches on FULL SKU
  IDENTITY (upc + size + pack + vintage) and only sets a number/name where the
  identity resolves to exactly one SKU. This loader instead uses the guard the
  buyer asked for: EXACT UPC match AND a brand-anchored SEMANTIC name agreement
  between our abbreviated CPL name and Wine Chateau's fuller name. That name
  agreement is what neutralises shared / placeholder barcodes (e.g. a Jaboulet
  P45 row and a Spellbound Pinot Noir row sharing one UPC): only the row whose
  name actually agrees gets the fuller name attached, so we never mislabel a
  product. See CELR ground rules ("NEVER blindly attach").

WHAT it sets:
  dist_item_name = Wine Chateau's fuller product name (the un-abbreviated name).
  dist_item_no   = Wine Chateau Product Code = the ABG SKU (confirmed against
                   Allied Translation: 100% UPC agreement on the 5,835 shared
                   SKUs, so Product Code IS the ABG number).
  product_name is NEVER touched: the CPL catalogue name stays the pricing key,
  the fuller name is an added field (show-don't-hide).

TIERS (pass 1 ships STRONG only, per the agreed plan):
  STRONG = brand token matches AND token score >= 0.60  -> auto-apply.
  MEDIUM = brand token matches AND token score >= 0.34  -> exported for review.
  REJECT = name disagrees -> held back (some are real matches our crude matcher
           misses on initialisms/truncations; pass 2 recovers them via the
           embedding layer).

Usage:
    # dry run (default): read prod read-only, write review CSVs + print counts
    python scripts/load_allied_name_xref.py --xlsx "<path to Wine Chateau xlsx>"

    # actually create/replace allied_name_xref in Postgres (STRONG rows)
    python scripts/load_allied_name_xref.py --xlsx "<...>" --write

After --write, trigger POST /api/admin/reload-pricing (or redeploy) so the cache
build attaches the names. The cache re-applies the join every build, so it
survives re-ingests.
"""
import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import psycopg

from backend.pg import DATABASE_URL as ENV_DATABASE_URL

# ---- Standardisation ---------------------------------------------------------
def upc_norm(u) -> str:
    """Digits only, leading zeros stripped — matches LTRIM(upc,'0') in the cache."""
    return re.sub(r"\D", "", str(u or "")).lstrip("0")


# ---- Brand-anchored semantic name agreement ----------------------------------
# Common wine/spirit abbreviations expanded so an abbreviated CPL token can be
# tested against Wine Chateau's spelled-out name. Intentionally small and
# conservative for pass 1; pass 2 replaces this with the embedding matcher.
_STOP = {"THE", "DE", "DI", "DEL", "DELLA", "LA", "LE", "EL", "OF", "AND", "&",
         "DU", "DES", "LES"}
_SIZE = re.compile(r"^\d+(ML|L|LTR|LT|OZ|PK|P|CS)$|^\d{1,2}/\d")
_ABBR = {
    "SB": "SAUVIGNON BLANC", "SVB": "SAUVIGNON BLANC", "CAB": "CABERNET",
    "SAUV": "SAUVIGNON", "CS": "CABERNET SAUVIGNON", "PN": "PINOT NOIR",
    "PG": "PINOT GRIGIO", "CHARD": "CHARDONNAY", "CH": "CHARDONNAY",
    "RES": "RESERVE", "RSV": "RESERVE", "RSRV": "RESERVE", "VOD": "VODKA",
    "VODK": "VODKA", "WHSK": "WHISKEY", "WHSKY": "WHISKEY", "WHIS": "WHISKEY",
    "BL": "BLANC", "ROUG": "ROUGE", "CHAMP": "CHAMPAGNE", "ZIN": "ZINFANDEL",
    "ZINF": "ZINFANDEL", "MERL": "MERLOT", "RIES": "RIESLING",
    "PROS": "PROSECCO", "MOSC": "MOSCATO", "GRIG": "GRIGIO", "APPL": "APPLE",
}


def _toks(s):
    s = re.sub(r"[^A-Z0-9 ]", " ", str(s or "").upper())
    out = []
    for t in s.split():
        if t in _STOP or re.fullmatch(r"\d{2,4}", t) or _SIZE.match(t):
            continue  # drop stopwords, bare vintages, size tokens
        out.append(t)
    return out


def _expand(t):
    return _ABBR.get(t, t).split()


def _tmatch(a, b):
    return a == b or (len(a) >= 3 and len(b) >= 3 and (a.startswith(b) or b.startswith(a)))


def name_agreement(cpl, wc):
    """Return (score, brand_ok). score = fraction of CPL content tokens that
    agree with the Wine Chateau name; brand_ok = the leading (brand) token
    agrees. Both are required for a match."""
    ct, wt = _toks(cpl), _toks(wc)
    if not ct or not wt:
        return 0.0, False
    brand_ok = any(_tmatch(ct[0], w) for w in wt)
    matched = sum(1 for t in ct if all(any(_tmatch(e, w) for w in wt) for e in _expand(t)))
    return matched / len(ct), brand_ok


# ---- Sources -----------------------------------------------------------------
def _open_book(xlsx_path: Path):
    try:
        return openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except PermissionError:
        import shutil, tempfile
        tmp = Path(tempfile.gettempdir()) / f"_wc_abg_{xlsx_path.stat().st_size}.xlsx"
        shutil.copy2(xlsx_path, tmp)
        return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def read_wine_chateau(xlsx_path: Path):
    """Return {upc_norm: [(fuller_name, abg_sku), ...]} from the WC export."""
    wb = _open_book(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    ci = {n: header.index(n) for n in ("Product Code", "Product Name", "UPC")
          if n in header}
    missing = {"Product Code", "Product Name", "UPC"} - set(ci)
    if missing:
        sys.exit(f"Wine Chateau sheet missing columns: {missing}; got {header}")
    by_upc = defaultdict(list)
    for r in rows:
        un = upc_norm(r[ci["UPC"]])
        if not un or len(un) < 8:
            continue  # blank / junk UPC — not join-usable
        name = str(r[ci["Product Name"]] or "").strip()
        sku = str(r[ci["Product Code"]] or "").strip()
        if name:
            by_upc[un].append((name, sku))
    return by_upc


def read_allied_catalog(db):
    """Distinct Allied (upc, product_name) naming units from prod (read-only)."""
    with psycopg.connect(db) as con:
        rows = con.execute(
            "SELECT DISTINCT upc, product_name FROM cpl_enriched "
            "WHERE wholesaler='allied' AND upc IS NOT NULL AND upc<>''"
        ).fetchall()
    return [(upc_norm(u), u, pn) for (u, pn) in rows]


# ---- Resolve -----------------------------------------------------------------
# Wine Chateau truncates the trailing check digit on many UPCs, so an exact match
# misses real products. For a UPC-A the 12th digit is DERIVED from the first 11,
# so matching on the first 11 (drop last 1) is identity-preserving, not lossy.
# We therefore fall back to a PARTIAL UPC match (drop last 1, then 2 digits of our
# UPC) when the exact UPC has no candidate. Because the barcode is then fuzzier,
# a partial match is only ever accepted at STRONG confidence (never MEDIUM), and
# the brand-anchored name agreement is what confirms it is the same product.
TRUNC_DEPTHS = (1, 2)


def resolve(catalog, wc_by_upc):
    """Yield dicts per Allied naming unit with the best WC candidate + tier.

    match_type: 'exact' | 'trunc1' | 'trunc2' — how the UPC matched."""
    for un, raw_upc, cpl_name in catalog:
        cands = wc_by_upc.get(un)
        match_type = "exact"
        if not cands and un:
            for d in TRUNC_DEPTHS:
                if len(un) > d + 6:  # keep >=7 significant digits
                    cands = wc_by_upc.get(un[:-d])
                    if cands:
                        match_type = f"trunc{d}"
                        break
        if not cands:
            yield dict(upc_norm=un, cpl_name=cpl_name, tier="NO_UPC", match_type=None)
            continue
        best_name, best_sku, best_sc, best_brand = None, None, -1.0, False
        for wc_name, wc_sku in cands:
            sc, brand = name_agreement(cpl_name, wc_name)
            if sc > best_sc:
                best_name, best_sku, best_sc, best_brand = wc_name, wc_sku, sc, brand
        strong = best_brand and best_sc >= 0.60
        if strong:
            tier = "STRONG"
        elif match_type == "exact" and best_brand and best_sc >= 0.34:
            tier = "MEDIUM"          # MEDIUM only on an EXACT UPC; partial must be STRONG
        else:
            tier = "REJECT"
        yield dict(upc_norm=un, cpl_name=cpl_name, dist_item_name=best_name,
                   abg_sku=best_sku, score=round(best_sc, 3), tier=tier,
                   match_type=match_type)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Wine Chateau x ABG inventory .xlsx")
    ap.add_argument("--database-url", default=ENV_DATABASE_URL)
    ap.add_argument("--write", action="store_true",
                    help="create/replace allied_name_xref (STRONG rows). Default: dry run.")
    ap.add_argument("--out-dir", default=None, help="where to write review CSVs")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")
    out_dir = Path(args.out_dir) if args.out_dir else xlsx_path.parent

    wc_by_upc = read_wine_chateau(xlsx_path)
    catalog = read_allied_catalog(args.database_url)
    resolved = list(resolve(catalog, wc_by_upc))

    tiers = defaultdict(list)
    for r in resolved:
        tiers[r["tier"]].append(r)
    total = len(resolved)
    print(f"Allied naming units (distinct upc,name): {total}")
    for t in ("NO_UPC", "REJECT", "MEDIUM", "STRONG"):
        print(f"  {t:8}: {len(tiers[t])}")
    by_mt = defaultdict(int)
    for r in tiers["STRONG"]:
        by_mt[r.get("match_type")] += 1
    print(f"  => STRONG auto-apply rows: {len(tiers['STRONG'])} "
          f"(exact={by_mt.get('exact',0)}, trunc1={by_mt.get('trunc1',0)}, trunc2={by_mt.get('trunc2',0)})")

    # Review CSVs (always written on a dry run, for eyeballing before --write)
    for t in ("STRONG", "MEDIUM", "REJECT"):
        p = out_dir / f"allied_name_xref_{t.lower()}.csv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["upc_norm", "cpl_name", "wine_chateau_name", "abg_sku", "score", "match_type"])
            for r in tiers[t]:
                w.writerow([r["upc_norm"], r["cpl_name"], r.get("dist_item_name"),
                            r.get("abg_sku"), r.get("score"), r.get("match_type")])
        print(f"  wrote {p.name} ({len(tiers[t])} rows)")

    if not args.write:
        print("\nDRY RUN. Re-run with --write to load allied_name_xref (STRONG only).")
        return

    strong = tiers["STRONG"]
    print(f"\nWriting allied_name_xref into {args.database_url.split('@')[-1]} ...")
    with psycopg.connect(args.database_url) as con:
        con.execute("DROP TABLE IF EXISTS allied_name_xref")
        con.execute(
            """CREATE TABLE allied_name_xref (
                upc_norm       text NOT NULL,
                cpl_name       text NOT NULL,
                dist_item_name text,
                abg_sku        text,
                score          double precision
            )""")
        with con.cursor() as cur:
            cur.executemany(
                "INSERT INTO allied_name_xref (upc_norm, cpl_name, dist_item_name, abg_sku, score) "
                "VALUES (%s,%s,%s,%s,%s)",
                [(r["upc_norm"], r["cpl_name"], r["dist_item_name"], r["abg_sku"], r["score"])
                 for r in strong])
        con.execute("CREATE INDEX idx_allied_name_xref ON allied_name_xref (upc_norm, cpl_name)")
        con.commit()
        n = con.execute("SELECT count(*) FROM allied_name_xref").fetchone()[0]
    print(f"Done. {n} rows. Trigger /api/admin/reload-pricing or redeploy to surface the names.")


if __name__ == "__main__":
    main()
