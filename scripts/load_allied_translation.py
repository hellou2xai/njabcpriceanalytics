#!/usr/bin/env python
"""Load Allied's authoritative SKU translation into Postgres as allied_sku_xref.

Source: "Data/ETL/Allied Translation.xlsx" — Allied Beverage Group's own item
number (SKU) against UPC + pack (case size) + volume + vintage. This sheet is the
TRUTH for Allied item numbers.

Unlike the legacy UPC-only sku_mapping, the catalogue is matched to this sheet on
the FULL SKU identity (UPC + vintage + pack + volume), because one UPC can carry
several SKUs (different vintages/packs/sizes). We keep ONLY the UNAMBIGUOUS
identities (those resolving to exactly one SKU); genuinely ambiguous ones are
dropped so a wrong number is never attached.

The cache build (backend/pricing_cache.py) joins cpl_enriched (Allied rows) to
this table on the same four normalised keys and sets dist_item_no = sku, so every
covered Allied row shows its real ABG number (same mechanism Fedway uses). Run
this once; a cache reload then surfaces the numbers, and it survives re-ingests
because the cache re-applies the join from this stable table every build.

Usage:
    python scripts/load_allied_translation.py
    python scripts/load_allied_translation.py --xlsx "Data/ETL/Allied Translation.xlsx"
    python scripts/load_allied_translation.py --database-url "postgresql://..."

Defaults to DATABASE_URL from the environment / .env. After it runs, trigger
POST /api/admin/reload-pricing (or redeploy) so the cache picks up the numbers.
"""
import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import psycopg

from backend.pg import DATABASE_URL as ENV_DATABASE_URL

DEFAULT_XLSX = "Data/ETL/Allied Translation.xlsx"


# ---- Normalisation (kept byte-aligned with the SQL in pricing_cache.py) ------
def upc_norm(u) -> str:
    """Digits only, leading zeros stripped — matches LTRIM(upc,'0') in the cache."""
    return re.sub(r"\D", "", str(u or "")).lstrip("0")


def vintage_norm(v) -> str:
    """'' for NV/blank; 4-digit year passthrough; 2-digit -> 20xx (<=30) else 19xx.
    Mirrors _VINTAGE_NORM_SQL in pricing_cache.py."""
    s = str(v if v is not None else "").strip().upper()
    if s in ("", "NAN", "NONE", "NV", "N/A", "NA"):
        return ""
    m = re.match(r"^(\d{4})", s)
    if m:
        return m.group(1)
    m2 = re.match(r"^(\d{2})$", s)
    if m2:
        n = int(m2.group(1))
        return ("20" if n <= 30 else "19") + m2.group(1)
    return ""


def pack_norm(p) -> str:
    """Case size as a plain integer string ('6', '12'); '' when unknown."""
    try:
        return str(int(float(p)))
    except (TypeError, ValueError):
        return ""


def size_ml(size, unit):
    """Volume in millilitres from the sheet's numeric size + unit (mL/L/floz/gal)."""
    try:
        v = float(size)
    except (TypeError, ValueError):
        return None
    u = str(unit if unit is not None else "").strip().lower()
    if u.startswith("ml"):
        return round(v)
    if u in ("l", "liter", "litre"):
        return round(v * 1000)
    if u in ("floz", "fl oz", "oz"):
        return round(v * 29.5735)
    if u.startswith("gal"):
        return round(v * 3785.41)
    return None


def _open_book(xlsx_path: Path):
    """Open the workbook, copying to a temp file first if the original is locked
    by OneDrive/Excel (a PermissionError on direct open is common on Windows)."""
    try:
        return openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except PermissionError:
        import shutil, tempfile
        tmp = Path(tempfile.gettempdir()) / f"_allied_tx_{xlsx_path.stat().st_size}.xlsx"
        shutil.copy2(xlsx_path, tmp)
        return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def read_xref(xlsx_path: Path):
    """Return {(upc_norm, size_ml, pack, vintage_norm): {sku: product_name}}."""
    wb = _open_book(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    header = None
    by_ident = defaultdict(dict)
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip().lower() if c is not None else "" for c in r]
            col = {name: header.index(name) for name in
                   ("sku", "product_name", "case_size", "size", "size_unit", "upc", "vintage_year")
                   if name in header}
            continue
        def g(name):
            j = col.get(name)
            return r[j] if j is not None and j < len(r) else None
        sku = str(g("sku") or "").strip()
        if not sku:
            continue
        un = upc_norm(g("upc"))
        if not un or len(un) < 6:
            continue  # no join-usable barcode
        ident = (un, size_ml(g("size"), g("size_unit")), pack_norm(g("case_size")),
                 vintage_norm(g("vintage_year")))
        by_ident[ident][sku] = str(g("product_name") or "").strip()
    return by_ident


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--database-url", default=ENV_DATABASE_URL)
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = Path(__file__).resolve().parent.parent / xlsx_path
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")

    by_ident = read_xref(xlsx_path)
    total = len(by_ident)
    ambiguous = [k for k, skus in by_ident.items() if len(skus) > 1]
    rows = [(un, ml, pk, vy, next(iter(skus)), next(iter(skus.values())))
            for (un, ml, pk, vy), skus in by_ident.items() if len(skus) == 1]
    print(f"identities: {total} | UNAMBIGUOUS (loaded): {len(rows)} | ambiguous (dropped): {len(ambiguous)}")

    db = args.database_url
    print(f"Writing allied_sku_xref into {db.split('@')[-1]}")
    with psycopg.connect(db) as con:
        con.execute("DROP TABLE IF EXISTS allied_sku_xref")
        con.execute(
            """CREATE TABLE allied_sku_xref (
                upc_norm     text NOT NULL,
                size_ml      integer,
                pack         text,
                vintage_norm text,
                sku          text NOT NULL,
                product_name text
            )"""
        )
        with con.cursor() as cur:
            cur.executemany(
                "INSERT INTO allied_sku_xref (upc_norm, size_ml, pack, vintage_norm, sku, product_name) "
                "VALUES (%s, %s, %s, %s, %s, %s)", rows)
        con.execute(
            "CREATE INDEX idx_allied_xref_ident ON allied_sku_xref "
            "(upc_norm, size_ml, pack, vintage_norm)")
        con.commit()
        n = con.execute("SELECT count(*) FROM allied_sku_xref").fetchone()[0]
    print(f"Done. {n} rows in allied_sku_xref. "
          "Trigger /api/admin/reload-pricing or redeploy to surface the numbers.")


if __name__ == "__main__":
    main()
