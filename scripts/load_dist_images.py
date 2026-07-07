#!/usr/bin/env python
"""Fetch distributor-supplied product images, re-host them in R2 (PERMANENT
storage, same as the Go-UPC images), and record the R2 URL in `dist_image` so
the cache can fill images the Go-UPC enrichment is missing.

Sources:
  - Allied: Wine Chateau x ABG inventory, "Primary Image (Bottle Shot)", by UPC.
  - Fedway: Fedway BR2 export, "Image URL", by PRODUCT SKU (= dist_item_no).

We only fetch images for products that LACK a Go-UPC image (so we don't re-host
what we already have). Each image is downloaded from the distributor URL and
uploaded to R2 under `dist/<wholesaler>/<key>.<ext>`; dist_image stores the R2
public URL, never the volatile distributor URL. The cache
(backend/pricing_cache.py) sets cpl_enriched.dist_image_url from this table only
where the row has no Go-UPC image, and the serving layer falls back to it. So a
distributor image never overrides a Go-UPC image or an admin override.

REQUIRES the R2_* env vars (same as scripts/enrich_products.py). Run it where
those are set (they are not in every dev env).

Usage:
    python scripts/load_dist_images.py                 # dry run (counts, no fetch)
    python scripts/load_dist_images.py --write         # fetch -> R2 -> load dist_image
    python scripts/load_dist_images.py --write --limit 50   # small test batch
"""
import argparse
import csv
import os
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl
import psycopg
import urllib.request
import backend.pg
import backend.r2 as r2
from backend.pg import DATABASE_URL as LOCAL

_EXT = {"image/jpeg": "jpg", "image/jpg": "jpg", "image/png": "png", "image/webp": "webp", "image/gif": "gif"}


def download_image(url: str):
    """(bytes, content_type) or None. Distributor image hosts, best-effort."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "celr-image-fetch/1.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            ctype = (resp.headers.get("Content-Type") or "image/jpeg").split(";")[0].strip().lower()
            data = resp.read()
        if data and ctype.startswith("image/") and len(data) > 512:
            return data, ctype
    except Exception:
        return None
    return None

ALLIED_XLSX = "Data/ETL/Wine Chateau x ABG Inventory 07062026.xlsx"
FEDWAY_CSV = "Data/ETL/Fedway BR2 2026-07.csv"


def unorm(u):
    return re.sub(r"\D", "", str(u or "")).lstrip("0")


def _http(u):
    u = (u or "").strip()
    return u if u.lower().startswith("http") else None


def read_allied(path: Path):
    """(wholesaler, upc_norm, sku_norm, image_url) rows, keyed by UPC."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    h = [str(c).strip() for c in next(it)]
    ui, ii = h.index("UPC"), h.index("Primary Image (Bottle Shot)")
    out = {}
    for r in it:
        img = _http(r[ii] if ii < len(r) else None)
        un = unorm(r[ui] if ui < len(r) else None)
        if img and un and len(un) >= 8:
            out[un] = ("allied", un, None, img)
    return list(out.values())


def read_fedway(path: Path):
    """(wholesaler, upc_norm, sku_norm, image_url) rows, keyed by SKU."""
    out = {}
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        for row in csv.DictReader(f):
            img = _http(row.get("Image URL"))
            sk = re.sub(r"\D", "", row.get("PRODUCT SKU") or "").lstrip("0")
            if img and sk:
                out[sk] = ("fedway", None, sk, img)
    return list(out.values())


def _targets(db):
    """Return (allied_upc_norms, fedway_sku_norms) that LACK a Go-UPC image."""
    with psycopg.connect(db) as con:
        au = {r[0] for r in con.execute("""
            SELECT DISTINCT LTRIM(regexp_replace(c.upc,'[^0-9]','','g'),'0') un
            FROM cpl_enriched c WHERE c.wholesaler='allied' AND c.upc IS NOT NULL AND c.upc<>''
              AND NOT EXISTS (SELECT 1 FROM product_enrichment pe
                              WHERE pe.upc=LTRIM(c.upc,'0') AND pe.image_url IS NOT NULL AND pe.image_url<>'')
        """).fetchall()}
        fs = {r[0] for r in con.execute("""
            SELECT DISTINCT LTRIM(CAST(c.dist_item_no AS VARCHAR),'0') sn
            FROM cpl_enriched c WHERE c.wholesaler='fedway' AND c.dist_item_no IS NOT NULL
              AND CAST(c.dist_item_no AS VARCHAR) NOT IN ('','0','None')
              AND NOT EXISTS (SELECT 1 FROM product_enrichment pe
                              WHERE pe.upc=LTRIM(c.upc,'0') AND pe.image_url IS NOT NULL AND pe.image_url<>'')
        """).fetchall()}
    return au, fs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()
    root = Path(__file__).resolve().parent.parent

    allied = {r[1]: r for r in read_allied(root / ALLIED_XLSX)}   # upc_norm -> row
    fedway = {r[2]: r for r in read_fedway(root / FEDWAY_CSV)}    # sku_norm -> row
    print(f"file images -> allied(by UPC): {len(allied)} | fedway(by SKU): {len(fedway)}")

    prod = os.environ.get("RENDER_EXTERNAL_DATABASE_URL")
    au, fs = _targets(prod or LOCAL)
    # products missing a Go-UPC image AND for which we have a distributor image
    todo = [allied[u] for u in au if u in allied] + [fedway[s] for s in fs if s in fedway]
    if args.limit:
        todo = todo[:args.limit]
    print(f"missing-image products with a distributor image to re-host: {len(todo)} "
          f"(allied {sum(1 for t in todo if t[0]=='allied')}, fedway {sum(1 for t in todo if t[0]=='fedway')})")

    if not args.write:
        print("DRY RUN. Re-run with --write to fetch -> R2 -> load dist_image.")
        return
    if not r2.R2_ENABLED:
        sys.exit("R2 not configured (R2_* env vars). Run where the Go-UPC image "
                 "enrichment runs, or set the R2_* vars in .env.")

    rows = []
    ok = fail = 0
    for i, (wh, un, sn, src_url) in enumerate(todo):
        dl = download_image(src_url)
        if not dl:
            fail += 1
        else:
            content, ctype = dl
            ext = _EXT.get(ctype, "jpg")
            key = f"dist/{wh}/{(un or sn)}.{ext}"
            try:
                r2_url = r2.upload_bytes(key, content, ctype)
                rows.append((wh, un, sn, r2_url))
                ok += 1
            except Exception as e:  # noqa: BLE001
                fail += 1
                print(f"  R2 upload failed {wh}/{un or sn}: {str(e)[:70]}")
        if (i + 1) % 100 == 0:
            print(f"  {i+1}/{len(todo)} re-hosted (ok={ok}, fail={fail})", flush=True)
        time.sleep(0.02)
    print(f"re-hosted {ok} images to R2 ({fail} failed)")

    for db, lbl in [(prod, "prod"), (LOCAL, "local")]:
        if not db:
            print(f"{lbl}: no URL, skipped"); continue
        with psycopg.connect(db) as con:
            con.execute("DROP TABLE IF EXISTS dist_image")
            con.execute("""CREATE TABLE dist_image (
                wholesaler text NOT NULL, upc_norm text, sku_norm text, image_url text NOT NULL)""")
            with con.cursor() as cur:
                cur.executemany(
                    "INSERT INTO dist_image (wholesaler, upc_norm, sku_norm, image_url) VALUES (%s,%s,%s,%s)", rows)
            con.execute("CREATE INDEX idx_dist_image_upc ON dist_image (wholesaler, upc_norm)")
            con.execute("CREATE INDEX idx_dist_image_sku ON dist_image (wholesaler, sku_norm)")
            con.commit()
            n = con.execute("SELECT count(*) FROM dist_image").fetchone()[0]
        print(f"  dist_image -> {lbl}: {n} rows")


if __name__ == "__main__":
    main()
