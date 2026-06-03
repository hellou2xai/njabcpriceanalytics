#!/usr/bin/env python
"""Load the Allied (ABG) SKU <-> UPC translation table into Postgres.

Source: the monthly "RIPs Combos <month>.xlsx" master from Allied Beverage Group.
It carries Allied's internal item number (ABG SKU) against the product UPC and
brand registration. ABG SKU is the clean key: each SKU maps to exactly one UPC.
The reverse (UPC -> SKU) is one-to-many, so display-time disambiguation is done
later (see backend/enrichment_join.attach_sku_mapping); this script only loads
the raw relationship.

The distributor is stored as the application's own wholesaler code, 'allied',
so it joins straight onto cpl_enriched.wholesaler. The same UPC appears under
other distributors, so consumers MUST gate on wholesaler = 'allied'.

Usage:
    python scripts/load_sku_mapping.py
    python scripts/load_sku_mapping.py --xlsx "Master Data Files/RIPs Combos April 2026.xlsx"
    python scripts/load_sku_mapping.py --database-url "postgresql://user:pw@host/db?sslmode=require"

Defaults to DATABASE_URL from the environment / .env. After it runs, trigger
POST /api/admin/reload-pricing (or redeploy) so the cache picks up the table.
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import psycopg

from backend.pg import DATABASE_URL as ENV_DATABASE_URL

DISTRIBUTOR = "allied"  # the app's wholesaler code for Allied Beverage Group


def _norm(x) -> str:
    return "" if x is None else str(x).strip()


def _upc_norm(u: str) -> str:
    """Match the catalogue join key: LTRIM(upc, '0')."""
    return u.lstrip("0")


def read_rows(xlsx_path: Path):
    """Return {abg_sku: (upc, upc_norm, brand_reg, item_name)} from both sheets.

    ABG SKU is unique across the file, so one row per SKU. RIPs is read first so
    its RIP DESCRIPTION wins as the item name when a SKU is in both sheets.
    Sheet column order is fixed by Allied's template.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: dict[str, tuple] = {}

    def take(sheet_name, sku_i, upc_i, brand_i, name_i):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # header
                continue
            sku = _norm(r[sku_i]); upc = _norm(r[upc_i])
            brand = _norm(r[brand_i]); name = _norm(r[name_i])
            if not sku:
                continue
            out.setdefault(sku, (upc, _upc_norm(upc), brand, name))

    # RIPs:   ABG SKU, RIP CODE, UPC CODE, BRAND, ..., RIP DESCRIPTION(6)
    take("RIPs April 2026", 0, 2, 3, 6)
    # Combos: ABG SKU, Combo Code, UPC Code, ..., Individ Prods(5), Brand Registration(6)
    take("Combos April 2026", 0, 2, 6, 5)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="Master Data Files/RIPs Combos April 2026.xlsx")
    ap.add_argument("--database-url", default=ENV_DATABASE_URL,
                    help="Target Postgres URL (defaults to env DATABASE_URL)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = Path(__file__).resolve().parent.parent / xlsx_path
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")

    rows = read_rows(xlsx_path)
    print(f"Read {len(rows)} distinct ABG SKUs from {xlsx_path.name}")
    valid = sum(1 for (_u, un, _b, _n) in rows.values() if len(un) >= 6)
    print(f"  with a join-usable UPC (>=6 digits): {valid}")

    db = args.database_url
    print(f"Writing sku_mapping into {db.split('@')[-1]}")
    with psycopg.connect(db) as con:
        con.execute("DROP TABLE IF EXISTS sku_mapping")
        con.execute(
            """CREATE TABLE sku_mapping (
                distributor text NOT NULL,
                abg_sku     text NOT NULL,
                upc         text,
                upc_norm    text,
                brand_reg   text,
                item_name   text,
                PRIMARY KEY (distributor, abg_sku)
            )"""
        )
        with con.cursor() as cur:
            cur.executemany(
                "INSERT INTO sku_mapping (distributor, abg_sku, upc, upc_norm, brand_reg, item_name) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                [(DISTRIBUTOR, sku, upc, un, brand, name)
                 for sku, (upc, un, brand, name) in rows.items()],
            )
        con.execute(
            "CREATE INDEX idx_sku_mapping_upc ON sku_mapping (distributor, upc_norm)"
        )
        con.commit()
        n = con.execute("SELECT count(*) FROM sku_mapping").fetchone()[0]
    print(f"Done. {n} rows in sku_mapping. "
          "Trigger /api/admin/reload-pricing or redeploy to refresh the cache.")


if __name__ == "__main__":
    main()
