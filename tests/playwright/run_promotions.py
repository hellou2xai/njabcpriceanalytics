"""
Promotions pages: end-to-end UI vs API vs Postgres validation.

Two stages per record:
  Stage A (scrape):  Playwright opens each Promotions page, iterates every
                     card, opens the detail modal, extracts every visible
                     field, then closes the modal.
  Stage B (validate): For each scraped record, fetch the canonical truth from
                     (1) the JSON API and (2) Postgres, and produce a row of
                     pass/fail verdicts for the spreadsheet.

Output: one openpyxl workbook with one sheet per page plus a Summary sheet,
written under tests/playwright/results/.
"""
import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from playwright.sync_api import Page, TimeoutError as PWTimeout, sync_playwright
import psycopg
from psycopg.rows import dict_row
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

from dotenv import load_dotenv
load_dotenv(REPO_ROOT / ".env")

API_BASE = os.getenv("API_BASE", "http://127.0.0.1:8000")
FRONTEND_BASE = os.getenv("FRONTEND_BASE", "http://localhost:5173")
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://celr:celrdev@localhost:5432/celr_dev")
DEFAULT_EMAIL = "sambit.tripathy@gmail.com"
DEFAULT_PASSWORD = "Cuttack10!"

PRICE_TOL = 0.02
PCT_TOL = 1.0  # UI rounds discount % to nearest integer, API keeps 2 decimals

PAGES = {
    "time-sensitive": {
        "route": "/time-sensitive",
        "endpoint": "/api/deals/time-sensitive",
        "params": {"limit": 20000},
        "sheet": "Time-Sensitive",
    },
    "major-discounts": {
        "route": "/major-discounts",
        "endpoint": "/api/deals/discounts",
        "params": {"limit": 1000, "sort": "total_savings_per_case"},
        "sheet": "Major-Discounts",
    },
    "price-drops": {
        "route": "/price-drops",
        "endpoint": "/api/analytics/price-movers",
        "params": {"direction": "down", "limit": 5000, "validity": "all"},
        "sheet": "Price-Drops",
    },
    "price-increases": {
        "route": "/price-increases",
        "endpoint": "/api/analytics/price-movers",
        "params": {"direction": "up", "limit": 5000, "validity": "all"},
        "sheet": "Price-Increases",
    },
}


# ---------------------------------------------------------------------------
# Stage A: scraping
# ---------------------------------------------------------------------------

@dataclass
class CardData:
    page: str
    sample_index: int
    wholesaler: str = ""
    upc: str = ""
    product_name: str = ""
    unit_volume: str = ""
    unit_qty: str = ""
    vintage: str = ""
    edition: str = ""
    card_now_price: float | None = None
    card_was_price: float | None = None
    card_save_amount: float | None = None
    card_discount_pct: float | None = None
    card_has_rip_badge: bool = False
    card_has_closeout_badge: bool = False
    card_text_dump: str = ""
    modal_opened: bool = False
    modal_list_price: float | None = None
    modal_effective_case: float | None = None
    modal_effective_bottle: float | None = None
    modal_discount_tier_count: int = 0
    modal_rip_tier_count: int = 0
    modal_text_dump: str = ""
    scrape_error: str = ""


def login(api_base: str, email: str, password: str) -> tuple[str, dict]:
    r = requests.post(
        f"{api_base}/api/auth/login",
        json={"email": email, "password": password},
        timeout=15,
    )
    r.raise_for_status()
    body = r.json()
    return body["token"], body["user"]


def _seed_local_storage(page: Page, token: str, user: dict) -> None:
    user_json = json.dumps(user)
    script = (
        "() => {"
        f"  localStorage.setItem('lpb_auth_token', {json.dumps(token)});"
        f"  localStorage.setItem('lpb_auth_user', {json.dumps(user_json)});"
        "}"
    )
    page.evaluate(script)


def _parse_money(text: str) -> float | None:
    if not text:
        return None
    m = re.search(r"\$([0-9][0-9,]*\.?\d*)", text)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except ValueError:
        return None


def _parse_pct(text: str) -> float | None:
    if not text:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def _parse_vintage(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"Vintage\s+(\S+)", text)
    return m.group(1) if m else ""


def _safe_text(locator) -> str:
    try:
        return (locator.text_content(timeout=500) or "").strip()
    except PWTimeout:
        return ""
    except Exception:
        return ""


def _extract_card_fields(card_el, data: CardData) -> None:
    data.wholesaler = card_el.get_attribute("data-ctx-wholesaler") or ""
    data.upc = (card_el.get_attribute("data-ctx-upc") or "").lstrip("0")
    data.product_name = card_el.get_attribute("data-ctx-product") or ""
    data.unit_volume = card_el.get_attribute("data-ctx-volume") or ""
    sub_text = _safe_text(card_el.locator(".deal-card-sub").first)
    data.vintage = _parse_vintage(sub_text)
    data.card_now_price = _parse_money(_safe_text(card_el.locator(".deal-now").first))
    data.card_was_price = _parse_money(_safe_text(card_el.locator(".deal-was").first))
    save_text = _safe_text(card_el.locator(".deal-save").first)
    data.card_save_amount = _parse_money(save_text)
    data.card_discount_pct = _parse_pct(save_text)
    data.card_has_rip_badge = card_el.locator(".source-badge.source-rip").count() > 0
    data.card_has_closeout_badge = card_el.locator(".tag.tag-orange").count() > 0
    data.card_text_dump = (_safe_text(card_el) or "")[:500]


def _extract_modal_fields(page: Page, data: CardData) -> None:
    try:
        overlay = page.locator(".modal-overlay").first
        overlay.wait_for(state="visible", timeout=6000)
        modal = overlay.locator(".modal").first
        modal.wait_for(state="visible", timeout=3000)
    except PWTimeout:
        data.scrape_error = (data.scrape_error + " | modal did not appear").strip(" |")
        return
    data.modal_opened = True
    # The PriceBreakdown table is rendered after the async product-detail
    # fetch; wait briefly so we don't read a loading-state empty modal.
    try:
        modal.locator(".pb-table").first.wait_for(state="visible", timeout=5000)
    except PWTimeout:
        data.scrape_error = (data.scrape_error + " | pb-table not rendered").strip(" |")
    # The price-movers modal loads three months in parallel (prev/this/next),
    # and the table renders progressively. Wait until the "This month" column
    # header is present so we don't grab a transient single-column view.
    try:
        modal.locator(".pb-table thead th", has_text=re.compile(r"This month", re.IGNORECASE)).first.wait_for(state="visible", timeout=5000)
    except PWTimeout:
        # Single-month products don't get a "This month" header — that's OK,
        # extraction falls back to column 1 which is the only value column.
        pass

    modal_text = _safe_text(modal)
    data.modal_text_dump = modal_text[:1500]

    # The PriceBreakdown is a multi-column table when prev/next editions exist:
    # column 0 = label, columns 1..N = one per edition ("Last month", "This
    # month", "Next month"). Pick the "This month" column when present, else
    # column 1 (single-side view).
    headers = []
    try:
        headers = modal.locator(".pb-table thead th").all_text_contents()
    except Exception:
        headers = []
    target_col = None
    for i, h in enumerate(headers):
        if "This month" in (h or ""):
            target_col = i
            break
    if target_col is None:
        target_col = 1 if len(headers) >= 2 else None

    if target_col is not None:
        # Find rows by label rather than position; the table may have section
        # headers, fallback rows, or single-column views that shift indexing.
        try:
            list_row = modal.locator(".pb-table tbody tr").filter(
                has_text=re.compile(r"List price", re.IGNORECASE)
            ).first
            list_cell = list_row.locator("td").nth(target_col)
            data.modal_list_price = _parse_money(_safe_text(list_cell))
        except Exception:
            pass
        try:
            eff_row = modal.locator(".pb-table tbody tr.pb-total").first
            eff_cell = eff_row.locator("td").nth(target_col)
            data.modal_effective_case = _parse_money(_safe_text(eff_cell))
        except Exception:
            pass

    try:
        data.modal_discount_tier_count = modal.locator("tr.pb-tier-row").filter(
            has_text=re.compile(r"\+\s*(cs|btl)", re.IGNORECASE)
        ).count()
    except Exception:
        data.modal_discount_tier_count = 0
    try:
        data.modal_rip_tier_count = modal.locator("tr.pb-tier-row").filter(
            has_text=re.compile(r"^Buy\s+\d", re.IGNORECASE)
        ).count()
    except Exception:
        data.modal_rip_tier_count = 0


def _open_and_close_modal(page: Page, card_el, data: CardData) -> None:
    try:
        card_el.scroll_into_view_if_needed(timeout=2000)
    except Exception:
        pass
    # Click the card's product name — a stable, button-free hit target.
    try:
        target = card_el.locator(".deal-card-name").first
        if target.count() == 0:
            target = card_el
        target.click(timeout=3000)
    except PWTimeout:
        data.scrape_error = (data.scrape_error + " | click timeout").strip(" |")
        return
    except Exception as e:
        data.scrape_error = (data.scrape_error + f" | click error: {type(e).__name__}").strip(" |")
        return
    _extract_modal_fields(page, data)
    # Close: click the ✕ button (more reliable than Escape, which depends on focus).
    try:
        close_btn = page.locator(".modal-close").first
        if close_btn.count() > 0:
            close_btn.click(timeout=2000)
        else:
            page.keyboard.press("Escape")
        page.locator(".modal-overlay").first.wait_for(state="hidden", timeout=3000)
    except Exception:
        pass


def scrape_page(
    page: Page,
    page_key: str,
    edition: str,
    max_samples: int,
    progress_every: int = 25,
    do_modal: bool = True,
) -> list[CardData]:
    cfg = PAGES[page_key]
    print(f"[scrape] {page_key} -> {FRONTEND_BASE}{cfg['route']}")
    page.goto(f"{FRONTEND_BASE}{cfg['route']}", wait_until="domcontentloaded")
    try:
        page.locator(".deal-cards").first.wait_for(state="visible", timeout=20000)
    except PWTimeout:
        print(f"[scrape] no .deal-cards container on {page_key}")
        return []

    samples: list[CardData] = []
    seen_keys: set[tuple] = set()
    page_idx = 0
    while len(samples) < max_samples:
        try:
            page.locator(".deal-card").first.wait_for(state="visible", timeout=10000)
        except PWTimeout:
            print(f"[scrape] no cards visible on page index {page_idx}")
            break

        cards = page.locator(".deal-card")
        count = cards.count()
        if count == 0:
            break

        for i in range(count):
            if len(samples) >= max_samples:
                break
            data = CardData(page=page_key, sample_index=len(samples), edition=edition)
            card_el = cards.nth(i)
            try:
                _extract_card_fields(card_el, data)
            except Exception as e:
                data.scrape_error = f"extract_card_fields: {type(e).__name__}: {e}"

            key = (data.wholesaler, data.upc, data.vintage, data.unit_qty, data.unit_volume)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            if do_modal:
                _open_and_close_modal(page, card_el, data)
            samples.append(data)
            if len(samples) % progress_every == 0:
                print(f"[scrape] {page_key}: {len(samples)} samples")

        # Try next page via the toolbar's Next button.
        next_btn = page.locator("button[title='Next page']").first
        try:
            disabled = next_btn.get_attribute("disabled") is not None or next_btn.is_disabled()
        except Exception:
            disabled = True
        if disabled:
            break
        try:
            next_btn.click(timeout=2000)
            page.wait_for_timeout(400)
            page_idx += 1
        except Exception:
            break

    print(f"[scrape] {page_key}: collected {len(samples)} samples")
    return samples


# ---------------------------------------------------------------------------
# Stage B: validation
# ---------------------------------------------------------------------------

@dataclass
class ValidationResult:
    page: str = ""
    sample_index: int = 0
    wholesaler: str = ""
    upc: str = ""
    product_name: str = ""
    vintage: str = ""
    unit_qty: str = ""
    unit_volume: str = ""
    edition: str = ""
    api_match_found: bool = False
    # Plain-English columns the reviewer reads first.
    screen_shows: str = ""
    db_says: str = ""
    plain_explanation: str = ""
    # Per-check verdicts (still useful for power-users debugging a failure).
    ui_vs_api_effective: str = ""
    ui_vs_api_frontline: str = ""
    ui_vs_api_save: str = ""
    ui_vs_api_pct: str = ""
    ui_vs_api_rip_flag: str = ""
    api_vs_db_effective: str = ""
    api_vs_db_frontline: str = ""
    api_vs_db_rip_savings: str = ""
    modal_vs_db_list: str = ""
    modal_vs_api_effective: str = ""
    page_specific_rule: str = ""
    overall: str = ""
    failures: list[str] = field(default_factory=list)
    api_payload: dict = field(default_factory=dict)
    db_payload: dict = field(default_factory=dict)
    notes: str = ""


def _close_to(a: float | None, b: float | None, tol: float) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def _verdict(a, b, tol) -> str:
    if a is None and b is None:
        return "n/a"
    if a is None or b is None:
        return f"MISSING ({a} vs {b})"
    return "PASS" if abs(a - b) <= tol else f"FAIL ({a} vs {b})"


def _flag_verdict(a: bool, b: bool) -> str:
    return "PASS" if a == b else f"FAIL ({a} vs {b})"


def _fetch_api_list(page_key: str) -> list[dict]:
    cfg = PAGES[page_key]
    r = requests.get(f"{API_BASE}{cfg['endpoint']}", params=cfg["params"], timeout=60)
    r.raise_for_status()
    body = r.json()
    return body if isinstance(body, list) else body.get("rows", body.get("items", []))


def _index_api(rows: list[dict]) -> dict[tuple, dict]:
    out: dict[tuple, dict] = {}
    for row in rows:
        upc = (row.get("upc") or "").lstrip("0")
        key = (
            row.get("wholesaler", ""),
            upc,
            row.get("vintage") or "",
            str(row.get("unit_qty") or ""),
            row.get("unit_volume") or "",
        )
        out.setdefault(key, row)
    return out


def _api_lookup(api_idx: dict[tuple, dict], data: CardData) -> dict | None:
    """Card has wholesaler+upc+volume from data-ctx; vintage from sub-text.
    unit_qty isn't a card attribute, so match by the fields we know."""
    candidates = [
        cand for cand_key, cand in api_idx.items()
        if cand_key[0] == data.wholesaler and cand_key[1] == data.upc
        and (cand_key[4] == data.unit_volume or not data.unit_volume)
    ]
    if not candidates:
        return None
    if data.vintage:
        for c in candidates:
            if (c.get("vintage") or "") == data.vintage:
                return c
    return candidates[0]


def _norm_int_str(v) -> str:
    """Normalise '6', '6.0', 6, 6.0, '06' -> '6'. Empty for None/blank."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    if not s:
        return ""
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def _db_lookup_by_api(con: psycopg.Connection, api_row: dict) -> dict | None:
    """Look up the DB row matching the API row's canonical key.

    cpl_enriched stores some text columns inconsistently across editions:
    - unit_qty: '6' OR '6.0'
    - vintage: '2018' OR '2018.0' OR '18'
    So we compare each side after normalising to an integer-string ('6'),
    and accept any DB row whose normalised values match the API's."""
    upc = (api_row.get("upc") or "").lstrip("0")
    edition = api_row.get("edition") or ""
    api_unit_qty = _norm_int_str(api_row.get("unit_qty"))
    api_vintage = _norm_int_str(api_row.get("vintage"))
    unit_volume = api_row.get("unit_volume") or ""

    product_name = (api_row.get("product_name") or "").strip()
    sql = """
        SELECT * FROM cpl_enriched
         WHERE wholesaler = %s
           AND LTRIM(upc, '0') = %s
           AND edition = %s
           AND COALESCE(unit_volume, '') = %s
    """
    rows = con.execute(
        sql, (api_row.get("wholesaler", ""), upc, edition, unit_volume),
    ).fetchall()
    if not rows:
        return None

    # Filter by exact product_name first — UPC alone is not unique (UPC '0' is
    # shared across many SKUs, and some legitimate UPCs appear under multiple
    # product names like "GREENALLS DRY GIN" vs "GREENALLS DRY GIN JIGGER 6&6").
    if product_name:
        named = [r for r in rows if (r.get("product_name") or "").strip() == product_name]
        if named:
            rows = named

    candidates = [r for r in rows if _norm_int_str(r.get("unit_qty")) == api_unit_qty]
    if not candidates:
        candidates = rows  # fall back to unit_volume-only match
    if len(candidates) == 1:
        return candidates[0]
    # Tiebreak by vintage after normalisation (handles 2018 / 2018.0 / 18 / 2018).
    if api_vintage:
        for r in candidates:
            if _norm_int_str(r.get("vintage")) == api_vintage:
                return r
    return candidates[0]


def _db_has_rip(con: psycopg.Connection, wholesaler: str, rip_code: str, edition: str) -> bool:
    """Cheap presence check: a RIP row exists in `rip` for this code+edition.

    cpl_enriched.rip_code is often a space-separated list when a SKU qualifies
    for multiple RIP programs at once (e.g. '10404 70001'). Split on whitespace
    so any one of the listed codes counts as present."""
    if not rip_code:
        return False
    codes = [c for c in str(rip_code).split() if c and c != "0"]
    if not codes:
        return False
    placeholders = ",".join(["%s"] * len(codes))
    row = con.execute(
        f"SELECT 1 FROM rip WHERE wholesaler = %s AND edition = %s "
        f"AND rip_code IN ({placeholders}) LIMIT 1",
        (wholesaler, edition, *codes),
    ).fetchone()
    return row is not None


def _page_specific(page_key: str, api_row: dict | None) -> str:
    if not api_row:
        return "n/a"
    if page_key == "time-sensitive":
        d = api_row.get("days_to_expire")
        return "PASS" if isinstance(d, (int, float)) else f"FAIL (days_to_expire={d!r})"
    if page_key == "major-discounts":
        p = api_row.get("discount_pct")
        return "PASS" if isinstance(p, (int, float)) and p > 0 else f"FAIL (discount_pct={p!r})"
    if page_key in ("price-drops", "price-increases"):
        # Card headline uses the headline_period's delta; that's the one the
        # user sees, so that's the one the rule checks.
        headline = api_row.get("headline_period") or "next"
        if headline == "next":
            delta = api_row.get("next_delta")
        else:
            delta = api_row.get("cur_delta") or api_row.get("case_delta")
        if not isinstance(delta, (int, float)):
            return f"FAIL (delta={delta!r})"
        if page_key == "price-drops":
            return "PASS" if delta < 0 else f"FAIL (delta={delta} not negative)"
        return "PASS" if delta > 0 else f"FAIL (delta={delta} not positive)"
    return "n/a"


def _expected_card_now(page_key: str, api_row: dict) -> float | None:
    """What .deal-now (the bold/coloured headline price) should show, per page."""
    if page_key in ("time-sensitive", "major-discounts"):
        return api_row.get("effective_case_price")
    # PriceMovers card defaults to 'next' headline; show next-month price.
    headline = api_row.get("headline_period") or "next"
    if headline == "next":
        return api_row.get("next_case_price") or api_row.get("effective_case_price")
    return api_row.get("case_price")


def _expected_card_was(page_key: str, api_row: dict) -> float | None:
    """What .deal-was (the struck-through 'was' price) should show."""
    if page_key in ("time-sensitive", "major-discounts"):
        return api_row.get("frontline_case_price")
    headline = api_row.get("headline_period") or "next"
    if headline == "next":
        return api_row.get("case_price")
    return api_row.get("prev_case_price")


def _expected_save_and_pct(page_key: str, api_row: dict) -> tuple[float | None, float | None]:
    """Card's .deal-save 'save $X · Y%' element, per page."""
    if page_key in ("time-sensitive", "major-discounts"):
        return api_row.get("total_savings_per_case"), api_row.get("discount_pct")
    # PriceMovers shows delta (signed) and delta-%.
    headline = api_row.get("headline_period") or "next"
    if headline == "next":
        d = api_row.get("next_delta")
        p = api_row.get("next_delta_pct")
    else:
        d = api_row.get("cur_delta") or api_row.get("case_delta")
        p = api_row.get("cur_delta_pct") or api_row.get("case_delta_pct")
    # The card prints |d| via money(); pct printed signed-absolute. Compare magnitudes.
    return (abs(d) if isinstance(d, (int, float)) else None,
            abs(p) if isinstance(p, (int, float)) else None)


def _fmt_money(v) -> str:
    if v is None or (isinstance(v, float) and v != v):
        return "—"
    try:
        return f"${float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(v) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):.1f}%"
    except (TypeError, ValueError):
        return str(v)


def _build_screen_summary(data: CardData) -> str:
    """One-line plain-English description of what the card+modal showed."""
    parts: list[str] = []
    if data.card_now_price is not None:
        parts.append(f"now {_fmt_money(data.card_now_price)}/cs")
    if data.card_was_price is not None:
        parts.append(f"was {_fmt_money(data.card_was_price)}/cs")
    if data.card_save_amount is not None:
        save_part = f"save {_fmt_money(data.card_save_amount)}/cs"
        if data.card_discount_pct is not None:
            save_part += f" ({data.card_discount_pct:.0f}% off)"
        parts.append(save_part)
    if data.card_has_rip_badge:
        parts.append("RIP rebate stacks")
    if data.card_has_closeout_badge:
        parts.append("Closeout")
    head = " · ".join(parts) if parts else "card showed no price"
    if data.modal_opened and data.modal_list_price is not None:
        head += (
            f"  ||  Modal: list {_fmt_money(data.modal_list_price)}/cs"
            f", final {_fmt_money(data.modal_effective_case)}/cs"
            f", {data.modal_discount_tier_count} CPL tier(s), {data.modal_rip_tier_count} RIP tier(s)"
        )
    elif not data.modal_opened:
        head += "  ||  Modal: did not open"
    return head


def _build_db_summary(db_row: dict | None, rip_present: bool, api_row: dict | None) -> str:
    """One-line description of what the DB / canonical source says."""
    if not db_row:
        return "no matching DB row"
    front = db_row.get("frontline_case_price")
    best = db_row.get("best_case_price")
    parts = [f"list {_fmt_money(front)}/cs"]
    if best is not None and front is not None and float(best) < float(front) - 0.01:
        parts.append(f"best after CPL discount {_fmt_money(best)}/cs")
    else:
        parts.append("no CPL discount")
    rip_sav = api_row.get("rip_savings") if api_row else None
    if rip_present and rip_sav and rip_sav > 0:
        parts.append(f"RIP rebate −{_fmt_money(rip_sav)}/cs")
    elif rip_present:
        parts.append("RIP rebate present")
    else:
        parts.append("no RIP rebate")
    if (db_row.get("closeout_permit") or "") not in ("", None, "0"):
        parts.append("Closeout permit on file")
    expected_eff = None
    if best is not None:
        expected_eff = float(best) - float(rip_sav or 0)
    if expected_eff is not None:
        parts.append(f"expected final {_fmt_money(expected_eff)}/cs")
    pct = None
    if front and front > 0 and expected_eff is not None:
        pct = (float(front) - expected_eff) / float(front) * 100
    if pct is not None and pct > 0.5:
        parts.append(f"total saving {_fmt_pct(pct)}")
    return " · ".join(parts)


def _build_plain_explanation(
    data: CardData,
    api_row: dict | None,
    db_row: dict | None,
    rip_present: bool,
    expected_now: float | None,
    expected_was: float | None,
) -> tuple[str, str]:
    """Return (one-sentence verdict, next-step hint)."""
    if not api_row:
        return ("No API row matched the card — backend may have dropped it.",
                "Re-check the card's UPC/wholesaler against the endpoint response.")
    if not db_row:
        return ("Card's product was not found in the cpl_enriched DB table.",
                "Confirm Postgres restore is current; check vintage/unit_qty mapping.")

    front_db = db_row.get("frontline_case_price")
    best_db = db_row.get("best_case_price")
    rip_sav = api_row.get("rip_savings") if api_row else None
    expected_eff_db = (float(best_db) - float(rip_sav or 0)) if best_db is not None else None

    problems: list[str] = []
    is_mover = data.page in ("price-drops", "price-increases")

    # 1. Effective price on screen matches the page's expected value.
    #    On time-sensitive / major-discounts that's DB's current-month derived
    #    effective. On price-drops/increases the card shows the *next-month*
    #    price (next_case_price), so the comparison is API-side.
    if data.card_now_price is not None:
        if is_mover:
            if expected_now is not None and abs(data.card_now_price - expected_now) > PRICE_TOL:
                problems.append(
                    f"Screen 'now' shows {_fmt_money(data.card_now_price)} but the next-month price"
                    f" the API reports is {_fmt_money(expected_now)}"
                )
        else:
            if expected_eff_db is not None and abs(data.card_now_price - expected_eff_db) > PRICE_TOL:
                cpl_amt = (front_db or 0) - (best_db or front_db or 0)
                problems.append(
                    f"Screen shows {_fmt_money(data.card_now_price)} but DB computes {_fmt_money(expected_eff_db)}"
                    f" (list {_fmt_money(front_db)} − CPL discount {_fmt_money(cpl_amt)}"
                    f" − RIP {_fmt_money(rip_sav or 0)})"
                )

    # 2. 'Was' price on screen vs the page's expected was-value.
    if data.card_was_price is not None:
        if is_mover:
            if expected_was is not None and abs(data.card_was_price - expected_was) > PRICE_TOL:
                problems.append(
                    f"Screen 'was' {_fmt_money(data.card_was_price)} but the API's prior-period price"
                    f" is {_fmt_money(expected_was)}"
                )
        else:
            if front_db is not None and abs(data.card_was_price - front_db) > PRICE_TOL:
                problems.append(
                    f"Screen lists 'was' {_fmt_money(data.card_was_price)} but DB list is {_fmt_money(front_db)}"
                )

    # 3. RIP badge on screen vs DB.
    if data.card_has_rip_badge != bool(rip_present):
        problems.append(
            f"Screen RIP badge {'ON' if data.card_has_rip_badge else 'OFF'},"
            f" DB rip row {'EXISTS' if rip_present else 'MISSING'}"
        )

    # 4. Modal list price vs DB list.
    #    On price-mover pages the modal's "This month" column corresponds to
    #    the calendar-current edition, while api_row.edition is the destination
    #    edition (= one month ahead). DB front for the destination edition is
    #    usually identical to the prior edition's list, but if it differs we'd
    #    incorrectly flag a real-life list change as a UI bug. Compare against
    #    the API-side frontline ('frontline_case_price' is keyed to api_row's
    #    edition; the prior edition's list would be 'frontline_prev_case_price').
    if data.modal_opened and data.modal_list_price is not None:
        if is_mover:
            expected_modal_list = api_row.get("frontline_prev_case_price") or front_db
        else:
            expected_modal_list = front_db
        if (expected_modal_list is not None
                and abs(data.modal_list_price - expected_modal_list) > PRICE_TOL):
            problems.append(
                f"Modal list {_fmt_money(data.modal_list_price)} ≠ DB list {_fmt_money(expected_modal_list)}"
            )

    # 5. Modal "Price after RIP" vs the expected calendar-current effective.
    #    For price-movers, "This month" in the modal = api_row.case_price
    #    (the source-of-truth for the current edition's effective). For other
    #    pages, it = expected_eff_db (the row's own edition is current).
    if data.modal_opened and data.modal_effective_case is not None:
        expected_modal_eff = api_row.get("case_price") if is_mover else expected_eff_db
        if (expected_modal_eff is not None
                and abs(data.modal_effective_case - expected_modal_eff) > PRICE_TOL):
            problems.append(
                f"Modal final price {_fmt_money(data.modal_effective_case)} ≠ this-month effective {_fmt_money(expected_modal_eff)}"
            )

    if not problems:
        # Mention deltas for the price-mover pages so it's clear the rule fired.
        if data.page in ("price-drops", "price-increases"):
            headline = api_row.get("headline_period") or "next"
            delta = api_row.get("next_delta") if headline == "next" else api_row.get("cur_delta")
            if delta is not None:
                dir_word = "drop" if delta < 0 else "rise"
                return (
                    f"Match. Card correctly shows a {dir_word} of {_fmt_money(abs(delta))}/cs"
                    f" from {_fmt_money(_expected_card_was(data.page, api_row))} to {_fmt_money(expected_now)}.",
                    "",
                )
        return (
            f"Match. Screen final price {_fmt_money(data.card_now_price)}/cs equals DB "
            f"({_fmt_money(front_db)} list − discounts/RIP = {_fmt_money(expected_eff_db)}).",
            "",
        )

    verdict = "Mismatch. " + ". ".join(problems) + "."
    return (verdict, "Manually compare the card UI to cpl_enriched + rip rows for this UPC+edition.")


def validate(samples_by_page: dict[str, list[CardData]], db_url: str) -> dict[str, list[ValidationResult]]:
    api_idx_by_page = {pk: _index_api(_fetch_api_list(pk)) for pk in samples_by_page}
    results: dict[str, list[ValidationResult]] = {pk: [] for pk in samples_by_page}
    with psycopg.connect(db_url, row_factory=dict_row) as con:
        for page_key, samples in samples_by_page.items():
            api_idx = api_idx_by_page[page_key]
            for data in samples:
                res = ValidationResult(
                    page=page_key,
                    sample_index=data.sample_index,
                    wholesaler=data.wholesaler,
                    upc=data.upc,
                    product_name=data.product_name,
                    vintage=data.vintage,
                    unit_qty=data.unit_qty,
                    unit_volume=data.unit_volume,
                    edition=data.edition,
                )
                api_row = _api_lookup(api_idx, data)
                db_row = _db_lookup_by_api(con, api_row) if api_row else None
                if api_row:
                    res.api_match_found = True
                    # Fill in unit_qty/edition from API since the card lacks them.
                    if not res.unit_qty:
                        res.unit_qty = str(api_row.get("unit_qty") or "")
                    if not res.edition:
                        res.edition = api_row.get("edition") or res.edition
                    res.api_payload = {k: api_row.get(k) for k in (
                        "effective_case_price", "frontline_case_price",
                        "total_savings_per_case", "discount_pct", "rip_savings",
                        "has_rip", "has_closeout", "case_price", "prev_case_price",
                        "next_case_price", "next_delta", "next_delta_pct",
                        "cur_delta", "cur_delta_pct", "headline_period",
                        "days_to_expire",
                    ) if k in api_row}
                if db_row:
                    res.db_payload = {k: db_row.get(k) for k in (
                        "frontline_case_price", "best_case_price",
                        "discount_1_amt", "discount_2_amt", "discount_3_amt",
                        "rip_code", "closeout_permit", "from_date", "to_date",
                    )}

                expected_now = _expected_card_now(page_key, api_row) if api_row else None
                expected_was = _expected_card_was(page_key, api_row) if api_row else None
                expected_save, expected_pct = (_expected_save_and_pct(page_key, api_row) if api_row else (None, None))
                rip_flag_api = bool(api_row.get("has_rip")) if api_row else False

                res.ui_vs_api_effective = _verdict(data.card_now_price, expected_now, PRICE_TOL)
                res.ui_vs_api_frontline = (
                    _verdict(data.card_was_price, expected_was, PRICE_TOL)
                    if data.card_was_price is not None else "n/a"
                )
                res.ui_vs_api_save = (
                    _verdict(data.card_save_amount, expected_save, PRICE_TOL)
                    if data.card_save_amount is not None and expected_save is not None else "n/a"
                )
                res.ui_vs_api_pct = (
                    _verdict(data.card_discount_pct, expected_pct, PCT_TOL)
                    if data.card_discount_pct is not None and expected_pct is not None else "n/a"
                )
                res.ui_vs_api_rip_flag = _flag_verdict(data.card_has_rip_badge, rip_flag_api)

                front_db = db_row.get("frontline_case_price") if db_row else None
                best_db = db_row.get("best_case_price") if db_row else None
                front_api = api_row.get("frontline_case_price") if api_row else None
                eff_api = api_row.get("effective_case_price") if api_row else None
                rip_savings_api = api_row.get("rip_savings") if api_row else None

                res.api_vs_db_frontline = _verdict(front_api, front_db, PRICE_TOL)
                # Effective vs DB: API.effective should equal DB.best - rip_savings.
                # When has_rip=true but the endpoint hides rip_savings (price-movers
                # does), the derived comparison is ambiguous; mark n/a instead of
                # failing the row on a known endpoint-shape limitation.
                if best_db is None:
                    res.api_vs_db_effective = "n/a"
                elif rip_flag_api and rip_savings_api is None:
                    res.api_vs_db_effective = "n/a (rip_savings hidden)"
                else:
                    expected_eff_db = float(best_db) - float(rip_savings_api or 0)
                    res.api_vs_db_effective = _verdict(eff_api, expected_eff_db, PRICE_TOL)
                db_has_rip = _db_has_rip(con, data.wholesaler, (db_row or {}).get("rip_code") or "", res.edition)
                res.api_vs_db_rip_savings = _flag_verdict(rip_flag_api, db_has_rip)

                # For price-mover pages the modal's "This month" column shows
                # the current calendar edition (one before api_row.edition).
                # Compare against api_row.case_price / prior-edition list to
                # avoid false fails when the destination edition has different
                # numbers than the prior (this-month) one.
                is_mover_page = page_key in ("price-drops", "price-increases")
                if data.modal_opened:
                    modal_list_expected = (
                        api_row.get("frontline_prev_case_price") if (is_mover_page and api_row) else front_db
                    )
                    modal_eff_expected = (
                        api_row.get("case_price") if (is_mover_page and api_row) else eff_api
                    )
                    res.modal_vs_db_list = _verdict(data.modal_list_price, modal_list_expected, PRICE_TOL)
                    res.modal_vs_api_effective = _verdict(data.modal_effective_case, modal_eff_expected, PRICE_TOL)
                else:
                    res.modal_vs_db_list = "n/a (no modal)"
                    res.modal_vs_api_effective = "n/a (no modal)"

                res.page_specific_rule = _page_specific(page_key, api_row)

                fail_fields: list[str] = []
                for field_name in (
                    "ui_vs_api_effective", "ui_vs_api_frontline", "ui_vs_api_save",
                    "ui_vs_api_pct", "ui_vs_api_rip_flag",
                    "api_vs_db_effective", "api_vs_db_frontline", "api_vs_db_rip_savings",
                    "modal_vs_db_list", "modal_vs_api_effective",
                    "page_specific_rule",
                ):
                    v = getattr(res, field_name)
                    if isinstance(v, str) and v.startswith("FAIL"):
                        fail_fields.append(field_name)
                res.failures = fail_fields
                if not res.api_match_found:
                    res.overall = "FAIL (no API match)"
                elif fail_fields:
                    res.overall = "FAIL"
                else:
                    res.overall = "PASS"

                # Plain-English columns.
                res.screen_shows = _build_screen_summary(data)
                res.db_says = _build_db_summary(db_row, db_has_rip, api_row)
                explanation, next_step = _build_plain_explanation(
                    data, api_row, db_row, db_has_rip, expected_now, expected_was,
                )
                res.plain_explanation = explanation
                if not data.modal_opened:
                    res.notes = "modal not extracted"
                elif next_step:
                    res.notes = next_step
                elif fail_fields:
                    res.notes = "investigate: " + ",".join(fail_fields)

                results[page_key].append(res)
    return results


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill("solid", fgColor="DCFCE7")
FAIL_FILL = PatternFill("solid", fgColor="FECACA")
NEUTRAL_FILL = PatternFill("solid", fgColor="F3F4F6")

SHEET_COLUMNS = [
    "Sample #", "Overall", "Plain explanation",
    "What screen shows", "What DB says",
    "Product", "UPC", "Wholesaler", "Vintage", "Unit Qty", "Unit Volume", "Edition",
    "API match",
    "UI vs API: effective", "UI vs API: frontline",
    "UI vs API: save $", "UI vs API: %", "UI vs API: RIP flag",
    "API vs DB: frontline", "API vs DB: effective", "API vs DB: rip presence",
    "Modal vs DB: list", "Modal vs API: effective",
    "Page-specific rule",
    "Failure fields", "Next step",
    "API payload", "DB payload",
]


def _value_for_col(res: ValidationResult, col: str) -> Any:
    mapping = {
        "Sample #": res.sample_index,
        "Plain explanation": res.plain_explanation,
        "What screen shows": res.screen_shows,
        "What DB says": res.db_says,
        "Wholesaler": res.wholesaler,
        "UPC": res.upc,
        "Product": res.product_name,
        "Vintage": res.vintage,
        "Unit Qty": res.unit_qty,
        "Unit Volume": res.unit_volume,
        "Edition": res.edition,
        "API match": "yes" if res.api_match_found else "no",
        "Overall": res.overall,
        "UI vs API: effective": res.ui_vs_api_effective,
        "UI vs API: frontline": res.ui_vs_api_frontline,
        "UI vs API: save $": res.ui_vs_api_save,
        "UI vs API: %": res.ui_vs_api_pct,
        "UI vs API: RIP flag": res.ui_vs_api_rip_flag,
        "API vs DB: frontline": res.api_vs_db_frontline,
        "API vs DB: effective": res.api_vs_db_effective,
        "API vs DB: rip presence": res.api_vs_db_rip_savings,
        "Modal vs DB: list": res.modal_vs_db_list,
        "Modal vs API: effective": res.modal_vs_api_effective,
        "Page-specific rule": res.page_specific_rule,
        "Failure fields": ", ".join(res.failures),
        "Next step": res.notes,
        "API payload": json.dumps(res.api_payload, default=str),
        "DB payload": json.dumps(res.db_payload, default=str),
    }
    return mapping[col]


def _verdict_fill(value: Any) -> PatternFill | None:
    if not isinstance(value, str):
        return None
    if value == "PASS":
        return PASS_FILL
    if value.startswith("FAIL"):
        return FAIL_FILL
    if value in ("n/a", "n/a (no modal)"):
        return NEUTRAL_FILL
    return None


def _safe_save(wb: Workbook, out_path: Path) -> Path:
    """Save wb to out_path; if that file is locked (open in Excel / held by
    OneDrive sync), fall back to a sibling -alt-<n>.xlsx so the run never
    crashes on a transient file lock. Returns the path actually written."""
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        for i in range(1, 12):
            alt = out_path.with_name(f"{out_path.stem}-alt-{i}{out_path.suffix}")
            try:
                wb.save(alt)
                print(f"[warn] {out_path.name} was locked; wrote {alt.name} instead")
                return alt
            except PermissionError:
                continue
        raise


def write_workbook(out_path: Path, results: dict[str, list[ValidationResult]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows: list[list[Any]] = []
    for page_key, rows in results.items():
        sheet_name = PAGES[page_key]["sheet"]
        ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(SHEET_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for row_idx, res in enumerate(rows, start=2):
            for col_idx, header in enumerate(SHEET_COLUMNS, start=1):
                value = _value_for_col(res, header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                fill = _verdict_fill(value)
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        ws.freeze_panes = "D2"
        widths = [
            8,   # Sample #
            10,  # Overall
            70,  # Plain explanation
            70,  # What screen shows
            70,  # What DB says
            32,  # Product
            14,  # UPC
            12,  # Wholesaler
            8,   # Vintage
            8,   # Unit Qty
            10,  # Unit Volume
            9,   # Edition
            10,  # API match
            20,  # UI vs API: effective
            20,  # UI vs API: frontline
            18,  # UI vs API: save $
            16,  # UI vs API: %
            20,  # UI vs API: RIP flag
            22,  # API vs DB: frontline
            22,  # API vs DB: effective
            22,  # API vs DB: rip presence
            22,  # Modal vs DB: list
            24,  # Modal vs API: effective
            24,  # Page-specific rule
            30,  # Failure fields
            32,  # Next step
            60,  # API payload
            60,  # DB payload
        ]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Wrap text on the three plain-English columns so reviewers can read them.
        for col_letter in ("C", "D", "E"):
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row_idx}"].alignment = Alignment(
                    vertical="top", wrap_text=True,
                )
        # Bigger row height so the wrapped text is visible.
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 60

        passes = sum(1 for r in rows if r.overall == "PASS")
        fails = sum(1 for r in rows if r.overall.startswith("FAIL"))
        no_api = sum(1 for r in rows if not r.api_match_found)
        no_modal = sum(1 for r in rows if r.notes == "modal not extracted")
        summary_rows.append([sheet_name, len(rows), passes, fails, no_api, no_modal])

    ws = wb.create_sheet("Summary", 0)
    headers = ["Page", "Samples", "Pass", "Fail", "No API match", "Modal not extracted"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r_idx, row in enumerate(summary_rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.column_dimensions["A"].width = 22
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    _safe_save(wb, out_path)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def detect_edition(api_base: str) -> str:
    try:
        r = requests.get(f"{api_base}/api/deals/discounts", params={"limit": 1}, timeout=15)
        r.raise_for_status()
        rows = r.json()
        if rows:
            return rows[0].get("edition") or ""
    except Exception:
        pass
    return datetime.now().strftime("%Y-%m")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pages", default="all",
                    help="comma-separated subset of: " + ",".join(PAGES) + " (default: all)")
    ap.add_argument("--limit", type=int, default=10000, help="max samples per page (default: 10000, effectively all)")
    ap.add_argument("--headless", action="store_true", default=True)
    ap.add_argument("--headed", dest="headless", action="store_false")
    ap.add_argument("--no-modal", dest="modal", action="store_false",
                    help="skip clicking each card to open its modal (5x faster)")
    ap.add_argument("--output", default=None, help="output .xlsx path (default: timestamped)")
    ap.add_argument("--email", default=DEFAULT_EMAIL)
    ap.add_argument("--password", default=DEFAULT_PASSWORD)
    ap.add_argument("--checkpoint-every", type=int, default=200,
                    help="rewrite the xlsx after every N validated samples (default: 200)")
    args = ap.parse_args()

    if args.pages == "all":
        page_keys = list(PAGES.keys())
    else:
        page_keys = [p.strip() for p in args.pages.split(",") if p.strip() in PAGES]
        if not page_keys:
            print("No valid pages requested.", file=sys.stderr)
            sys.exit(2)

    out_dir = Path(__file__).parent / "results"
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(args.output) if args.output else (out_dir / f"promotions_{stamp}.xlsx")

    print(f"[setup] API={API_BASE} FE={FRONTEND_BASE} pages={page_keys} limit={args.limit}")
    token, user = login(API_BASE, args.email, args.password)
    print(f"[setup] logged in as {user.get('email')} (admin={user.get('is_admin')})")

    edition = detect_edition(API_BASE)
    print(f"[setup] edition={edition}")

    samples: dict[str, list[CardData]] = {}
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=args.headless)
        try:
            context = browser.new_context(viewport={"width": 1500, "height": 1000})
            # Seed localStorage before any page script runs. Also suppress the
            # one-time welcome-tour modal, which otherwise overlays every page
            # and intercepts our card clicks.
            init = (
                f"localStorage.setItem('lpb_auth_token', {json.dumps(token)});"
                f"localStorage.setItem('lpb_auth_user', {json.dumps(json.dumps(user))});"
                "localStorage.setItem('celr_welcome_tour_never', '1');"
            )
            context.add_init_script(init)
            page = context.new_page()
            page.set_default_timeout(8000)
            for pk in page_keys:
                start = time.time()
                samples[pk] = scrape_page(page, pk, edition, args.limit, do_modal=args.modal)
                print(f"[scrape] {pk} done in {time.time()-start:.1f}s")
                # Per-page checkpoint: validate what we have so far and rewrite
                # the workbook so a crash during the next page doesn't lose work.
                try:
                    partial = validate({k: samples[k] for k in samples}, DATABASE_URL)
                    write_workbook(out_path, partial)
                    print(f"[checkpoint] wrote {out_path}")
                except Exception as e:
                    print(f"[checkpoint] skipped: {e}")
            context.close()
        finally:
            browser.close()

    print("[validate] final pass DB + API comparisons")
    results = validate(samples, DATABASE_URL)

    print(f"[report] writing {out_path}")
    write_workbook(out_path, results)

    for pk, rows in results.items():
        passes = sum(1 for r in rows if r.overall == "PASS")
        fails = sum(1 for r in rows if r.overall.startswith("FAIL"))
        print(f"[summary] {pk}: {len(rows)} samples, pass={passes}, fail={fails}")
    print(f"[done] {out_path}")


if __name__ == "__main__":
    main()
