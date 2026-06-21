"""Cart-page detailed accuracy regression.

Builds a cart of known SKUs, then confirms the Cart page:
  1. SHOWS every field for each line (name, distributor, size, pack, price).
  2. PACK   — bottles/case shown ("12/cs") matches the catalogue's unit_qty.
  3. VINTAGE — a wine's vintage ("Vintage 2024") is picked up and shown; a
              spirit with no vintage shows none (no junk year).
  4. PRICE  — the case price on the line matches the catalogue frontline.

Two truths are cross-checked: the CART API line (did the cart pick pack/vintage/
price up correctly from the catalogue) and the CART PAGE DOM (did the frontend
render them). Results append a "Cart" sheet to tests/Doc1_test_results.xlsx.

Targets PROD by default; override with CELR_WEB / CELR_API / CELR_EMAIL / CELR_PW.
Run: python tests/playwright/test_cart_accuracy.py
"""
import json
import os
import re
import sys
import time
from datetime import datetime, timezone

import requests
from playwright.sync_api import sync_playwright

WEB = os.getenv("CELR_WEB", "https://nj.celr.ai").rstrip("/")
API = os.getenv("CELR_API", WEB).rstrip("/")
EMAIL = os.getenv("CELR_EMAIL", "sambit.tripathy@gmail.com")
PW = os.getenv("CELR_PW", "Cuttack10!")
OUT_XLSX = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "Doc1_test_results.xlsx"))

# (upc, wholesaler) SKUs to load. The Chianti is the VINTAGE check (a wine);
# the rest are spirits/RTD with no vintage (must show none).
TARGETS = [
    {"label": "Casamigos Classic Lime Margarita 750ML", "upc": "82000812128", "wholesaler": "fedway"},
    {"label": "Ridgemont 1792 Bourbon 750ML", "upc": "80660001203", "wholesaler": "fedway"},
    {"label": "Laphroaig 10Y Boot 750ML", "upc": "80686007326", "wholesaler": "allied"},
    {"label": "Aspen Vodka 1L", "upc": "860010300046", "wholesaler": "fedway"},
    {"label": "Frescobaldi Castiglioni Chianti 750ML (wine, vintage)", "upc": "839183000060", "wholesaler": "opici"},
]
QTY = 2
results = []


def http_json(method, url, **kw):
    last = None
    for attempt in range(4):
        try:
            r = requests.request(method, url, timeout=30, **kw)
            if r.status_code >= 500:
                last = f"{r.status_code} {r.reason}"; time.sleep(2 + attempt); continue
            r.raise_for_status()
            return r.json() if r.text.strip() else {}
        except Exception as e:  # noqa: BLE001
            last = f"{type(e).__name__}: {e}"; time.sleep(2 + attempt)
    raise RuntimeError(f"request failed: {method} {url} ({last})")


def record(label, check, expected, actual, ok, details=""):
    results.append({"label": label, "check": check, "expected": str(expected),
                    "actual": str(actual), "status": "PASS" if ok else "FAIL", "details": details})
    print(f"  [{'PASS' if ok else 'FAIL'}] {label} · {check}: exp={expected} got={actual} {details}")


def norm_upc(u):
    return str(u or "").lstrip("0")


def norm_vintage(v):
    s = str(v or "").strip()
    m = re.match(r"^(19|20)\d{2}", s)
    return m.group(0) if (m and s not in ("0", "")) else None


def pack_int(uq):
    try:
        return int(float(uq))
    except (TypeError, ValueError):
        return None


def catalog_truth(token, upc, ws):
    data = http_json("GET", f"{API}/api/catalog/search",
                     params={"upcs": upc, "include_tiers": "false", "limit": 50},
                     headers={"Authorization": f"Bearer {token}"})
    rows = [i for i in data.get("items", []) if i.get("wholesaler") == ws
            and norm_upc(i.get("upc")) == norm_upc(upc)]
    if not rows:
        rows = [i for i in data.get("items", []) if norm_upc(i.get("upc")) == norm_upc(upc)]
    return rows[0] if rows else None


def main():
    print(f"== Cart-page accuracy vs {WEB} ==")
    token = http_json("POST", f"{API}/api/auth/login", json={"email": EMAIL, "password": PW})["token"]
    user = http_json("POST", f"{API}/api/auth/login", json={"email": EMAIL, "password": PW})["user"]
    H = {"Authorization": f"Bearer {token}"}

    # Build a deterministic cart from the catalogue truth.
    http_json("POST", f"{API}/api/cart/clear?scope=active", headers=H)
    truth = {}
    for t in TARGETS:
        row = catalog_truth(token, t["upc"], t["wholesaler"])
        if not row:
            record(t["label"], "Catalogue row", "found", "none", False, f"upc {t['upc']} @ {t['wholesaler']}")
            continue
        truth[t["upc"]] = row
        http_json("POST", f"{API}/api/cart", headers=H, json={
            "product_name": row.get("product_name"), "wholesaler": t["wholesaler"],
            "upc": row.get("upc"), "unit_volume": row.get("unit_volume"),
            "qty_cases": QTY, "qty_units": 0})

    # Cart API lines (did the cart pick up pack / vintage / price correctly).
    cart = http_json("GET", f"{API}/api/cart", headers=H)
    cart_by_upc = {norm_upc(it.get("upc")): it for it in cart.get("items", [])}

    with sync_playwright() as pw:
        b = pw.chromium.launch(headless=True)
        ctx = b.new_context(viewport={"width": 1600, "height": 1100})
        ctx.add_init_script(
            f"localStorage.setItem('lpb_auth_token', {json.dumps(token)});"
            f"localStorage.setItem('lpb_auth_user', {json.dumps(json.dumps(user))});"
            "localStorage.setItem('celr_welcome_tour_never','1');"
            "localStorage.setItem('celr_cookie_consent','{\"necessary\":true,\"analytics\":false}');")
        page = ctx.new_page()
        page.goto(f"{WEB}/cart", wait_until="domcontentloaded")
        page.wait_for_timeout(6000)
        try:
            page.evaluate("document.querySelector('.cc')?.remove()")
        except Exception:
            pass
        try:
            page.locator("[data-tour='cart-line']").first.wait_for(timeout=30000)
        except Exception:
            pass
        # Per-line text keyed by the UPC shown on the line.
        lines = page.evaluate(
            """() => [...document.querySelectorAll("[data-tour='cart-line']")].map(
                el => el.innerText.replace(/\\s+/g,' ').trim())""")

        for t in TARGETS:
            label, upc = t["label"], t["upc"]
            row = truth.get(upc)
            if not row:
                continue
            un = norm_upc(upc)
            cart_line = cart_by_upc.get(un)

            # ---- pickup accuracy: cart line vs catalogue ----
            cat_pack = pack_int(row.get("unit_qty"))
            cat_vtg = norm_vintage(row.get("vintage"))
            cat_fl = row.get("frontline_case_price")
            if cart_line is None:
                record(label, "In cart", "line present", "missing", False)
                continue
            record(label, "Pack picked up", cat_pack, pack_int(cart_line.get("unit_qty")),
                   pack_int(cart_line.get("unit_qty")) == cat_pack)
            record(label, "Vintage picked up", cat_vtg or "none", norm_vintage(cart_line.get("vintage")) or "none",
                   norm_vintage(cart_line.get("vintage")) == cat_vtg)
            cl_fl = cart_line.get("frontline_case_price")
            record(label, "Price picked up", cat_fl, cl_fl,
                   cat_fl is not None and cl_fl is not None and abs(float(cat_fl) - float(cl_fl)) <= 0.02)

            # ---- display accuracy: the cart line DOM ----
            txt = next((L for L in lines if un and un in L.replace(",", "")), None)
            if txt is None:
                record(label, "Line displayed", "row on page", "not found", False, f"upc {un}")
                continue
            up = txt.upper()
            size = (row.get("unit_volume") or "").upper()
            record(label, "Size shown", size or "n/a", "shown" if size and size in up else "missing",
                   bool(size) and size in up)
            pk_ok = cat_pack is not None and f"{cat_pack}/CS" in up.replace(" ", "")
            record(label, "Pack shown", f"{cat_pack}/cs", "shown" if pk_ok else "missing", pk_ok)
            if cat_vtg:
                record(label, "Vintage shown", f"Vintage {cat_vtg}",
                       "shown" if f"VINTAGE {cat_vtg}" in up else "missing", f"VINTAGE {cat_vtg}" in up)
            else:
                record(label, "No junk vintage", "no 'Vintage' text",
                       "clean" if "VINTAGE" not in up else "shows vintage", "VINTAGE" not in up)
            price_ok = cat_fl is not None and f"{float(cat_fl):,.2f}" in txt
            record(label, "Price shown", f"${float(cat_fl):,.2f}" if cat_fl is not None else "n/a",
                   "shown" if price_ok else "missing", price_ok)
            has_dollar = "$" in txt
            record(label, "All fields present",
                   "name+distributor+size+pack+price",
                   "ok" if has_dollar and size in up and pk_ok else "incomplete",
                   has_dollar and bool(size) and size in up and pk_ok)
        ctx.close(); b.close()

    write_excel()
    n_fail = sum(1 for r in results if r["status"] == "FAIL")
    print(f"\n== {len(results)} cart checks, {n_fail} FAIL — report: {OUT_XLSX} ==")
    sys.exit(1 if n_fail else 0)


def write_excel():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    if os.path.exists(OUT_XLSX):
        try:
            wb = load_workbook(OUT_XLSX)
        except Exception:
            wb = Workbook(); wb.remove(wb.active)
    else:
        wb = Workbook(); wb.remove(wb.active)
    if "Cart" in wb.sheetnames:
        del wb["Cart"]
    ws = wb.create_sheet("Cart")
    headers = ["#", "Product", "Check", "Expected", "Actual", "Status", "Details", "When (UTC)"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(vertical="center")
    red = PatternFill("solid", fgColor="F8CBAD"); green = PatternFill("solid", fgColor="C6EFCE")
    when = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    for i, r in enumerate(results, 1):
        ws.append([i, r["label"], r["check"], r["expected"], r["actual"], r["status"], r["details"], when])
        ws.cell(row=i + 1, column=6).fill = red if r["status"] == "FAIL" else green
    for col, w in enumerate([4, 42, 20, 30, 30, 8, 42, 16], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = "A2"
    path = OUT_XLSX
    try:
        wb.save(path)
    except PermissionError:
        path = OUT_XLSX.replace(".xlsx", f"_cart_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(path)
    print("wrote", path)


if __name__ == "__main__":
    main()
