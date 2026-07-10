"""Discover-Deals price audit against PROD — find where a deal card's shown price
disagrees with the product-details page.

WHAT IT CHECKS
  The Discover page renders precomputed `deal_grid` cards (via
  /api/catalog/discover-deals) with NO client math — the card literally prints
  btl_1cs / rip_per_case / qd_save_per_case. The product-details page renders the
  LIVE pricing engine (via /api/catalog/product/{w}/{name}: discount_tiers +
  rip_tiers). If deal_grid is stale or out of sync with the live engine, the
  card shows a price the product page contradicts — that is the bug class we hunt.

  For up to N products drawn from the SAME category rails the page shows, we
  compare, per product:
    1) 1 CS bottle price   card.btl_1cs           vs product.frontline_unit_price
    2) RIP  (per case + $) card.rip_per_case/amt  vs product.rip_tiers[code,qty]
    3) QD   (per case)     card.qd_save_per_case   vs product.discount_tiers[qty]

  A Playwright pass first loads the LIVE /discover page and confirms the rendered
  DOM card values equal the discover-deals API values (so "shown price" == the
  API field the audit compares). That anchors the API-level audit to what a
  human actually sees on the page.

  This is a DIAGNOSTIC (understand the bugs) — it does not assert/exit non-zero.
  Every product + every mismatch is written to
  tests/discover_price_audit.xlsx (and .csv), with a categorized summary printed.

Run:  python tests/playwright/test_discover_price_audit.py [N]
Env:  CELR_WEB / CELR_API / CELR_EMAIL / CELR_PW  (default prod + owner)
"""
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from urllib.parse import quote

import requests

WEB = os.getenv("CELR_WEB", "https://nj.celr.ai").rstrip("/")
API = os.getenv("CELR_API", WEB).rstrip("/")
EMAIL = os.getenv("CELR_EMAIL", "sambit.tripathy@gmail.com")
PW = os.getenv("CELR_PW", "Cuttack10!")
N_TARGET = int(sys.argv[1]) if (len(sys.argv) > 1 and sys.argv[1].isdigit()) else 200
# FULL_COVERAGE (default): audit every deal card across all distributors with deals.
# Pass a number (e.g. 200) for a per-distributor sample instead.
FULL_COVERAGE = not (len(sys.argv) > 1 and sys.argv[1].isdigit())
DISTRIBUTORS = ["allied", "fedway", "opici"] if not FULL_COVERAGE else [
    "a_and_a", "allied", "banville", "david_bowler", "douglas_polaner", "fedway",
    "gallo", "independence_wine", "kramer", "massanois", "michael_skurnik", "monsieur",
    "opici", "other_brothers", "regal_wine", "shore_point", "trivin", "wilson_daniels",
    "wine_enterprises", "winebow",
]
_BIG = 10 ** 9
_DL = os.path.join(os.path.expanduser("~"), "Downloads")
OUT_XLSX = os.path.join(_DL if os.path.isdir(_DL) else os.path.dirname(__file__),
                        "discover_price_audit.xlsx")
OUT_CSV = os.path.join(_DL if os.path.isdir(_DL) else os.path.dirname(__file__),
                       "discover_price_audit.csv")


def http_json(method, url, **kw):
    """requests with retries on transient 5xx / network blips (prod 502 on deploy)."""
    last = None
    for attempt in range(4):
        try:
            r = requests.request(method, url, timeout=40, **kw)
            if r.status_code == 404:
                return {"__status__": 404}
            if r.status_code >= 500:
                last = f"{r.status_code} {r.reason}"
                time.sleep(1.5 + attempt)
                continue
            r.raise_for_status()
            return r.json()
        except Exception as e:  # noqa: BLE001
            last = f"{type(e).__name__}: {e}"
            time.sleep(1.5 + attempt)
    return {"__error__": last}


def login():
    b = http_json("POST", f"{API}/api/auth/login", json={"email": EMAIL, "password": PW})
    return b.get("token"), b.get("user")


# ---- numeric helpers -------------------------------------------------------
def num(x):
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def close(a, b, rel=0.002, absolute=0.02):
    """True if a≈b within max(absolute, rel*value). None==None ok; one-None fails."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= max(absolute, rel * max(abs(a), abs(b)))


def unit_family(u):
    u = str(u or "").strip().lower()
    return "bottle" if u.startswith("b") else "case"


def norm_upc(u):
    return re.sub(r"\D", "", str(u or "")).lstrip("0")


def split_codes(code):
    if not code:
        return []
    return [c for c in re.split(r"[\s,;/]+", str(code).strip()) if c and c.lower() not in ("none", "nan", "0")]


_STOP = {"the", "and", "of", "with", "pk", "ml", "l", "bottle", "btl", "case", "cs", "1pk", "6pk", "12pk"}


def _words(s):
    return {w for w in re.findall(r"[a-z0-9]+", str(s or "").lower()) if w not in _STOP and len(w) > 1}


def _name_jaccard(a, b):
    """Loose brand/name overlap so size/format wording differences don't false-flag,
    but a genuinely different product (barcode reuse) does."""
    wa, wb = _words(a), _words(b)
    if not wa or not wb:
        return 1.0
    return len(wa & wb) / len(wa | wb)


# ---------------------------------------------------------------------------
# 1) Enumerate deal cards from the SAME rails the Discover page renders.
# ---------------------------------------------------------------------------
def collect_cards(edition_hint=None, divisions=None, target=None):
    target = target or N_TARGET
    cats = http_json("GET", f"{API}/api/catalog/top-categories")
    rails = [*(cats.get("spirits") or []), *(cats.get("wine") or [])] if isinstance(cats, dict) else []
    seen = {}
    order = []

    def add(items):
        for d in items or []:
            key = (d.get("primary_wholesaler"), norm_upc(d.get("upc")),
                   d.get("unit_volume"), str(d.get("unit_qty")), str(d.get("vintage")))
            if key in seen:
                continue
            seen[key] = d
            order.append(key)

    base = {}
    if edition_hint:
        base["edition"] = edition_hint
    if divisions:
        base["divisions"] = divisions

    SIZES = ["375ML", "750ML", "1L", "1.75L", "355ML", "50ML", "700ML", "720ML", "19.2OZ", "3L"]

    def pull(params):
        """One page; if it hits the 200 cap, sub-split by size to page past it."""
        r = http_json("GET", f"{API}/api/catalog/discover-deals{_qs(params)}")
        items = r.get("items") or []
        add(items)
        if len(items) >= 200:  # capped — recover the rest per size
            for sz in SIZES:
                rr = http_json("GET", f"{API}/api/catalog/discover-deals{_qs({**params, 'sizes': sz})}")
                add(rr.get("items"))

    # 1) Flat pull (all categories) across several sorts — different sorts surface
    #    different top-200 slices, so a >200 bucket is largely recovered.
    for srt in ("net", "pct", "name", "case", "rip", "qd"):
        if len(order) >= target:
            break
        pull({"sort": srt, "limit": 200, **base})

    # 2) Every category rail (spirit_category / grapes / product_type), like the page.
    for rail in rails:
        if len(order) >= target:
            break
        pr = rail.get("params") or {}
        params = {"sort": "name", "limit": 200, **base}
        for k in ("spirit_category", "grapes", "product_type"):
            if pr.get(k):
                params[k] = pr[k]
        if "spirit_category" not in params and "grapes" not in params and "product_type" not in params and pr.get("q"):
            s = http_json("GET", f"{API}/api/catalog/search?q={quote(pr['q'])}&limit=120")
            upcs = ",".join({str(i.get("upc")) for i in (s.get("items") or []) if i.get("upc")})
            if not upcs:
                continue
            params["upcs"] = upcs
        pull(params)

    return [seen[k] for k in order[:target]]


def _qs(params):
    if not params:
        return ""
    return "?" + "&".join(f"{k}={quote(str(v))}" for k, v in params.items())


# ---------------------------------------------------------------------------
# 2) Compare one card vs its product-details page (the live pricing engine).
# ---------------------------------------------------------------------------
def audit_card(d):
    # Report label uses the friendly display name; the URL + identity check use the
    # RAW product_name — exactly what the card deep-link (cardHref n=product_name)
    # sends, so the audit resolves the same row a real click would.
    name = d.get("display_name") or d.get("product_name") or ""
    card_raw = d.get("product_name") or ""
    ident = {
        "wholesaler": d.get("primary_wholesaler"),
        "product": name,
        "upc": d.get("upc"),
        "edition": d.get("edition"),
        "size": d.get("unit_volume"),
        "pack": d.get("unit_qty"),
        "vintage": d.get("vintage"),
        "rip_code": d.get("rip_code"),
    }
    q = {}
    if d.get("edition"):
        q["edition"] = d["edition"]
    if d.get("upc"):
        q["upc"] = d["upc"]
    if d.get("unit_volume"):
        q["unit_volume"] = d["unit_volume"]
    if d.get("unit_qty") not in (None, ""):
        q["unit_qty"] = d["unit_qty"]
    if d.get("vintage") not in (None, ""):
        q["vintage"] = d["vintage"]
    # NOTE: the card deep-link (cardHref) does NOT send rip_code, so the product
    # page resolves RIP from the row's own code. Passing a malformed multi-token
    # code (e.g. "VKDA 6100") as an override makes _split_rip_codes mis-split it
    # and drops all tiers — a test-only false positive. Mirror cardHref: omit it.
    url = f"{API}/api/catalog/product/{quote(str(d.get('primary_wholesaler') or ''))}/{quote(card_raw or name)}{_qs(q)}"
    resp = http_json("GET", url)

    findings = []  # dict rows

    def flag(check, severity, shown, page, detail):
        findings.append({**ident, "check": check, "severity": severity,
                         "shown_on_card": shown, "product_page": page, "detail": detail})

    if not isinstance(resp, dict) or resp.get("__status__") == 404 or resp.get("__error__") or not resp.get("product"):
        why = "404 — product page does not resolve" if resp.get("__status__") == 404 else (
            resp.get("__error__") or "no product in response")
        flag("FETCH", "HIGH", "(card visible)", "(page missing)", why)
        return findings, True  # counted as fetch-fail

    p = resp.get("product") or {}
    dts = resp.get("discount_tiers") or []
    rts = resp.get("rip_tiers") or []

    # --- 0) DID THE RIGHT PAGE OPEN? (identity of the resolved row) ---
    # Compare RAW name vs RAW name: the card's display_name is Go-UPC-enriched
    # ("The Balvenie 21 Year…") while the page returns the distributor's raw name
    # ("BALVENIE 21YR 3P") — comparing those two always false-flags. The card's own
    # raw product_name IS the page's raw product_name for the same row.
    page_name = p.get("product_name") or ""
    if norm_upc(p.get("upc")) and norm_upc(d.get("upc")) and norm_upc(p.get("upc")) != norm_upc(d.get("upc")):
        flag("OPENS_WRONG_UPC", "HIGH", d.get("upc"), p.get("upc"),
             "product page resolved a different UPC than the card")
    elif card_raw and page_name and _name_jaccard(card_raw, page_name) < 0.34:
        flag("OPENS_WRONG_PRODUCT", "HIGH", card_raw, page_name,
             "card links to a product page showing a different-named product (barcode reuse?)")
    ps = str(p.get("unit_volume") or "")
    if d.get("unit_volume") and ps and ps != str(d.get("unit_volume")):
        flag("OPENS_WRONG_SIZE", "MED", d.get("unit_volume"), ps,
             "product page opened a different size than the card")

    # --- 1) 1-CS bottle price ---
    shown_1cs = num(d.get("btl_1cs"))
    page_1cs = num(p.get("frontline_unit_price"))
    if shown_1cs is not None and not close(shown_1cs, page_1cs):
        flag("1CS_BOTTLE", "HIGH", shown_1cs, page_1cs,
             f"card 1-case bottle {shown_1cs} vs product page {page_1cs}")

    # --- 2) RIP (per-case rebate + total $) ---
    if d.get("has_rip"):
        r_pc = num(d.get("rip_per_case"))
        r_amt = num(d.get("rip_amount"))
        r_qty = num(d.get("rip_qty"))
        want_codes = set(c.lower() for c in split_codes(d.get("rip_code")))
        # candidate tiers on the product page (respect the code the card advertises)
        cands = rts
        if want_codes:
            coded = [t for t in rts if str(t.get("code") or "").lower() in want_codes]
            if coded:
                cands = coded
        # exact match: a tier at the same qty whose per-case rebate agrees
        by_qty = [t for t in cands if r_qty is not None and close(num(t.get("qty")), r_qty, absolute=0.5)]
        pc_match = [t for t in (by_qty or cands) if close(num(t.get("per_case_savings")), r_pc)]
        if not rts:
            flag("RIP", "HIGH", f"{r_pc}/cs (qty {r_qty})", "(no RIP tiers)",
                 "card advertises a RIP the product page has none of")
        elif not pc_match:
            best = max(cands or rts, key=lambda t: num(t.get("qty")) or 0, default=None)
            flag("RIP", "HIGH", f"{r_pc}/cs (qty {r_qty}, code {d.get('rip_code')})",
                 (f"{num(best.get('per_case_savings'))}/cs (qty {num(best.get('qty'))}, "
                  f"code {best.get('code')})" if best else "(no matching tier)"),
                 "card RIP per-case not found on product page at that qty/code")
        else:
            # per-case agrees; cross-check the total $ (amount) too
            t = pc_match[0]
            if r_amt is not None and not close(r_amt, num(t.get("amount")), absolute=1.0):
                flag("RIP_AMT", "MED", r_amt, num(t.get("amount")),
                     f"RIP total $ differs (card {r_amt} vs page {num(t.get('amount'))})")

    # --- 3) QD (per-case saving at the featured bulk quantity) ---
    if d.get("has_qd"):
        q_pc = num(d.get("qd_save_per_case"))
        q_qty = num(d.get("qd_qty"))
        # product QD tiers: quantity + amount_per_case (per case). Exclude the
        # 1-case entry QD (the card excludes it from the featured chip).
        bulk = [t for t in dts if not (num(t.get("quantity")) == 1)]
        at_qty = [t for t in dts if q_qty is not None and close(num(t.get("quantity")), q_qty, absolute=0.5)]
        pc_match = [t for t in (at_qty or bulk or dts) if close(num(t.get("amount_per_case")), q_pc)]
        if not dts:
            flag("QD", "HIGH", f"{q_pc}/cs (qty {q_qty})", "(no QD tiers)",
                 "card advertises a QD the product page has none of")
        elif not pc_match:
            best = max(bulk or dts, key=lambda t: num(t.get("quantity")) or 0, default=None)
            flag("QD", "HIGH", f"{q_pc}/cs (qty {q_qty})",
                 (f"{num(best.get('amount_per_case'))}/cs (qty {num(best.get('quantity'))})"
                  if best else "(no matching tier)"),
                 "card QD per-case not found on product page at that qty")

    return findings, False


# ---------------------------------------------------------------------------
# 3) Playwright DOM anchor — prove the rendered card == the deal-grid API value.
# ---------------------------------------------------------------------------
def playwright_dom_check(token, user):
    out = {"scraped": 0, "matched": 0, "mismatches": [], "error": None}
    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:  # noqa: BLE001
        out["error"] = f"playwright not installed: {e}"
        return out
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(viewport={"width": 1600, "height": 1200})
            ctx.add_init_script(
                f"localStorage.setItem('lpb_auth_token', {json.dumps(token)});"
                f"localStorage.setItem('lpb_auth_user', {json.dumps(json.dumps(user))});"
                "localStorage.setItem('celr_welcome_tour_never', '1');"
            )
            page = ctx.new_page()
            page.goto(f"{WEB}/discover", wait_until="networkidle", timeout=60000)
            try:
                page.wait_for_selector("a.disc-card", timeout=30000)
            except Exception:
                out["error"] = "no deal cards rendered on /discover"
                browser.close()
                return out
            page.wait_for_timeout(2500)
            cards = page.query_selector_all("a.disc-card")
            # (a) DOM shows the deal_grid value verbatim?  (b) does CLICKING a card
            # open the correct product page (matching name)?
            targets = []  # (card_name, href)
            for c in cards[:40]:
                href = c.get_attribute("href") or ""
                bp = c.query_selector(".disc-fav-prices")
                nm_el = c.query_selector(".disc-card-name")
                card_name = (nm_el.inner_text() if nm_el else "").split("\n")[0].strip()
                txt = bp.inner_text() if bp else ""
                m = re.search(r"\$([\d,]+\.\d{2})", txt)
                if not m or "u=" not in href:
                    continue
                out["scraped"] += 1
                dom_1cs = float(m.group(1).replace(",", ""))
                upc = re.search(r"[?&]u=([^&]+)", href)
                size = re.search(r"[?&]s=([^&]+)", href)
                qp = {"upcs": requests.utils.unquote(upc.group(1)), "limit": 5}
                api = http_json("GET", f"{API}/api/catalog/discover-deals{_qs(qp)}")
                api_1cs = None
                for it in api.get("items", []):
                    if size and it.get("unit_volume") != requests.utils.unquote(size.group(1)):
                        continue
                    api_1cs = num(it.get("btl_1cs"))
                    break
                if api_1cs is not None and abs(api_1cs - dom_1cs) <= 0.02:
                    out["matched"] += 1
                elif api_1cs is not None:
                    out["mismatches"].append({"upc": upc.group(1), "dom": dom_1cs, "api": api_1cs})
                if card_name and len(targets) < 15:
                    targets.append((card_name, href))

            # Click-through: navigate to each card's product page, read the title.
            out["click"] = {"tested": 0, "ok": 0, "wrong": [], "failed": []}
            for card_name, href in targets:
                out["click"]["tested"] += 1
                try:
                    page.goto(f"{WEB}{href}", wait_until="domcontentloaded", timeout=40000)
                    try:
                        page.wait_for_selector("h1.pd-title", timeout=20000)
                    except Exception:
                        pass
                    page.wait_for_timeout(600)
                    h = page.query_selector("h1.pd-title")
                    # strip the "(1.75L · 6 btl/cs)" size parenthetical the headline appends
                    raw = (h.inner_text() if h else "")
                    page_name = re.sub(r"\s*\(.*$", "", raw.split("\n")[0]).strip()
                    if page_name and _name_jaccard(card_name, page_name) >= 0.34:
                        out["click"]["ok"] += 1
                    elif page_name:
                        out["click"]["wrong"].append({"card": card_name, "page": page_name})
                    else:
                        out["click"]["failed"].append(card_name)
                except Exception as e:  # noqa: BLE001
                    out["click"]["failed"].append(f"{card_name} ({type(e).__name__})")
            browser.close()
    except Exception as e:  # noqa: BLE001
        out["error"] = f"{type(e).__name__}: {e}"
    return out


# ---------------------------------------------------------------------------
def write_reports(all_findings, per_dist_stats, click_summary):
    import csv
    cols = ["distributor", "wholesaler", "product", "upc", "edition", "size", "pack",
            "vintage", "rip_code", "check", "severity", "shown_on_card", "product_page", "detail"]
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in all_findings:
            w.writerow({k: r.get(k, "") for k in cols})
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        # ---- Summary sheet ----
        ws = wb.active
        ws.title = "summary"
        ws.append([f"Discover-Deals price audit vs product-details page — {WEB}"])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([f"generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}  ·  target {N_TARGET}/distributor"])
        ws.append([])
        ws.append(["Distributor", "Products audited", "Products w/ issue", "Findings",
                   "HIGH", "1CS", "RIP", "QD", "OPENS_WRONG*", "Page-fetch fails"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for dist, s in per_dist_stats.items():
            ws.append([dist, s["n"], s["prod_issues"], s["findings"], s["high"],
                       s.get("1CS", 0), s.get("RIP", 0), s.get("QD", 0),
                       s.get("OPENS", 0), s["fetch_fails"]])
        ws.append([])
        ws.append(["Click-through test (did the correct product page open?)"])
        ws[ws.max_row][0].font = Font(bold=True)
        if click_summary:
            ws.append([f"  tested {click_summary.get('tested', 0)} card clicks · "
                       f"{click_summary.get('ok', 0)} opened the right product · "
                       f"{len(click_summary.get('wrong', []))} opened a DIFFERENT product · "
                       f"{len(click_summary.get('failed', []))} failed to load"])
            for mm in click_summary.get("wrong", [])[:20]:
                ws.append([f"     card '{mm['card']}'  →  page '{mm['page']}'"])
        else:
            ws.append(["  (Playwright click-through not run)"])

        # ---- Mismatches sheet ----
        ms = wb.create_sheet("mismatches")
        ms.append(cols)
        for cell in ms[1]:
            cell.font = Font(bold=True)
        ms.freeze_panes = "A2"
        red = PatternFill("solid", fgColor="FFC7CE")
        yel = PatternFill("solid", fgColor="FFEB9C")
        for r in all_findings:
            ms.append([str(r.get(k, "")) for k in cols])
            fill = red if r.get("severity") == "HIGH" else yel
            for cell in ms[ms.max_row]:
                cell.fill = fill
        widths = {"C": 40, "M": 40, "N": 46, "L": 22}
        for col, wd in widths.items():
            ms.column_dimensions[col].width = wd
        wb.save(OUT_XLSX)
        return OUT_XLSX
    except Exception as e:  # noqa: BLE001
        print(f"  [xlsx] openpyxl unavailable ({e}); wrote CSV only")
        return OUT_CSV


def _bucket(check):
    if check.startswith("OPENS"):
        return "OPENS"
    if check.startswith("1CS"):
        return "1CS"
    if check.startswith("RIP"):
        return "RIP"
    if check.startswith("QD"):
        return "QD"
    return check


def main():
    t0 = time.time()
    scope = f"FULL COVERAGE · {len(DISTRIBUTORS)} distributors" if FULL_COVERAGE else \
        f"target {N_TARGET}/distributor: {', '.join(DISTRIBUTORS)}"
    print(f"[{datetime.now(timezone.utc):%H:%M:%S}] Discover price audit → {WEB}  ({scope})")
    token, user = login()
    if not token:
        print("  login failed — DOM/click pass skipped; API audit continues")

    # Playwright DOM + click-through anchor (once; best-effort).
    click_summary = None
    if token:
        print("[dom] loading LIVE /discover, verifying shown==API and clicking cards…")
        dom = playwright_dom_check(token, user)
        if dom.get("error"):
            print(f"[dom] skipped/failed: {dom['error']}")
        else:
            print(f"[dom] scraped {dom['scraped']} cards, {dom['matched']} match API 1-CS price, "
                  f"{len(dom['mismatches'])} DOM≠API")
            click_summary = dom.get("click")
            if click_summary:
                print(f"[click] {click_summary['tested']} clicks · {click_summary['ok']} right product · "
                      f"{len(click_summary['wrong'])} WRONG product · {len(click_summary['failed'])} load-fail")
                for mm in click_summary["wrong"][:10]:
                    print(f"        WRONG: card '{mm['card']}' → page '{mm['page']}'")

    all_findings = []
    per_dist_stats = {}
    lock = threading.Lock()
    prog = {"done": 0, "total": 0}
    stop = threading.Event()

    def heartbeat():
        while not stop.wait(60):
            with lock:
                el = int(time.time() - t0)
                hi = sum(1 for f in all_findings if f["severity"] == "HIGH")
                print(f"  … [{el//60}m{el%60:02d}s] {prog['done']}/{prog['total']} audited · "
                      f"{len(all_findings)} findings ({hi} HIGH)")
    hb = threading.Thread(target=heartbeat, daemon=True)
    hb.start()

    tgt = _BIG if FULL_COVERAGE else N_TARGET
    seen_global = set()  # a card can belong to several distributors' groups; audit once
    for dist in DISTRIBUTORS:
        lim = "ALL" if FULL_COVERAGE else N_TARGET
        print(f"\n[collect:{dist}] pulling up to {lim} deal cards…")
        cards = collect_cards(divisions=dist, target=tgt)
        # dedup across distributor groups by SKU identity
        uniq = []
        for c in cards:
            k = (c.get("primary_wholesaler"), norm_upc(c.get("upc")),
                 c.get("unit_volume"), str(c.get("unit_qty")), str(c.get("vintage")))
            if k in seen_global:
                continue
            seen_global.add(k)
            uniq.append(c)
        cards = uniq
        print(f"[collect:{dist}] {len(cards)} unique cards "
              f"(editions {sorted(set(c.get('edition') for c in cards))})")
        with lock:
            prog["total"] += len(cards)
        dist_findings, ffails = [], 0
        with ThreadPoolExecutor(max_workers=8) as ex:
            futs = {ex.submit(audit_card, c): c for c in cards}
            for fut in as_completed(futs):
                fnd, ff = fut.result()
                for f in fnd:
                    f["distributor"] = dist
                with lock:
                    dist_findings.extend(fnd)
                    all_findings.extend(fnd)
                    ffails += 1 if ff else 0
                    prog["done"] += 1
        # per-distributor rollup
        stat = {"n": len(cards), "findings": len(dist_findings), "fetch_fails": ffails,
                "high": sum(1 for f in dist_findings if f["severity"] == "HIGH"),
                "prod_issues": len({(f["upc"], f["size"]) for f in dist_findings})}
        for f in dist_findings:
            b = _bucket(f["check"])
            stat[b] = stat.get(b, 0) + 1
        per_dist_stats[dist] = stat
        print(f"[done:{dist}] {len(dist_findings)} findings "
              f"({stat['high']} HIGH) across {stat['prod_issues']} products")
    stop.set()

    # Fold the Playwright click-through "wrong product opened" cases in as rows so
    # they land in the Excel alongside the price mismatches.
    if click_summary:
        for mm in click_summary.get("wrong", []):
            all_findings.append({
                "distributor": "dom-sample", "wholesaler": "", "product": mm["card"],
                "upc": "", "edition": "", "size": "", "pack": "", "vintage": "",
                "rip_code": "", "check": "OPENS_WRONG_CLICK", "severity": "HIGH",
                "shown_on_card": mm["card"], "product_page": mm["page"],
                "detail": "clicking the card opened a different-named product page (barcode reuse / vintage / enrichment mislabel)",
            })

    el = int(time.time() - t0)
    path = write_reports(all_findings, per_dist_stats, click_summary)

    by_check = {}
    for f in all_findings:
        by_check[f["check"]] = by_check.get(f["check"], 0) + 1

    print("\n" + "=" * 80)
    print(f"AUDIT COMPLETE in {el//60}m{el%60:02d}s — {sum(s['n'] for s in per_dist_stats.values())} products")
    print(f"{'distributor':12s} {'audited':>8s} {'issues':>7s} {'HIGH':>5s} "
          f"{'1CS':>4s} {'RIP':>4s} {'QD':>4s} {'OPENS':>6s} {'404':>4s}")
    for dist, s in per_dist_stats.items():
        print(f"{dist:12s} {s['n']:>8d} {s['prod_issues']:>7d} {s['high']:>5d} "
              f"{s.get('1CS', 0):>4d} {s.get('RIP', 0):>4d} {s.get('QD', 0):>4d} "
              f"{s.get('OPENS', 0):>6d} {s['fetch_fails']:>4d}")
    print("\nFindings by exact check:")
    for k in sorted(by_check, key=lambda k: -by_check[k]):
        print(f"   {k:22s} {by_check[k]}")
    if click_summary:
        print(f"\nClick-through: {click_summary['ok']}/{click_summary['tested']} cards opened the "
              f"correct product page; {len(click_summary['wrong'])} opened a different product.")
    print(f"\nReport (Excel) written: {path}")
    print("=" * 80)


if __name__ == "__main__":
    main()
